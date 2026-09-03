"""Daily calendar email - two Outlook drafts, straight from Bloomberg, no Excel.

    Draft 1  Macro calendar    central-bank decisions across the US and Asia, then
                               the high-relevance US data calendar, next 3 months
    Draft 2  HSI earnings      constituents reporting in the next 91 days, with
                               the conference-call date and time

Built for Jupyter.  Load it in one cell:

    %run daily_email.py

then call one of these in the next cell:

    demo()                          both emails from sample data, no Bloomberg
    test()                          run the self-tests
    probe()                         check fields and tickers on the terminal
    run()                           the real thing: two Outlook drafts
    run(excel='C:/path/cal.xlsx')   from the workbook instead
    compare('C:/path/cal.xlsx')     Bloomberg against the workbook

The emails are drawn inside the notebook and saved to the out folder.  Nothing is
ever sent.  None of the six raises: a problem is printed as a sentence saying what
to do next, and the kernel keeps running.

Everything is in this one file:
    CONFIG      recipients, subjects, the exclusion list, tickers   <- edit here
    ENGINE      the Bloomberg session, the three pulls, the HTML, Outlook
    FIXTURE     a fake terminal, so demo() and test() run on any machine
    TESTS       ~90 checks over parsing, filtering, sorting and rendering
    JUPYTER     the six functions above
    MAIN        the same thing from a command prompt, if you ever want it
"""

import argparse
import datetime as dt
import html as _html
import os
import re
import sys
from email.message import EmailMessage
from email.utils import format_datetime

try:                                    # normal: next to the script
    HERE = os.path.dirname(os.path.abspath(__file__))
except NameError:                       # pasted straight into a notebook cell
    HERE = os.getcwd()


def _in_jupyter():
    """True inside a notebook.  Jupyter needs different manners: argparse must
    never see the kernel's own argv, and sys.exit() would kill the kernel."""
    try:
        from IPython import get_ipython
        return type(get_ipython()).__name__ == 'ZMQInteractiveShell'
    except Exception:
        return False


IN_JUPYTER = _in_jupyter()

# ================================================================ CONFIG ===
RECIPIENTS = 'oscar.chan1@nomura.com; gordon.ho1@nomura.com'
GREETING = 'Hi Oscar, Gordon,'

BBG_HOST = 'localhost'
BBG_PORT = 8194
EVENT_SPINS = 240              # bounded drain per request: ~2 min worst case at 500 ms
CHUNK = 50                     # securities per ReferenceDataRequest

LOOKAHEAD_DAYS = 91            # Excel: range(0D, 3M) and btoday()+91d
OUT_DIR = os.path.join(HERE, 'out')   # an .html and an .eml copy of every draft
FOOTER_NOTES = True            # list what the API could NOT return in the email footer
RELEVANCE_MIN = 0              # drop US events whose RELEVANCE_VALUE is below this (0 = off)

# Times.  Bloomberg hands back release times in the terminal's own time-zone
# setting (TZDF), so by default they are shown exactly as returned.  If --probe
# shows New York times for US data, set TIME_IN_LOCAL_TZ = True and every row is
# converted from the release's local zone to TIME_TZ_OUT (the date moves too).
TIME_IN_LOCAL_TZ = False
TIME_TZ_OUT = 'Asia/Hong_Kong'
US_TZ = 'America/New_York'

HSI_INDEX = 'HSI Index'

# The two drafts.  {date} = run date, {n_hsi} = HSI names reporting.
EMAILS = [
    dict(key='macro',
         subject='Macro calendar - {date}: central banks + US data (next 3M)',
         intro=('Central-bank decisions across the US and Asia, then the high-relevance US '
                'data calendar, both for the next 3 months. Survey = Bloomberg consensus '
                'median. Prior = last print (Current = the policy rate today).'),
         sections=['central_banks', 'us_eco']),
    dict(key='earnings',
         subject='HSI earnings calendar - {date}: {n_hsi} constituents reporting in the next 91 days',
         intro=('HSI constituents with an expected report date inside the next 91 days, '
                'earliest first, with the conference-call date and time where Bloomberg '
                'has one.'),
         sections=['hsi_earnings']),
]

# Field mnemonics, in order of preference.  Every candidate is requested; the
# first one Bloomberg actually returns is used, and --probe reports which.
ECO_FIELDS = {
    'date':      ['ECO_FUTURE_RELEASE_DATE', 'ECO_FUTURE_RELEASE_DATE_TIME', 'ECO_RELEASE_DT'],
    'time':      ['ECO_FUTURE_RELEASE_TIME', 'ECO_RELEASE_TIME'],
    'survey':    ['BN_SURVEY_MEDIAN'],
    'prior':     ['PX_LAST'],
    'relevance': ['RELEVANCE_VALUE'],
    'bbg_name':  ['NAME'],
}
# 'date' falls back to the ERN_ANN_DT_AND_PER bulk field, which v33_hk_basket_full.py
# already pulls on this desk - the one earnings mnemonic here that is proven, not assumed.
EARN_FIELDS = {
    'name':      ['NAME'],
    'date':      ['EXPECTED_REPORT_DT'],
    'time':      ['EXPECTED_REPORT_TIME'],
    'call_date': ['EARNINGS_CONF_CALL_DT', 'CONF_CALL_DT'],
    'call_time': ['EARNINGS_CONF_CALL_TIME', 'CONF_CALL_TIME'],
}
EARN_BULK_DATE = 'ERN_ANN_DT_AND_PER'

# US calendar universe = what BQL's relevancy=HIGH returns, as tickers.  The
# label is the ECO event name (what the email shows and what EXCLUDE_EVENTS
# matches).  Add a row to widen, delete a row to narrow; --probe prints the
# Bloomberg NAME beside each so a mislabelled row is obvious.
US_ECO_TICKERS = [
    ('NFP TCH Index',   'Change in Nonfarm Payrolls'),
    ('NFP PCH Index',   'Change in Private Payrolls'),
    ('USURTOT Index',   'Unemployment Rate'),
    ('AHE MOM% Index',  'Average Hourly Earnings MoM'),
    ('AHE YOY% Index',  'Average Hourly Earnings YoY'),
    ('PRUSTOT Index',   'Labor Force Participation Rate'),
    ('ADP CHNG Index',  'ADP Employment Change'),
    ('INJCJC Index',    'Initial Jobless Claims'),
    ('INJCSP Index',    'Continuing Claims'),
    ('JOLTTOTL Index',  'JOLTS Job Openings'),
    ('CPI CHNG Index',  'CPI MoM'),
    ('CPI YOY Index',   'CPI YoY'),
    ('CPUPXCHG Index',  'CPI Ex Food and Energy MoM'),
    ('CPI XYOY Index',  'CPI Ex Food and Energy YoY'),
    ('FDIDFDMO Index',  'PPI Final Demand MoM'),
    ('FDIDFDYO Index',  'PPI Final Demand YoY'),
    ('FDIDSGMO Index',  'PPI Ex Food and Energy MoM'),
    ('FDIDSGYO Index',  'PPI Ex Food and Energy YoY'),
    ('PCE DEFM Index',  'PCE Price Index MoM'),
    ('PCE DEFY Index',  'PCE Price Index YoY'),
    ('PCE CMOM Index',  'Core PCE Price Index MoM'),
    ('PCE CYOY Index',  'Core PCE Price Index YoY'),
    ('PITLCHNG Index',  'Personal Income'),
    ('PCE CRCH Index',  'Personal Spending'),
    ('RSTAMOM Index',   'Retail Sales Advance MoM'),
    ('RSTAXMOM Index',  'Retail Sales Ex Auto MoM'),
    ('RSTAXAG Index',   'Retail Sales Ex Auto and Gas'),
    ('RSTACGRP Index',  'Retail Sales Control Group'),
    ('GDP CQOQ Index',  'GDP Annualized QoQ'),
    ('GDPCTOT% Index',  'Personal Consumption'),
    ('GDP PIQQ Index',  'GDP Price Index'),
    ('GDPCPCEC Index',  'Core PCE Price Index QoQ'),
    ('NAPMPMI Index',   'ISM Manufacturing'),
    ('NAPMPRIC Index',  'ISM Prices Paid'),
    ('NAPMNMI Index',   'ISM Services Index'),
    ('MPMIUSMA Index',  'S&P Global US Manufacturing PMI'),
    ('MPMIUSSA Index',  'S&P Global US Services PMI'),
    ('MPMIUSCA Index',  'S&P Global US Composite PMI'),
    ('FDTR Index',      'FOMC Rate Decision (Upper Bound)'),
    ('IP CHNG Index',   'Industrial Production MoM'),
    ('CPTICHNG Index',  'Capacity Utilization'),
    ('DGNOCHNG Index',  'Durable Goods Orders'),
    ('DGNOXTCH Index',  'Durables Ex Transportation'),
    ('CGNOXAI% Index',  'Cap Goods Orders Nondef Ex Air'),
    ('TMNOCHNG Index',  'Factory Orders'),
    ('NHSPSTOT Index',  'Housing Starts'),
    ('NHSPATOT Index',  'Building Permits'),
    ('NHSLTOT Index',   'New Home Sales'),
    ('ETSLTOTL Index',  'Existing Home Sales'),
    ('USPHTMOM Index',  'Pending Home Sales MoM'),
    ('HPIMMOM Index',   'FHFA House Price Index MoM'),
    ('SPCS20Y% Index',  'S&P CoreLogic CS 20-City YoY NSA'),
    ('CONSSENT Index',  'U. of Mich. Sentiment'),
    ('CONSP1MD Index',  'U. of Mich. 1 Yr Inflation'),
    ('CONSPXMD Index',  'U. of Mich. 5-10 Yr Inflation'),
    ('CONCCONF Index',  'Conf. Board Consumer Confidence'),
    ('USTBTOT Index',   'Trade Balance'),
    ('MBAVCHNG Index',  'MBA Mortgage Applications'),
    ('EMPRGBCI Index',  'Empire Manufacturing'),
    ('OUTFGAF Index',   'Philadelphia Fed Business Outlook'),
    ('RCHSINDX Index',  'Richmond Fed Manufact. Index'),
    ('DFEDGBA Index',   'Dallas Fed Manf. Activity'),
    ('CHPMINDX Index',  'MNI Chicago PMI'),
    ('CFNAI Index',     'Chicago Fed Nat Activity Index'),
    ('LEI CHNG Index',  'Leading Index'),
    ('MWINCHNG Index',  'Wholesale Inventories MoM'),
    ('IMP1CHNG Index',  'Import Price Index MoM'),
    ('FDDSSD Index',    'Federal Budget Balance'),
    ('SBOITOTL Index',  'NFIB Small Business Optimism'),
    ('TICTTOT Index',   'Total Net TIC Flows'),
    ('TICTLONG Index',  'Net Long-term TIC Flows'),
    ('ECI SA% Index',   'Employment Cost Index'),
    ('PRODNFR% Index',  'Nonfarm Productivity'),
    ('COSTNFR% Index',  'Unit Labor Costs'),
    ('CNSTTMOM Index',  'Construction Spending MoM'),
    ('USCABAL Index',   'Current Account Balance'),
]

# Your VBA excludeEvents dictionary, verbatim.  Matched against the event
# label (case and spacing ignored) in every source, API or Excel.
EXCLUDE_EVENTS = [
    'Building Permits',
    'MBA Mortgage Applications',
    'Continuing Claims',
    'Durable Goods Orders',
    'Durables Ex Transportation',
    'Chicago Fed Nat Activity Index',
    'Wholesale Inventories MoM',
    'U. of Mich. Sentiment',
    'Dallas Fed Manf. Activity',
    'FHFA House Price Index MoM',
    'MNI Chicago PMI',
    'Conf. Board Consumer Confidence',
    'JOLTS Job Openings',
    'ISM Prices Paid',
    'Factory Orders',
    'S&P Global US Services PMI',
    'S&P Global US Composite PMI',
    'Federal Budget Balance',
    'NFIB Small Business Optimism',
    'Total Net TIC Flows',
    'Net Long-term TIC Flows',
    'Empire Manufacturing',
    'Philadelphia Fed Business Outlook',
    'Retail Sales Ex Auto MoM',
    'Pending Home Sales MoM',
    'Import Price Index MoM',
    'Capacity Utilization',
    'Leading Index',
    'Richmond Fed Manufact. Index',
    'S&P Global US Manufacturing PMI',
]

# Central banks: (country, bank, policy-rate ticker, event label, local time zone).
# The ticker's next ECO release date is the next decision.  Rows marked VERIFY
# are the ones I could not confirm off-terminal; --probe checks each and searches
# for the right symbol when one is rejected.
CENTRAL_BANKS = [
    ('US', 'Fed',  'FDTR Index',     'FOMC Rate Decision (Upper Bound)', 'America/New_York'),
    ('TW', 'CBC',  'TARDR Index',    'CBC Benchmark Interest Rate',      'Asia/Taipei'),      # VERIFY
    ('KR', 'BoK',  'KORP7D Index',   'BoK Base Rate',                    'Asia/Seoul'),
    ('MY', 'BNM',  'MAOPRATE Index', 'BNM Overnight Policy Rate',        'Asia/Kuala_Lumpur'),
    ('TH', 'BoT',  'BTRR1DAY Index', 'BoT Benchmark Interest Rate',      'Asia/Bangkok'),
    ('ID', 'BI',   'IDBIRRPO Index', 'BI-Rate',                          'Asia/Jakarta'),     # VERIFY
    ('IN', 'RBI',  'INRPYLDP Index', 'RBI Repurchase Rate',              'Asia/Kolkata'),
    ('CN', 'PBoC', 'CHLR1Y Index',   '1-Year Loan Prime Rate',           'Asia/Shanghai'),    # VERIFY
    ('CN', 'PBoC', 'CHLR5Y Index',   '5-Year Loan Prime Rate',           'Asia/Shanghai'),    # VERIFY
    ('JP', 'BoJ',  'BOJDTR Index',   'BoJ Target Rate',                  'Asia/Tokyo'),
    ('AU', 'RBA',  'RBATCTR Index',  'RBA Cash Rate Target',             'Australia/Sydney'),
]

# Excel source: None = auto-detect the three BQL blocks by their header row.
# Or pin them: [('Sheet1', 'A1', 'us_eco'), ('Sheet1', 'N1', 'central_banks'),
#               ('Sheet2', 'A1', 'hsi_earnings')]   (cell = first header cell)
EXCEL_BLOCKS = None

# BQuant source (--bql): your three queries in string form.  Only runs
# inside BQuant, where `import bql` works; the Desktop API has no BQL service.
BQL_QUERIES = {
    'us_eco': "get(calendar(dates=range(0D, 3M), relevancy=HIGH)) for('US Country')",
    'central_banks': (
        "get(calendar(type=central_banks,relevancy=medium,dates=range(0d,+3m)).RELEASE_DATE, "
        "calendar(type=central_banks,relevancy=medium,dates=range(0d,+3m)).RELEASE_TIME, "
        "calendar(type=central_banks,relevancy=medium,dates=range(0d,+3m)).EVENT_NAME, "
        "calendar(type=central_banks,relevancy=medium,dates=range(0d,+3m)).SURVEY_MEDIAN, "
        "calendar(type=central_banks,relevancy=medium,dates=range(0d,+3m)).PRIOR) "
        "for(['US Country','TW Country','KR Country','MY Country','TH Country','ID Country',"
        "'IN Country','CN Country','JP Country','AU Country'])"),
    'hsi_earnings': (
        "get(id().value, name().value, expected_report_dt().value, expected_report_time().value, "
        "earnings_conf_call_dt().value, earnings_conf_call_time().value) "
        "for(filter(members('HSI Index'), expected_report_dt >= btoday() "
        "and expected_report_dt <= btoday()+91d))"),
}

SECTION_TITLES = {
    'central_banks': 'Central banks - next decisions',
    'us_eco':        'US economic calendar - high relevance',
    'hsi_earnings':  'HSI constituents - expected earnings',
}
# ============================================================ END CONFIG ===


# ================================================================ ENGINE ===


# -------------------------------------------------------------- helpers ---
def norm(s):
    """'  U. of Mich.  Sentiment ' -> 'u. of mich. sentiment' for exclude matching."""
    return re.sub(r'\s+', ' ', str(s or '')).strip().lower()


_EXCLUDE = {norm(e) for e in EXCLUDE_EVENTS}


def is_excluded(event):
    return norm(event) in _EXCLUDE


def flatten_fields(spec):
    seen, out = set(), []
    for cands in spec.values():
        for f in cands:
            if f not in seen:
                seen.add(f)
                out.append(f)
    return out


def as_date(v):
    """date / datetime / Excel serial / 'YYYY-MM-DD' -> date, else None."""
    if v is None or v == '':
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        if 20000 < v < 80000:                       # Excel serial
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).date()
        return None
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y',
                '%m/%d/%Y', '%d-%b-%Y', '%d %b %Y', '%Y%m%d'):
        try:
            return dt.datetime.strptime(s[:len(fmt) + 2] if 'T' in fmt else s, fmt).date()
        except ValueError:
            continue
    return None


def as_time(v):
    """time / datetime / Excel fraction / '08:30' -> 'HH:MM'; other text kept ('Bef-mkt')."""
    if v is None or v == '':
        return None
    if isinstance(v, dt.datetime):
        return v.strftime('%H:%M') if (v.hour or v.minute) else None
    if isinstance(v, dt.time):
        return v.strftime('%H:%M')
    if isinstance(v, (int, float)):
        if 0 <= v < 1:
            mins = int(round(v * 24 * 60))
            return f'{mins // 60:02d}:{mins % 60:02d}'
        return None
    s = str(v).strip()
    m = re.match(r'^(\d{1,2}):(\d{2})', s)
    if m:
        return f'{int(m.group(1)):02d}:{m.group(2)}'
    return s or None


def time_key(t):
    """Sort key: before-market < clock times < after-market < unknown."""
    if not t:
        return 24 * 60 + 2
    m = re.match(r'^(\d{2}):(\d{2})', t)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    tl = t.lower()
    if 'bef' in tl or 'pre' in tl:
        return -1
    if 'aft' in tl or 'post' in tl:
        return 24 * 60 + 1
    return 24 * 60


def shift_tz(date, time_str, tz_in, tz_out):
    """Move a local release time into tz_out.  Text times ('Bef-mkt') pass through."""
    if not (date and time_str and tz_in and tz_out):
        return date, time_str
    m = re.match(r'^(\d{2}):(\d{2})$', time_str)
    if not m:
        return date, time_str
    try:
        from zoneinfo import ZoneInfo
        local = dt.datetime(date.year, date.month, date.day, int(m.group(1)),
                            int(m.group(2)), tzinfo=ZoneInfo(tz_in))
        moved = local.astimezone(ZoneInfo(tz_out))
        return moved.date(), moved.strftime('%H:%M')
    except Exception:
        return date, time_str


def fmt_num(v, decimals=None):
    if v is None or v == '':
        return '\u2013'
    if isinstance(v, str):
        try:
            v = float(v.replace(',', ''))
        except ValueError:
            return v
    if decimals is not None:
        return f'{v:,.{decimals}f}'
    s = f'{v:,.4f}'.rstrip('0').rstrip('.')
    return '0' if s in ('', '-0') else s


def fmt_date(d):
    return d.strftime('%a %d %b') if d else '\u2013'


def day_tag(d, today):
    n = (d - today).days
    if n == 0:
        return 'TODAY'
    if n == 1:
        return 'tomorrow'
    if n < 0:
        return f'{-n}d ago'
    return f'in {n} days'


def pick(rec, candidates):
    for f in candidates:
        v = rec.get(f)
        if v is not None and v != '':
            return v
    return None


def pick_date(rec, date_fields, time_fields, today):
    """First candidate date that is today or later; a datetime also supplies the time."""
    when, when_time = None, None
    for f in date_fields:
        v = rec.get(f)
        d = as_date(v)
        if d is not None and d >= today:
            when = d
            if isinstance(v, dt.datetime):
                when_time = as_time(v)
            break
    if when is None:
        return None, None
    t = pick(rec, time_fields)
    if t is not None:
        when_time = as_time(t) or when_time
    return when, when_time


def element_value(el):
    """Python value of a response element: scalars as-is, bulk fields as list of dicts."""
    if el.isArray():
        if el.isComplexType():
            return [element_dict(el.getValueAsElement(i)) for i in range(el.numValues())]
        return [el.getValue(i) for i in range(el.numValues())]
    if el.isComplexType():
        return element_dict(el)
    if el.isNull():
        return None
    try:
        return el.getValue()
    except Exception:
        return el.getValueAsString()


def element_dict(el):
    return {str(el.getElement(i).name()): element_value(el.getElement(i))
            for i in range(el.numElements())}


class Section:
    def __init__(self, key, rows, notes=None):
        self.key = key
        self.title = SECTION_TITLES.get(key, key)
        self.rows = rows
        self.notes = notes or []
        self.n_excluded = None       # us_eco: rows dropped by EXCLUDE_EVENTS this run


# ------------------------------------------------------------ bloomberg ---
class Bloomberg:
    """One long-lived session, re-used for every request (the price_alarm pattern).

    ref() is the same request/response snapshot as =BDP(); nothing here needs
    BQL.  Invalid mnemonics and dead tickers are collected, not raised: the
    section builders decide what a gap means and the footer says what is missing.
    """

    def __init__(self, host=BBG_HOST, port=BBG_PORT, blpapi_module=None):
        self.host, self.port = host, port
        self.blpapi = blpapi_module
        self.session = None
        self.services = {}
        self.field_errors = {}          # mnemonic -> Bloomberg's message (first seen)
        self.bad_securities = {}        # security -> message

    def connect(self):
        if self.blpapi is None:
            try:
                import blpapi
            except ImportError:
                raise RuntimeError(
                    'blpapi is not installed in this Python - run from the prompt the desk '
                    'notebooks use, or: pip install blpapi --index-url='
                    'https://blpapi.bloomberg.com/repository/releases/python/simple/')
            self.blpapi = blpapi
        opts = self.blpapi.SessionOptions()
        opts.setServerHost(self.host)
        opts.setServerPort(self.port)
        self.session = self.blpapi.Session(opts)
        if not self.session.start():
            raise RuntimeError('cannot start a Bloomberg session - is the terminal running '
                               'and logged in on this machine?')
        return self

    def close(self):
        try:
            if self.session is not None:
                self.session.stop()
        except Exception:
            pass
        self.session, self.services = None, {}

    def service(self, name):
        if name not in self.services:
            if not self.session.openService(name):
                raise RuntimeError(f'cannot open {name}')
            self.services[name] = self.session.getService(name)
        return self.services[name]

    def _drain(self, request, take):
        """Send one request, hand every message to take(msg), stop at the final RESPONSE."""
        self.session.sendRequest(request)
        for _ in range(EVENT_SPINS):                 # bounded: never hang the desk
            ev = self.session.nextEvent(500)
            for msg in ev:
                take(msg)
            if ev.eventType() == self.blpapi.Event.RESPONSE:
                return
        raise RuntimeError('Bloomberg did not answer within ~2 min - check the terminal session')

    def ref(self, securities, fields):
        """ReferenceDataRequest -> {security: {field: value}} (bulk fields: list of dicts)."""
        out = {}
        svc = self.service('//blp/refdata')
        securities = list(securities)

        def take(msg):
            if msg.hasElement('responseError'):
                err = msg.getElement('responseError')
                raise RuntimeError('Bloomberg responseError: ' + (
                    err.getElementAsString('message') if err.hasElement('message') else '?'))
            if not msg.hasElement('securityData'):
                return
            arr = msg.getElement('securityData')
            for i in range(arr.numValues()):
                sd = arr.getValueAsElement(i)
                sec = sd.getElementAsString('security')
                if sd.hasElement('securityError'):
                    err = sd.getElement('securityError')
                    self.bad_securities[sec] = (err.getElementAsString('message')
                                                if err.hasElement('message') else 'security error')
                    continue
                if sd.hasElement('fieldExceptions'):
                    fe = sd.getElement('fieldExceptions')
                    for j in range(fe.numValues()):
                        x = fe.getValueAsElement(j)
                        fid = x.getElementAsString('fieldId')
                        msg_txt = 'field exception'
                        if x.hasElement('errorInfo'):
                            info = x.getElement('errorInfo')
                            parts = [info.getElementAsString(k) for k in ('subcategory', 'message')
                                     if info.hasElement(k)]
                            msg_txt = ' / '.join(p for p in parts if p) or msg_txt
                        self.field_errors.setdefault(fid, msg_txt)
                row = {}
                if sd.hasElement('fieldData'):
                    fd = sd.getElement('fieldData')
                    for j in range(fd.numElements()):
                        el = fd.getElement(j)
                        row[str(el.name())] = element_value(el)
                out[sec] = row

        for i in range(0, len(securities), CHUNK):
            rq = svc.createRequest('ReferenceDataRequest')
            for s in securities[i:i + CHUNK]:
                rq.getElement('securities').appendValue(s)
            for f in fields:
                rq.getElement('fields').appendValue(f)
            self._drain(rq, take)
        return out

    def field_search(self, spec, limit=25):
        """//blp/apiflds FieldSearchRequest -> [(mnemonic, description)] - FLDS from Python."""
        svc = self.service('//blp/apiflds')
        rq = svc.createRequest('FieldSearchRequest')
        rq.set('searchSpec', spec)
        found = []

        def take(msg):
            if not msg.hasElement('fieldData'):
                return
            arr = msg.getElement('fieldData')
            for i in range(arr.numValues()):
                fd = arr.getValueAsElement(i)
                if not fd.hasElement('fieldInfo'):
                    continue
                info = fd.getElement('fieldInfo')
                mn = info.getElementAsString('mnemonic') if info.hasElement('mnemonic') else ''
                de = info.getElementAsString('description') if info.hasElement('description') else ''
                found.append((mn, de))

        self._drain(rq, take)
        return found[:limit]

    def instrument_search(self, query, yellow='YK_FILTER_INDX', limit=8):
        """//blp/instruments instrumentListRequest -> [(security, description)] - SECF from Python."""
        svc = self.service('//blp/instruments')
        rq = svc.createRequest('instrumentListRequest')
        rq.set('query', query)
        rq.set('maxResults', limit)
        rq.set('yellowKeyFilter', yellow)
        found = []

        def take(msg):
            if not msg.hasElement('results'):
                return
            arr = msg.getElement('results')
            for i in range(arr.numValues()):
                r = arr.getValueAsElement(i)
                found.append((r.getElementAsString('security'), r.getElementAsString('description')))

        self._drain(rq, take)
        return found

    def probe_bql_service(self, query):
        """Does this terminal expose BQL over the API?  Officially it does not (BQL lives
        in Excel and BQuant), but some installs report a //blp/bqlsvc.  Nothing is assumed:
        open it, print its schema, and if a request has an 'expression' slot, try the query."""
        lines = []
        name = '//blp/bqlsvc'
        try:
            opened = self.session.openService(name)
        except Exception as e:
            opened = False
            lines.append(f'{name}: openService raised {e!r}')
        if not opened:
            lines.append(f'{name}: not available (expected - BQL is not part of the Desktop API). '
                         'Verdict: stay on the refdata path, or run --bql inside BQuant.')
            return lines
        svc = self.session.getService(name)
        lines.append(f'{name}: OPENED. Operations and request schemas:')
        for i in range(svc.numOperations()):
            op = svc.getOperation(i)
            lines.append(f'  - {op.name()}')
            try:
                lines.append('      ' + str(op.requestDefinition()).replace('\n', '\n      '))
            except Exception as e:
                lines.append(f'      (schema print failed: {e!r})')
        for i in range(svc.numOperations()):
            op_name = str(svc.getOperation(i).name())
            try:
                rq = svc.createRequest(op_name)
                el = rq.asElement()
                slot = next((str(el.getElement(j).name()) for j in range(el.numElements())
                             if str(el.getElement(j).name()).lower() in ('expression', 'query', 'bql')),
                            None)
                if slot is None:
                    continue
                rq.set(slot, query)
                lines.append(f'  Sent {op_name}.{slot} = {query[:80]}...')
                dump = []
                self._drain(rq, lambda m: dump.append(m.toString()))
                for d in dump[:3]:
                    lines.append('      ' + d[:3000].replace('\n', '\n      '))
                lines.append('  Verdict: bring this output back - if it is the same table Excel shows, '
                             'the three BQL strings can run verbatim via the API.')
            except Exception as e:
                lines.append(f'  {op_name}: attempt failed: {e!r}')
        return lines


# -------------------------------------------------------------- sections ---
def field_notes(bbg, spec, records):
    """What a rejected mnemonic actually cost.

    A mnemonic Bloomberg refuses is only worth mentioning if nothing else in its
    role returned data - otherwise the fallback covered it and saying so is noise.
    A role that produced nothing at all is the line that matters: that column is
    blank in the email and the reader should know why."""
    notes = []
    for role, cands in spec.items():
        rejected = [f for f in cands if f in bbg.field_errors]
        got = any(rec.get(f) not in (None, '') for rec in records for f in cands)
        if got or not rejected:
            continue
        notes.append(f'"{role}" is blank: Bloomberg returned nothing for '
                     + ', '.join(f'{f} [{bbg.field_errors[f]}]' for f in rejected)
                     + ' - correct the mnemonic in the script config (run --probe for the list).')
    return notes


def pull_us_eco(bbg, today):
    fields = flatten_fields(ECO_FIELDS)
    tickers = [t for t, _ in US_ECO_TICKERS]
    data = bbg.ref(tickers, fields)
    horizon = today + dt.timedelta(days=LOOKAHEAD_DAYS)
    rows, no_date, excluded, low_rel = [], [], [], []
    for tkr, label in US_ECO_TICKERS:
        rec = data.get(tkr)
        if rec is None:
            continue                                  # dead ticker: bbg.bad_securities has it
        when, when_time = pick_date(rec, ECO_FIELDS['date'], ECO_FIELDS['time'], today)
        if when is None:
            no_date.append(tkr)
            continue
        if when > horizon:
            continue
        if TIME_IN_LOCAL_TZ:
            when, when_time = shift_tz(when, when_time, US_TZ, TIME_TZ_OUT)
        rel = pick(rec, ECO_FIELDS['relevance'])
        row = dict(date=when, time=when_time, country='US', bank='', event=label,
                   bbg_name=pick(rec, ECO_FIELDS['bbg_name']),
                   survey=pick(rec, ECO_FIELDS['survey']), prior=pick(rec, ECO_FIELDS['prior']),
                   relevance=rel, ticker=tkr)
        if is_excluded(label):
            excluded.append(label)
            continue
        if RELEVANCE_MIN and rel is not None and float(rel) < RELEVANCE_MIN:
            low_rel.append(label)
            continue
        rows.append(row)
    notes = []
    dead = [t for t in tickers if t in bbg.bad_securities]
    if dead:
        notes.append(f'{len(dead)} US ticker(s) unknown to Bloomberg, not shown: ' + ', '.join(dead))
    if no_date:
        notes.append(f'{len(no_date)} US series returned no upcoming release date: '
                     + ', '.join(no_date[:8]) + (' ...' if len(no_date) > 8 else ''))
    notes += field_notes(bbg, ECO_FIELDS, list(data.values()))
    print(f'  us_eco: {len(rows)} rows kept, {len(excluded)} excluded by name, '
          f'{len(low_rel)} below RELEVANCE_MIN, {len(no_date)} without a date, {len(dead)} dead tickers')
    sec = Section('us_eco', sort_rows(rows), notes)
    sec.n_excluded = len(excluded)
    return sec


def pull_central_banks(bbg, today):
    fields = flatten_fields(ECO_FIELDS)
    tickers = [t for _, _, t, _, _ in CENTRAL_BANKS]
    data = bbg.ref(tickers, fields)
    horizon = today + dt.timedelta(days=LOOKAHEAD_DAYS)
    rows, no_date = [], []
    for country, bank, tkr, label, tz in CENTRAL_BANKS:
        rec = data.get(tkr)
        if rec is None:
            continue
        when, when_time = pick_date(rec, ECO_FIELDS['date'], ECO_FIELDS['time'], today)
        if when is None:
            no_date.append(tkr)
            continue
        if when > horizon:
            continue
        if TIME_IN_LOCAL_TZ:
            when, when_time = shift_tz(when, when_time, tz, TIME_TZ_OUT)
        rows.append(dict(date=when, time=when_time, country=country, bank=bank, event=label,
                         bbg_name=pick(rec, ECO_FIELDS['bbg_name']),
                         survey=pick(rec, ECO_FIELDS['survey']), prior=pick(rec, ECO_FIELDS['prior']),
                         relevance=pick(rec, ECO_FIELDS['relevance']), ticker=tkr))
    notes = []
    dead = [t for t in tickers if t in bbg.bad_securities]
    if dead:
        notes.append('Central-bank ticker(s) unknown to Bloomberg, not shown: ' + ', '.join(dead))
    if no_date:
        notes.append('No upcoming decision date returned for: ' + ', '.join(no_date))
    notes += field_notes(bbg, ECO_FIELDS, list(data.values()))
    notes.append('Only scheduled rate decisions are available through the API; speeches and '
                 'minutes (which BQL type=central_banks includes) are not tickers.')
    print(f'  central_banks: {len(rows)} rows, {len(no_date)} without a date, {len(dead)} dead tickers')
    return Section('central_banks', sort_rows(rows), notes)


def bulk_field(rec, field, *want):
    """Values out of a bulk field, matched by substring on the sub-element name.

    Element names differ between terminal versions ('Member Ticker and Exchange
    Code' vs 'Ticker'), which is why v33_hk_basket_full.py parses DVD_HIST_ALL by
    substring rather than by exact key.  Same trick here."""
    out = []
    for row in (rec.get(field) or []):
        if not isinstance(row, dict):
            continue
        hit = next((v for k, v in row.items()
                    if any(w in str(k).lower() for w in want)), None)
        if hit is None:
            hit = next(iter(row.values()), None)
        if hit not in (None, ''):
            out.append(hit)
    return out


def next_announced(rec, today):
    """Earliest announcement date at or after today, out of ERN_ANN_DT_AND_PER.

    The field returns history and estimated forward dates together, so past rows
    are dropped rather than trusted."""
    dates = [d for d in (as_date(v) for v in bulk_field(rec, EARN_BULK_DATE, 'announcement', 'date'))
             if d is not None and d >= today]
    return min(dates) if dates else None


def hsi_member_tickers(bbg):
    rec = bbg.ref([HSI_INDEX], ['INDX_MEMBERS']).get(HSI_INDEX, {})
    out = []
    for code in bulk_field(rec, 'INDX_MEMBERS', 'ticker'):
        code = str(code).strip()
        if code:
            out.append(code if code.upper().endswith(' EQUITY') else f'{code} Equity')
    return out


def pull_hsi_earnings(bbg, today):
    tickers = hsi_member_tickers(bbg)
    fields = flatten_fields(EARN_FIELDS) + [EARN_BULK_DATE]
    data = bbg.ref(tickers, fields) if tickers else {}
    horizon = today + dt.timedelta(days=LOOKAHEAD_DAYS)
    rows = []
    for tkr in tickers:
        rec = data.get(tkr)
        if rec is None:
            continue
        when = as_date(pick(rec, EARN_FIELDS['date']))
        from_bulk = False
        if when is None or when < today:
            when = next_announced(rec, today)       # proven bulk fallback
            from_bulk = when is not None
        if when is None or when < today or when > horizon:
            continue
        call_d = as_date(pick(rec, EARN_FIELDS['call_date']))
        rows.append(dict(date=when, time=as_time(pick(rec, EARN_FIELDS['time'])),
                         ticker=tkr[:-len(' Equity')] if tkr.upper().endswith(' EQUITY') else tkr,
                         name=pick(rec, EARN_FIELDS['name']) or '',
                         call_date=call_d, call_time=as_time(pick(rec, EARN_FIELDS['call_time'])),
                         estimated=from_bulk))
    notes = []
    if not tickers:
        notes.append(f'INDX_MEMBERS returned nothing for {HSI_INDEX}.')
    dead = [t for t in tickers if t in bbg.bad_securities]
    if dead:
        notes.append('Members Bloomberg could not resolve: ' + ', '.join(dead))
    notes += field_notes(bbg, EARN_FIELDS, list(data.values()))
    n_bulk = sum(1 for r in rows if r['estimated'])
    if n_bulk:
        notes.append(f'{n_bulk} date(s) marked "est." came from {EARN_BULK_DATE} because '
                     f'EXPECTED_REPORT_DT was empty for that name.')
    print(f'  hsi_earnings: {len(tickers)} members, {len(rows)} reporting inside {LOOKAHEAD_DAYS}d, '
          f'{len(dead)} unresolved')
    return Section('hsi_earnings', sorted(rows, key=lambda r: (r['date'], time_key(r['time']), r['ticker'])),
                   notes)


def sort_rows(rows):
    def rel(r):
        try:
            return -float(r.get('relevance') or 0)
        except (TypeError, ValueError):
            return 0
    return sorted(rows, key=lambda r: (r['date'], time_key(r['time']), rel(r), r['event']))


def build_sections_blpapi(today, blpapi_module=None):
    bbg = Bloomberg(blpapi_module=blpapi_module).connect()
    try:
        print('Pulling from Bloomberg ...')
        return {
            'central_banks': pull_central_banks(bbg, today),
            'us_eco': pull_us_eco(bbg, today),
            'hsi_earnings': pull_hsi_earnings(bbg, today),
        }, bbg
    finally:
        bbg.close()


# ---------------------------------------------------- excel / bql sources ---
CAL_SPEC = [('date', ['RELEASE_DATE', 'DATE']), ('time', ['RELEASE_TIME', 'TIME']),
            ('event', ['EVENT_NAME', 'EVENT', 'NAME']), ('survey', ['SURVEY']),
            ('prior', ['PRIOR']), ('relevance', ['RELEVANC']), ('id', ['=ID', 'COUNTRY']),
            ('ticker', ['TICKER'])]
EARN_SPEC = [('date', ['EXPECTED_REPORT_DT', 'REPORT_DT']), ('time', ['EXPECTED_REPORT_TIME', 'REPORT_TIME']),
             ('call_date', ['CONF_CALL_DT']), ('call_time', ['CONF_CALL_TIME']),
             ('ticker', ['=ID', 'TICKER']), ('name', ['NAME'])]


def map_columns(headers, spec):
    up = [str(h or '').strip().upper() for h in headers]
    taken, out = set(), {}
    for key, toks in spec:
        for tok in toks:
            exact = tok.startswith('=')
            tok = tok.lstrip('=')
            hit = next((i for i, h in enumerate(up) if i not in taken
                        and ((h == tok) if exact else (tok in h))), None)
            if hit is not None:
                out[key] = hit
                taken.add(hit)
                break
    return out


def classify_header(headers):
    up = ' | '.join(str(h or '').upper() for h in headers)
    if 'EXPECTED_REPORT' in up or 'CONF_CALL' in up:
        return 'hsi_earnings'
    if 'EVENT' in up or 'RELEASE_DATE' in up:
        return 'calendar'
    return None


def rows_from_table(headers, body, kind, today):
    """Rows in the same shape the API path produces, from any table with BQL-style headers."""
    horizon = today + dt.timedelta(days=LOOKAHEAD_DAYS)
    if kind == 'hsi_earnings':
        m = map_columns(headers, EARN_SPEC)
        rows = []
        for vals in body:
            g = lambda k: vals[m[k]] if k in m and m[k] < len(vals) else None
            when = as_date(g('date'))
            if when is None or when < today or when > horizon:
                continue
            tkr = str(g('ticker') or '').strip()
            tkr = tkr[:-len(' Equity')] if tkr.upper().endswith(' EQUITY') else tkr
            rows.append(dict(date=when, time=as_time(g('time')), ticker=tkr, name=str(g('name') or ''),
                             call_date=as_date(g('call_date')), call_time=as_time(g('call_time')),
                             estimated=False))
        return 'hsi_earnings', sorted(rows, key=lambda r: (r['date'], time_key(r['time']), r['ticker'])), m
    m = map_columns(headers, CAL_SPEC)
    rows, ids = [], set()
    for vals in body:
        g = lambda k: vals[m[k]] if k in m and m[k] < len(vals) else None
        raw_date = g('date')
        when = as_date(raw_date)
        if when is None or when < today or when > horizon:
            continue
        t = as_time(g('time')) or (as_time(raw_date) if isinstance(raw_date, dt.datetime) else None)
        ident = str(g('id') or '').strip()
        ids.add(ident)
        country = ident.replace('Country', '').strip() or 'US'
        event = str(g('event') or '').strip()
        if is_excluded(event):
            continue
        rows.append(dict(date=when, time=t, country=country, bank='', event=event, bbg_name=None,
                         survey=g('survey'), prior=g('prior'), relevance=g('relevance'),
                         ticker=str(g('ticker') or '')))
    key = 'central_banks' if len(ids - {''}) >= 2 else 'us_eco'
    return key, sort_rows(rows), m


def read_excel_sections(path, today):
    """openpyxl data_only: the values the BQL add-in last wrote into the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    blocks = []
    if EXCEL_BLOCKS:
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        for sheet, cell, kind in EXCEL_BLOCKS:
            ws = wb[sheet]
            col_letter, row0 = coordinate_from_string(cell)
            c0 = column_index_from_string(col_letter)
            grid = list(ws.iter_rows(min_row=row0, min_col=c0, values_only=True))
            hdr = list(grid[0])
            while hdr and hdr[-1] in (None, ''):
                hdr.pop()
            body = []
            for vals in grid[1:]:
                vals = list(vals[:len(hdr)])
                if all(v in (None, '') for v in vals):
                    break
                body.append(vals)
            blocks.append((sheet, cell, kind, hdr, body))
    else:
        for ws in wb.worksheets:
            grid = list(ws.iter_rows(values_only=True))
            claimed = set()                     # (row, col) cells already inside a block
            for r, rowvals in enumerate(grid):
                c = 0
                while c < len(rowvals):
                    if rowvals[c] in (None, '') or (r, c) in claimed:
                        c += 1
                        continue
                    c1 = c
                    while c1 + 1 < len(rowvals) and rowvals[c1 + 1] not in (None, ''):
                        c1 += 1
                    hdr = list(rowvals[c:c1 + 1])
                    kind = classify_header(hdr)
                    if kind:
                        body = []
                        for rr in range(r + 1, len(grid)):
                            vals = list(grid[rr][c:c1 + 1]) if len(grid[rr]) > c else []
                            vals += [None] * (len(hdr) - len(vals))
                            if all(v in (None, '') for v in vals):
                                break
                            body.append(vals)
                            claimed.update((rr, cc) for cc in range(c, c1 + 1))
                        blocks.append((ws.title, f'{openpyxl.utils.get_column_letter(c + 1)}{r + 1}',
                                       kind, hdr, body))
                    c = c1 + 1
    sections = {}
    for sheet, cell, kind, hdr, body in blocks:
        key, rows, colmap = rows_from_table(hdr, body, kind, today)
        if kind not in ('calendar', key):
            key = kind
        used = ', '.join(f'{k}<-{hdr[i]}' for k, i in colmap.items())
        print(f'  excel block {sheet}!{cell}: {key}, {len(body)} rows read, {len(rows)} kept  [{used}]')
        if key in sections:
            print(f'    (a second {key} block was found - the later one wins; pin EXCEL_BLOCKS to choose)')
        sections[key] = Section(key, rows, [f'Source: {os.path.basename(path)} ({sheet}!{cell})'])
    return sections


def build_sections_bql(today):
    """Inside BQuant only: run the three BQL strings verbatim."""
    try:
        import bql
    except ImportError:
        raise RuntimeError('`import bql` failed - the bql package only exists inside BQuant. '
                           'Run this on the terminal PC without --bql instead.')
    bq = bql.Service()
    sections = {}
    for key, query in BQL_QUERIES.items():
        res = bq.execute(query)
        frames = [r.df() for r in res]
        df = frames[0] if len(frames) == 1 else bql.combined_df(res)
        df = df.reset_index()
        headers = [str(c) for c in df.columns]
        body = df.values.tolist()
        kind = 'hsi_earnings' if key == 'hsi_earnings' else 'calendar'
        found_key, rows, colmap = rows_from_table(headers, body, kind, today)
        print(f'  bql {key}: {len(body)} rows, {len(rows)} kept; columns {headers}')
        sections[key] = Section(key, rows, ['Source: BQL (BQuant)'])
    return sections


# ------------------------------------------------------------ rendering ---
FONT = 'font-family:Calibri,Arial,Helvetica,sans-serif;'
NAVY, BAND, TODAY_BG, LINE, MUTED, STRIPE = '#1F3864', '#EDF1F7', '#FFF4D6', '#E1E5EC', '#6B7280', '#F8F9FB'
TH = (f'padding:6px 8px;background:{NAVY};color:#ffffff;font-size:10pt;font-weight:bold;'
      'text-align:{align};white-space:nowrap;')
TD = f'padding:5px 8px;border-bottom:1px solid {LINE};font-size:10pt;vertical-align:top;text-align:{{align}};'


def esc(s):
    return _html.escape('' if s is None else str(s), quote=True)


def band_row(d, today, ncols):
    bg = TODAY_BG if d == today else BAND
    tag = day_tag(d, today)
    return (f'<tr><td colspan="{ncols}" bgcolor="{bg}" style="padding:6px 8px;background:{bg};'
            f'font-size:10pt;font-weight:bold;color:{NAVY};border-bottom:1px solid {LINE};">'
            f'{esc(fmt_date(d))}'
            f'<span style="font-weight:normal;color:{MUTED};"> &nbsp;&middot;&nbsp; {esc(tag)}</span>'
            f'</td></tr>')


def html_table(cols, groups, today):
    """cols: [(header, width_px|None, align)], groups: [(date, [[cell,...], ...])]."""
    ncols = len(cols)
    head = ''.join(
        f'<th{" width=%d" % w if w else ""} style="{TH.format(align=a)}">{esc(h)}</th>'
        for h, w, a in cols)
    body = []
    for d, rows in groups:
        body.append(band_row(d, today, ncols))
        for i, cells in enumerate(rows):
            bg = STRIPE if i % 2 else '#ffffff'
            tds = ''.join(
                f'<td bgcolor="{bg}" style="{TD.format(align=cols[j][2])}background:{bg};'
                f'{"white-space:nowrap;" if cols[j][1] else ""}">{cells[j]}</td>'
                for j in range(ncols))
            body.append(f'<tr>{tds}</tr>')
    return (f'<table width="760" cellpadding="0" cellspacing="0" border="0" '
            f'style="border-collapse:collapse;width:760px;{FONT}">'
            f'<thead><tr>{head}</tr></thead><tbody>{"".join(body)}</tbody></table>')


def group_by_date(rows):
    groups, cur, cur_d = [], [], None
    for r in rows:
        if r['date'] != cur_d:
            if cur:
                groups.append((cur_d, cur))
            cur, cur_d = [], r['date']
        cur.append(r)
    if cur:
        groups.append((cur_d, cur))
    return groups


def render_calendar(section, today):
    is_cb = section.key == 'central_banks'
    if is_cb:
        cols = [('Time', 58, 'left'), ('Ctry', 44, 'left'), ('Bank', 52, 'left'),
                ('Decision', None, 'left'), ('Survey', 72, 'right'), ('Current', 72, 'right')]
    else:
        cols = [('Time', 58, 'left'), ('Event', None, 'left'), ('Survey', 80, 'right'),
                ('Prior', 80, 'right'), ('Rel.', 46, 'right')]
    groups = []
    for d, rows in group_by_date(section.rows):
        out = []
        for r in rows:
            dec = 2 if is_cb else None
            if is_cb:
                out.append([esc(r['time'] or '\u2013'), esc(r['country']), esc(r['bank']),
                            esc(r['event']), esc(fmt_num(r['survey'], dec)), esc(fmt_num(r['prior'], dec))])
            else:
                rel = r.get('relevance')
                rel_s = '\u2013' if rel in (None, '') else f'{float(rel):.0f}'
                out.append([esc(r['time'] or '\u2013'), esc(r['event']), esc(fmt_num(r['survey'])),
                            esc(fmt_num(r['prior'])),
                            f'<span style="color:{MUTED}">{esc(rel_s)}</span>'])
        groups.append((d, out))
    return html_table(cols, groups, today)


def render_earnings(section, today):
    cols = [('Ticker', 78, 'left'), ('Company', None, 'left'), ('Report time', 90, 'left'),
            ('Conference call', 150, 'left')]
    groups = []
    for d, rows in group_by_date(section.rows):
        out = []
        for r in rows:
            if r['call_date']:
                call = fmt_date(r['call_date']) + (f' {r["call_time"]}' if r['call_time'] else '')
            elif r['call_time']:
                call = r['call_time']
            else:
                call = '\u2013'
            when_cell = esc(r['time'] or '\u2013')
            if r.get('estimated'):
                when_cell += f'<span style="color:{MUTED};"> est.</span>'
            out.append([f'<b>{esc(r["ticker"])}</b>', esc(r['name']), when_cell, esc(call)])
        groups.append((d, out))
    return html_table(cols, groups, today)


def render_section(section, number, today):
    n = len(section.rows)
    if section.key == 'hsi_earnings':
        table = render_earnings(section, today)
        count = f'{n} name{"s" if n != 1 else ""} reporting in the next {LOOKAHEAD_DAYS} days'
    else:
        table = render_calendar(section, today)
        count = (f'{n} decision{"s" if n != 1 else ""} scheduled in the next {LOOKAHEAD_DAYS} days'
                 if section.key == 'central_banks' else
                 f'{n} release{"s" if n != 1 else ""} in the next {LOOKAHEAD_DAYS} days'
                 + (f' ({section.n_excluded} low-value release{"s" if section.n_excluded != 1 else ""} '
                    f'filtered out)' if section.n_excluded else ''))
    if n == 0:
        table = (f'<p style="color:{MUTED};font-size:10pt;margin:4px 0 0 0;">Nothing returned - see the '
                 f'notes at the bottom.</p>')
    return (f'<p style="font-size:13pt;font-weight:bold;color:{NAVY};margin:22px 0 2px 0;">'
            f'{number}. {esc(section.title)}</p>'
            f'<p style="font-size:9.5pt;color:{MUTED};margin:0 0 8px 0;">{esc(count)}</p>'
            f'{table}')


def render_email(email_cfg, sections, today, now, source_label):
    parts = [f'<div style="{FONT}font-size:11pt;color:#1a1a1a;max-width:780px;">',
             f'<p style="margin:0 0 10px 0;">{esc(GREETING)}</p>',
             f'<p style="margin:0 0 6px 0;">{esc(email_cfg["intro"])}</p>']
    notes = []
    for i, key in enumerate(email_cfg['sections'], start=1):
        sec = sections.get(key) or Section(key, [], [f'No {key} data was produced.'])
        parts.append(render_section(sec, i, today))
        notes += sec.notes
    parts.append(f'<p style="font-size:8.5pt;color:{MUTED};margin:18px 0 0 0;">'
                 f'Source: {esc(source_label)}, pulled {esc(now.strftime("%d %b %Y %H:%M"))}. '
                 f'Times as returned by Bloomberg'
                 f'{" (converted to " + esc(TIME_TZ_OUT) + ")" if TIME_IN_LOCAL_TZ else ""}. '
                 f'Generated by daily_email.py.</p>')
    if FOOTER_NOTES and notes:
        parts.append(f'<p style="font-size:8.5pt;color:{MUTED};margin:4px 0 0 0;">Data notes: '
                     + ' &nbsp;|&nbsp; '.join(esc(n) for n in notes) + '</p>')
    parts.append('</div>')
    n_hsi = len(sections['hsi_earnings'].rows) if 'hsi_earnings' in sections else 0
    subject = email_cfg['subject'].format(date=today.strftime('%a %d %b %Y'), n_hsi=n_hsi)
    return subject, ''.join(parts)


def full_document(subject, fragment):
    return (f'<html><head><meta charset="utf-8"><title>{esc(subject)}</title></head>'
            f'<body style="margin:0;padding:16px;background:#ffffff;">{fragment}</body></html>')


def splice_into_signature(signature_html, fragment):
    """Put our body above Outlook's default signature."""
    m = re.search(r'<body[^>]*>', signature_html or '', flags=re.I)
    if not m:
        return full_document('', fragment)
    return signature_html[:m.end()] + fragment + signature_html[m.end():]


# ------------------------------------------------------------- delivery ---
def write_files(key, subject, fragment, today, now):
    os.makedirs(OUT_DIR, exist_ok=True)
    stem = os.path.join(OUT_DIR, f'{today:%Y-%m-%d}_{key}')
    doc = full_document(subject, fragment)
    with open(stem + '.html', 'w', encoding='utf-8') as fh:
        fh.write(doc)
    msg = EmailMessage()
    msg['To'] = RECIPIENTS.replace(';', ',')
    msg['Subject'] = subject
    msg['Date'] = format_datetime(now.astimezone() if now.tzinfo is None else now)
    msg['X-Unsent'] = '1'              # Outlook opens an .eml with this header in compose mode
    msg.set_content('This message is HTML - open it in an HTML-capable client.')
    msg.add_alternative(doc, subtype='html')
    with open(stem + '.eml', 'wb') as fh:
        fh.write(bytes(msg))
    return stem + '.html', stem + '.eml'


def outlook_draft(subject, fragment, display):
    """Create the draft in Outlook.  Returns (ok, message)."""
    try:
        import win32com.client
    except ImportError:
        return False, 'pywin32 is not installed (pip install pywin32) - .eml written instead'
    try:
        app = win32com.client.Dispatch('Outlook.Application')
    except Exception as e:
        return False, f'Outlook not reachable ({e}) - .eml written instead'
    mail = app.CreateItem(0)                       # olMailItem
    mail.To = RECIPIENTS
    mail.Subject = subject
    try:
        mail.GetInspector                          # touching it makes Outlook insert the signature
        signature = mail.HTMLBody or ''
    except Exception:
        signature = ''
    mail.HTMLBody = splice_into_signature(signature, fragment)
    if display:
        mail.Display()
        return True, 'opened for review'
    mail.Save()
    return True, 'saved to Drafts'


# ----------------------------------------------------------------- probe ---
def _probe_report(today, blpapi_module=None):
    bbg = Bloomberg(blpapi_module=blpapi_module).connect()
    lines = [f'daily_email --probe  {dt.datetime.now():%Y-%m-%d %H:%M}  host {BBG_HOST}:{BBG_PORT}', '']
    try:
        # 1. US calendar: every candidate field on the whole universe
        fields = flatten_fields(ECO_FIELDS)
        tickers = [t for t, _ in US_ECO_TICKERS]
        data = bbg.ref(tickers, fields)
        lines.append('== US economic calendar: field mnemonics ==')
        for role, cands in ECO_FIELDS.items():
            for f in cands:
                have = sum(1 for t in tickers if f in data.get(t, {}))
                status = ('REJECTED ' + bbg.field_errors[f]) if f in bbg.field_errors else f'ok, {have}/{len(tickers)} tickers returned it'
                lines.append(f'  {role:10s} {f:32s} {status}')
        lines.append('')
        lines.append('== US economic calendar: tickers (label = what the email shows) ==')
        lines.append(f'  {"ticker":18s} {"label":36s} {"Bloomberg NAME":44s} {"next":10s} {"time":7s} '
                     f'{"survey":>9s} {"prior":>9s} {"rel":>4s}')
        for tkr, label in US_ECO_TICKERS:
            rec = data.get(tkr)
            if rec is None:
                lines.append(f'  {tkr:18s} {label:36s} !! {bbg.bad_securities.get(tkr, "no data")}')
                continue
            when, when_time = pick_date(rec, ECO_FIELDS['date'], ECO_FIELDS['time'], today)
            rel = pick(rec, ECO_FIELDS['relevance'])
            lines.append(f'  {tkr:18s} {label:36s} {str(pick(rec, ECO_FIELDS["bbg_name"]) or "")[:44]:44s} '
                         f'{str(when or "-"):10s} {str(when_time or "-"):7s} '
                         f'{fmt_num(pick(rec, ECO_FIELDS["survey"])):>9s} {fmt_num(pick(rec, ECO_FIELDS["prior"])):>9s} '
                         f'{("" if rel is None else f"{float(rel):.0f}"):>4s}'
                         f'{"   [excluded]" if is_excluded(label) else ""}')
        lines.append('')
        lines.append('  Time-zone check: if the NFP/CPI "time" above reads 08:30 you are seeing New York '
                     'time -> set TIME_IN_LOCAL_TZ = True.  If it reads 20:30 the terminal already '
                     'converts to HKT -> leave it False.')
        lines.append('')

        # 2. Central banks
        lines.append('== Central banks ==')
        cb_tickers = [t for _, _, t, _, _ in CENTRAL_BANKS]
        cb = bbg.ref(cb_tickers, fields)
        for country, bank, tkr, label, tz in CENTRAL_BANKS:
            rec = cb.get(tkr)
            if rec is None:
                lines.append(f'  {country} {bank:5s} {tkr:16s} !! {bbg.bad_securities.get(tkr, "no data")}')
                try:
                    hits = bbg.instrument_search(f'{country} {bank} policy rate')
                    for sec, desc in hits[:5]:
                        lines.append(f'        candidate: {sec:24s} {desc}')
                except Exception as e:
                    lines.append(f'        (instrument search failed: {e!r})')
                continue
            when, when_time = pick_date(rec, ECO_FIELDS['date'], ECO_FIELDS['time'], today)
            lines.append(f'  {country} {bank:5s} {tkr:16s} {str(pick(rec, ECO_FIELDS["bbg_name"]) or "")[:40]:40s} '
                         f'next {when or "-"} {when_time or ""}  current {fmt_num(pick(rec, ECO_FIELDS["prior"]), 2)}  '
                         f'survey {fmt_num(pick(rec, ECO_FIELDS["survey"]), 2)}')
        lines.append('')

        # 3. HSI earnings
        lines.append('== HSI earnings ==')
        members = hsi_member_tickers(bbg)
        lines.append(f'  INDX_MEMBERS({HSI_INDEX}): {len(members)} members')
        efields = flatten_fields(EARN_FIELDS) + [EARN_BULK_DATE]
        sample = members[:12]
        ed = bbg.ref(sample, efields) if sample else {}
        for role, cands in EARN_FIELDS.items():
            for f in cands:
                have = sum(1 for t in sample if f in ed.get(t, {}))
                status = ('REJECTED ' + bbg.field_errors[f]) if f in bbg.field_errors else f'ok, {have}/{len(sample)} of the sample returned it'
                lines.append(f'  {role:10s} {f:28s} {status}')
        have_bulk = sum(1 for t in sample if ed.get(t, {}).get(EARN_BULK_DATE))
        lines.append(f'  date(bulk)  {EARN_BULK_DATE:28s} '
                     + (('REJECTED ' + bbg.field_errors[EARN_BULK_DATE])
                        if EARN_BULK_DATE in bbg.field_errors
                        else f'ok, {have_bulk}/{len(sample)} of the sample returned it'))
        for t in sample:
            rec = ed.get(t, {})
            lines.append(f'    {t:16s} {str(rec.get("NAME", ""))[:30]:30s} report {rec.get("EXPECTED_REPORT_DT", "-")} '
                         f'{rec.get("EXPECTED_REPORT_TIME", "")}  call {pick(rec, EARN_FIELDS["call_date"]) or "-"} '
                         f'{pick(rec, EARN_FIELDS["call_time"]) or ""}  '
                         f'next announced {next_announced(rec, today) or "-"}')
        lines.append('')

        # 4. FLDS from Python: what Bloomberg calls the things we need
        lines.append('== Field search (FLDS) - use these to correct ECO_FIELDS / EARN_FIELDS ==')
        for spec in ('future release date', 'release time', 'survey median', 'relevance',
                     'expected report', 'conference call'):
            try:
                hits = bbg.field_search(spec, limit=8)
                lines.append(f'  "{spec}":')
                for mn, de in hits:
                    lines.append(f'      {mn:32s} {de}')
            except Exception as e:
                lines.append(f'  "{spec}": search failed: {e!r}')
        lines.append('')

        # 5. BQL over the API?
        lines.append('== BQL over the API? ==')
        lines += bbg.probe_bql_service(BQL_QUERIES['hsi_earnings'])
        lines.append('')
        if bbg.field_errors:
            lines.append('== All rejected mnemonics this run ==')
            for f, m in sorted(bbg.field_errors.items()):
                lines.append(f'  {f:32s} {m}')
    finally:
        bbg.close()
    report = '\n'.join(lines)
    print(report)
    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, f'probe_{today:%Y-%m-%d}.txt')
    with open(path, 'w', encoding='utf-8') as fh:
        fh.write(report)
    print(f'\nSaved: {path}')


def _parity_report(api_sections, xl_sections):
    """Row-by-row diff of the API path against the workbook's BQL output."""
    print('\n== Parity: API vs Excel ==')
    for key in ('central_banks', 'us_eco', 'hsi_earnings'):
        a = api_sections.get(key)
        x = xl_sections.get(key)
        if a is None or x is None:
            print(f'  {key}: {"API" if a is None else "Excel"} side missing - skipped')
            continue
        if key == 'hsi_earnings':
            ka = {(r['date'], norm(r['ticker'])) for r in a.rows}
            kx = {(r['date'], norm(r['ticker'])) for r in x.rows}
        else:
            ka = {(r['date'], norm(r['event'])) for r in a.rows}
            kx = {(r['date'], norm(r['event'])) for r in x.rows}
        both = ka & kx
        print(f'  {key}: API {len(ka)} rows, Excel {len(kx)} rows, {len(both)} identical (date+name)')
        for d, e in sorted(kx - ka)[:40]:
            print(f'     Excel only: {d} {e}')
        for d, e in sorted(ka - kx)[:40]:
            print(f'     API only  : {d} {e}')


# ============================================================== FIXTURE ===
# A fake Bloomberg terminal, so --demo and --test run on any machine.  It speaks
# enough of the API to exercise the real code path above: batched requests,
# the event drain, dead tickers, rejected mnemonics and bulk fields.
_TODAY_FIXTURE = dt.date.today()


def fixture_configure(today):
    global _TODAY_FIXTURE
    _TODAY_FIXTURE = today


class FakeEvent:
    TIMEOUT, RESPONSE, PARTIAL_RESPONSE = 0, 1, 2

    def __init__(self, messages, event_type=1):
        self._messages, self._type = messages, event_type

    def __iter__(self):
        return iter(self._messages)

    def eventType(self):
        return self._type


class FakeName(str):
    pass


class FakeElement:
    """Scalar, array-of-scalars, or complex (named children), like the real thing."""

    def __init__(self, name, value=None, children=None, array=None):
        self._name, self._value = name, value
        self._children = children                 # list[(name, FakeElement)] when complex
        self._array = array                       # list[FakeElement] when an array

    def name(self):
        return FakeName(self._name)

    def isArray(self):
        return self._array is not None

    def isComplexType(self):
        # A bulk field (INDX_MEMBERS) is an array whose values are complex, and
        # the real API reports isArray() and isComplexType() both True for it.
        if self._array is not None:
            return bool(self._array) and self._array[0]._children is not None
        return self._children is not None

    def isNull(self):
        return self._value is None and self._children is None and self._array is None

    def numValues(self):
        return len(self._array or [])

    def numElements(self):
        return len(self._children or [])

    def getValueAsElement(self, i):
        return self._array[i]

    def getElement(self, key):
        if isinstance(key, int):
            return self._children[key][1]
        for n, el in (self._children or []):
            if n == key:
                return el
        raise KeyError(key)

    def hasElement(self, key):
        return any(n == key for n, _ in (self._children or []))

    def getValue(self, i=None):
        return self._array[i]._value if i is not None else self._value

    def getValueAsString(self):
        return '' if self._value is None else str(self._value)

    def getElementAsString(self, key):
        v = self.getElement(key)._value
        return '' if v is None else str(v)

    def getElementAsFloat(self, key):
        return float(self.getElement(key)._value)

    def appendValue(self, v):
        self._array.append(FakeElement('item', v))

    def set(self, key, value):
        self._children = [(n, e) for n, e in (self._children or []) if n != key]
        self._children.append((key, FakeElement(key, value)))


def fake_complex(name, pairs):
    return FakeElement(name, children=[(n, e if isinstance(e, FakeElement) else FakeElement(n, e))
                                   for n, e in pairs])


def fake_array(name, items):
    return FakeElement(name, array=list(items))


class FakeMessage:
    def __init__(self, root):
        self._root = root

    def hasElement(self, key):
        return self._root.hasElement(key)

    def getElement(self, key):
        return self._root.getElement(key)

    def toString(self):
        return f'<fake message {self._root.name()}>'


class FakeRequest:
    def __init__(self, operation):
        self.operation = operation
        self._root = FakeElement('request', children=[
            ('securities', fake_array('securities', [])),
            ('fields', fake_array('fields', [])),
        ])

    def getElement(self, key):
        return self._root.getElement(key)

    def asElement(self):
        return self._root

    def set(self, key, value):
        self._root.set(key, value)

    @property
    def securities(self):
        return [e._value for e in self._root.getElement('securities')._array]

    @property
    def fields(self):
        return [e._value for e in self._root.getElement('fields')._array]


class FakeService:
    def __init__(self, name):
        self._name = name

    def createRequest(self, operation):
        return FakeRequest(operation)

    def numOperations(self):
        return 0


class FakeSessionOptions:
    def setServerHost(self, *a): pass
    def setServerPort(self, *a): pass


# ------------------------------------------------------------ fake market ---
UNKNOWN_SECURITY = {'XXXX FAKE Index'}          # exercises the securityError path
UNKNOWN_FIELDS = {'ECO_RELEASE_DT', 'CONF_CALL_DT', 'CONF_CALL_TIME', 'ECO_RELEASE_TIME'}

HSI_MEMBERS = [
    ('700 HK', 'TENCENT'), ('939 HK', 'CCB'), ('1299 HK', 'AIA'), ('941 HK', 'CHINA MOBILE'),
    ('9988 HK', 'BABA-W'), ('388 HK', 'HKEX'), ('1398 HK', 'ICBC'), ('3988 HK', 'BANK OF CHINA'),
    ('2318 HK', 'PING AN'), ('883 HK', 'CNOOC'), ('16 HK', 'SHK PPT'), ('1810 HK', 'XIAOMI-W'),
    ('3690 HK', 'MEITUAN-W'), ('2020 HK', 'ANTA SPORTS'), ('27 HK', 'GALAXY ENT'),
    ('1113 HK', 'CK ASSET'), ('2 HK', 'CLP HOLDINGS'), ('386 HK', 'SINOPEC'),
    ('9618 HK', 'JD-SW'), ('1211 HK', 'BYD'), ('6862 HK', 'HAIDILAO'), ('1928 HK', 'SANDS CHINA'),
]


def _bday(base, n):
    d, step = base, (1 if n >= 0 else -1)
    for _ in range(abs(n)):
        d += dt.timedelta(days=step)
        while d.weekday() >= 5:
            d += dt.timedelta(days=step)
    return d


def _eco_row(ticker, label, idx):
    offset = (idx * 5) % 88 + 2
    when = _bday(_TODAY_FIXTURE, offset)
    survey = round(0.1 + (idx % 17) * 0.13, 2)
    prior = round(survey - 0.05 * ((idx % 5) - 2), 2)
    return [('ECO_FUTURE_RELEASE_DATE', when.strftime('%Y-%m-%d')),
            ('ECO_FUTURE_RELEASE_TIME', '08:30' if idx % 3 == 0 else ('10:00' if idx % 3 == 1 else '14:00')),
            ('BN_SURVEY_MEDIAN', survey), ('PX_LAST', prior),
            ('RELEVANCE_VALUE', 99 - (idx % 40)), ('NAME', label.upper())]


def _cb_row(ticker, label, idx):
    when = _bday(_TODAY_FIXTURE, (idx * 9) % 80 + 3)
    rate = round(0.25 + idx * 0.55, 2)
    return [('ECO_FUTURE_RELEASE_DATE', when.strftime('%Y-%m-%d')),
            ('ECO_FUTURE_RELEASE_TIME', '14:00' if idx % 2 else '09:00'),
            ('BN_SURVEY_MEDIAN', rate), ('PX_LAST', rate),
            ('RELEVANCE_VALUE', 99), ('NAME', label.upper())]


def _earn_row(ticker, name, idx):
    when = _bday(_TODAY_FIXTURE, (idx * 4) % 86 + 1)
    call = _bday(when, 0)
    # ERN_ANN_DT_AND_PER returns history AND estimated forward dates together.
    bulk = fake_array('ERN_ANN_DT_AND_PER', [
        fake_complex('row', [('Announcement Date', d.strftime('%Y-%m-%d')),
                           ('Financial Period', 'FY26')])
        for d in (_bday(_TODAY_FIXTURE, -120), _bday(_TODAY_FIXTURE, -30), when)])
    # every 7th name has no scalar report date - only the bulk field, so the
    # fallback path is exercised on every fixture run
    scalar_missing = (idx % 7 == 3)
    return [('NAME', name),
            ('EXPECTED_REPORT_DT', None if scalar_missing else when.strftime('%Y-%m-%d')),
            ('EXPECTED_REPORT_TIME', None if scalar_missing else
             ['Bef-mkt', '12:00', 'Aft-mkt', '07:00'][idx % 4]),
            ('EARNINGS_CONF_CALL_DT', call.strftime('%Y-%m-%d') if idx % 5 else None),
            ('EARNINGS_CONF_CALL_TIME', '16:30' if idx % 5 else None),
            ('ERN_ANN_DT_AND_PER', bulk)]


def _fake_fields(security, wanted):
    """Everything the fixture knows, before the requested-field filter."""
    if security == HSI_INDEX:
        members = [fake_complex('INDX_MEMBERS',
                              [('Member Ticker and Exchange Code', t)]) for t, _ in HSI_MEMBERS]
        return [('INDX_MEMBERS', fake_array('INDX_MEMBERS', members))]
    for i, (t, label) in enumerate(US_ECO_TICKERS):
        if t == security:
            return _eco_row(t, label, i)
    for i, (_c, _b, t, label, _tz) in enumerate(CENTRAL_BANKS):
        if t == security:
            return _cb_row(t, label, i)
    base = security[:-len(' Equity')] if security.upper().endswith(' EQUITY') else security
    for i, (t, name) in enumerate(HSI_MEMBERS):
        if t == base:
            return _earn_row(t, name, i)
    return []


class FakeSession:
    def __init__(self, options=None):
        self._queue = []

    def start(self):
        return True

    def stop(self):
        pass

    def openService(self, name):
        return name in ('//blp/refdata', '//blp/apiflds', '//blp/instruments')

    def getService(self, name):
        return FakeService(name)

    def sendRequest(self, request):
        wanted = set(request.fields)
        sec_elements = []
        for sec in request.securities:
            if sec in UNKNOWN_SECURITY:
                sec_elements.append(fake_complex('securityData', [
                    ('security', sec),
                    ('securityError', fake_complex('securityError', [('message', 'Unknown/Invalid Security')])),
                ]))
                continue
            pairs = [(k, v) for k, v in _fake_fields(sec, wanted)
                     if k in wanted and v is not None]
            fd = fake_complex('fieldData', pairs)
            kids = [('security', sec), ('fieldData', fd)]
            bad = sorted(wanted & UNKNOWN_FIELDS)
            if bad:
                kids.append(('fieldExceptions', fake_array('fieldExceptions', [
                    fake_complex('fieldExceptions', [
                        ('fieldId', f),
                        ('errorInfo', fake_complex('errorInfo', [
                            ('subcategory', 'BAD_FLD'), ('message', 'Invalid Field')])),
                    ]) for f in bad])))
            sec_elements.append(fake_complex('securityData', kids))
        root = FakeElement('ReferenceDataResponse',
                       children=[('securityData', fake_array('securityData', sec_elements))])
        self._queue = [FakeEvent([FakeMessage(root)], FakeEvent.RESPONSE)]

    def nextEvent(self, timeout=0):
        return self._queue.pop(0) if self._queue else FakeEvent([], FakeEvent.TIMEOUT)


class FixtureAPI:
    """Stands in for the blpapi module itself."""
    SessionOptions = FakeSessionOptions
    Session = FakeSession
    Event = FakeEvent


# ================================================================ TESTS ===
# python daily_email.py --test     (no terminal, no Outlook, no network)
TEST_TODAY = dt.date(2026, 9, 4)
FAILURES = []


def check(name, cond, detail=''):
    if cond:
        print(f'  ok    {name}')
    else:
        print(f'  FAIL  {name}  {detail}')
        FAILURES.append(name)


# ------------------------------------------------------------ conversions ---
def test_dates_and_times():
    print('dates and times')
    check('ISO date', as_date('2026-09-04') == dt.date(2026, 9, 4))
    check('datetime keeps the date', as_date(dt.datetime(2026, 9, 4, 8, 30)) == dt.date(2026, 9, 4))
    check('Excel serial 46269 -> 2026-09-04', as_date(46269) == dt.date(2026, 9, 4),
          str(as_date(46269)))
    check('day/month/year', as_date('04/09/2026') == dt.date(2026, 9, 4))
    check('blank is None', as_date('') is None and as_date(None) is None)
    check('junk is None', as_date('n.a.') is None)
    check('clock text', as_time('8:30') == '08:30')
    check('Excel fraction 0.354166 -> 08:30', as_time(0.3541666667) == '08:30', as_time(0.3541666667))
    check('text time survives', as_time('Bef-mkt') == 'Bef-mkt')
    check('midnight datetime is not a time', as_time(dt.datetime(2026, 9, 4, 0, 0)) is None)
    check('time object', as_time(dt.time(14, 0)) == '14:00')


def test_sort_and_format():
    print('sorting and formatting')
    keys = [time_key('Bef-mkt'), time_key('08:30'), time_key('16:00'),
            time_key('Aft-mkt'), time_key(None)]
    check('before < clock < after < unknown', keys == sorted(keys), str(keys))
    check('thousands', fmt_num(1234567.0) == '1,234,567', fmt_num(1234567.0))
    check('trailing zeros trimmed', fmt_num(0.700) == '0.7', fmt_num(0.7))
    check('rate to 2dp', fmt_num(5.5, 2) == '5.50')
    check('None is a dash', fmt_num(None) == '–')
    check('zero prints as 0', fmt_num(0.0) == '0', fmt_num(0.0))
    check('negative survives', fmt_num(-0.25) == '-0.25', fmt_num(-0.25))
    check('today tag', day_tag(TEST_TODAY, TEST_TODAY) == 'TODAY')
    check('tomorrow tag', day_tag(TEST_TODAY + dt.timedelta(days=1), TEST_TODAY) == 'tomorrow')
    check('n-day tag', day_tag(TEST_TODAY + dt.timedelta(days=9), TEST_TODAY) == 'in 9 days')


def test_exclusions():
    print('exclusion list')
    check('exact name excluded', is_excluded('Building Permits'))
    check('case and spacing ignored', is_excluded('  building   permits '))
    check('kept name not excluded', not is_excluded('Change in Nonfarm Payrolls'))
    check('near-miss kept', not is_excluded('Retail Sales Advance MoM'))
    check('every VBA entry loaded', len(_EXCLUDE) == len(set(EXCLUDE_EVENTS)),
          f'{len(_EXCLUDE)} vs {len(EXCLUDE_EVENTS)}')
    labels = {lbl for _, lbl in US_ECO_TICKERS}
    unmatched = [e for e in EXCLUDE_EVENTS if e not in labels]
    check('every exclusion matches a configured label', not unmatched, str(unmatched))


def test_pick_date_skips_stale():
    print('date selection')
    rec = {'ECO_FUTURE_RELEASE_DATE': '2026-08-01', 'ECO_RELEASE_DT': '2026-09-20'}
    when, _ = pick_date(rec, ['ECO_FUTURE_RELEASE_DATE', 'ECO_RELEASE_DT'], [], TEST_TODAY)
    check('a past date falls through to the next candidate', when == dt.date(2026, 9, 20), str(when))
    rec = {'ECO_FUTURE_RELEASE_DATE': dt.datetime(2026, 9, 20, 8, 30)}
    when, t = pick_date(rec, ['ECO_FUTURE_RELEASE_DATE'], ['ECO_FUTURE_RELEASE_TIME'], TEST_TODAY)
    check('a datetime supplies the time', (when, t) == (dt.date(2026, 9, 20), '08:30'), f'{when} {t}')
    rec = {'ECO_FUTURE_RELEASE_DATE': '2026-09-20', 'ECO_FUTURE_RELEASE_TIME': '14:00'}
    _, t = pick_date(rec, ['ECO_FUTURE_RELEASE_DATE'], ['ECO_FUTURE_RELEASE_TIME'], TEST_TODAY)
    check('an explicit time field wins', t == '14:00')
    check('no usable date -> None', pick_date({}, ['A'], ['B'], TEST_TODAY) == (None, None))


def test_timezone_shift():
    print('time-zone conversion')
    d, t = shift_tz(dt.date(2026, 9, 4), '08:30', 'America/New_York', 'Asia/Hong_Kong')
    check('NY 08:30 -> HK 20:30 same day', (d, t) == (dt.date(2026, 9, 4), '20:30'), f'{d} {t}')
    d, t = shift_tz(dt.date(2026, 9, 4), '16:00', 'America/New_York', 'Asia/Hong_Kong')
    check('NY 16:00 rolls the date forward', (d, t) == (dt.date(2026, 9, 5), '04:00'), f'{d} {t}')
    d, t = shift_tz(dt.date(2026, 9, 4), 'Bef-mkt', 'America/New_York', 'Asia/Hong_Kong')
    check('text time passes through untouched', (d, t) == (dt.date(2026, 9, 4), 'Bef-mkt'))


# ------------------------------------------------------------- API path ----
def build(today=TEST_TODAY):
    fixture_configure(today)
    sections, bbg = build_sections_blpapi(today, blpapi_module=FixtureAPI)
    return sections, bbg


def test_api_path():
    print('Bloomberg path (fake blpapi)')
    sections, bbg = build()
    check('three sections', set(sections) == {'central_banks', 'us_eco', 'hsi_earnings'})

    eco = sections['us_eco']
    check('us_eco has rows', len(eco.rows) > 10, str(len(eco.rows)))
    check('nothing excluded survived', not [r for r in eco.rows if is_excluded(r['event'])])
    check('excluded rows were counted', eco.n_excluded > 0, str(eco.n_excluded))
    dates = [r['date'] for r in eco.rows]
    check('sorted by date', dates == sorted(dates))
    check('nothing in the past', all(d >= TEST_TODAY for d in dates))
    check('nothing past the horizon',
          all(d <= TEST_TODAY + dt.timedelta(days=LOOKAHEAD_DAYS) for d in dates))
    same_day = [r for r in eco.rows if r['date'] == dates[0]]
    check('same-day rows ordered by time',
          [time_key(r['time']) for r in same_day] == sorted(time_key(r['time']) for r in same_day))

    cb = sections['central_banks']
    check('central banks have rows', len(cb.rows) >= 5, str(len(cb.rows)))
    check('each names a bank', all(r['bank'] for r in cb.rows))
    check('countries look right', {r['country'] for r in cb.rows} <= {c for c, *_ in CENTRAL_BANKS})

    earn = sections['hsi_earnings']
    check('earnings have rows', len(earn.rows) >= 5, str(len(earn.rows)))
    check('tickers carry no Equity suffix', not [r for r in earn.rows if 'Equity' in r['ticker']])
    check('every row has a company name', all(r['name'] for r in earn.rows))
    check('sorted by date', [r['date'] for r in earn.rows] == sorted(r['date'] for r in earn.rows))
    check('some rows carry a conference call', any(r['call_date'] for r in earn.rows))
    check('the bulk-field fallback filled some dates', any(r['estimated'] for r in earn.rows))
    check('bulk-sourced dates are still in the window',
          all(TEST_TODAY <= r['date'] <= TEST_TODAY + dt.timedelta(days=LOOKAHEAD_DAYS)
              for r in earn.rows if r['estimated']))
    check('the fallback is disclosed in the notes',
          any('ERN_ANN_DT_AND_PER' in n for n in earn.notes), str(earn.notes))


def test_bulk_fields():
    print('bulk fields')
    rec = {'INDX_MEMBERS': [{'Member Ticker and Exchange Code': '700 HK'}, {'Ticker': '939 HK'}]}
    check('substring match handles both element names',
          bulk_field(rec, 'INDX_MEMBERS', 'ticker') == ['700 HK', '939 HK'],
          str(bulk_field(rec, 'INDX_MEMBERS', 'ticker')))
    odd = {'INDX_MEMBERS': [{'Something Unexpected': '1299 HK'}]}
    check('an unknown element name falls back to the first value',
          bulk_field(odd, 'INDX_MEMBERS', 'ticker') == ['1299 HK'])
    check('a missing bulk field is empty, not an error', bulk_field({}, 'X', 'y') == [])
    check('a scalar in place of a bulk field is ignored',
          bulk_field({'X': ['not a dict']}, 'X', 'y') == [])
    hist = {EARN_BULK_DATE: [{'Announcement Date': '2025-03-20'},
                                {'Announcement Date': '2026-11-12'},
                                {'Announcement Date': '2026-09-30'}]}
    check('the earliest FUTURE announcement wins',
          next_announced(hist, TEST_TODAY) == dt.date(2026, 9, 30),
          str(next_announced(hist, TEST_TODAY)))
    past = {EARN_BULK_DATE: [{'Announcement Date': '2025-03-20'}]}
    check('history alone yields nothing', next_announced(past, TEST_TODAY) is None)


def test_bad_security_and_field():
    print('error handling')
    fixture_configure(TEST_TODAY)
    bbg = Bloomberg(blpapi_module=FixtureAPI).connect()
    got = bbg.ref(['XXXX FAKE Index', 'FDTR Index'], ['PX_LAST', 'CONF_CALL_DT'])
    check('a dead ticker is recorded, not raised', 'XXXX FAKE Index' in bbg.bad_securities)
    check('the dead ticker returns no row', 'XXXX FAKE Index' not in got)
    check('the good ticker still comes back', 'FDTR Index' in got)
    check('a rejected mnemonic is recorded', 'CONF_CALL_DT' in bbg.field_errors)
    quiet = field_notes(bbg, {'call_date': ['EARNINGS_CONF_CALL_DT', 'CONF_CALL_DT']},
                           [{'EARNINGS_CONF_CALL_DT': '2026-09-20'}])
    check('a covered fallback produces no note', quiet == [], str(quiet))
    loud = field_notes(bbg, {'call_date': ['CONF_CALL_DT']}, [{}])
    check('a role with no data at all is flagged', len(loud) == 1 and 'call_date' in loud[0], str(loud))
    bbg.close()


def test_chunking():
    print('request chunking')
    fixture_configure(TEST_TODAY)
    bbg = Bloomberg(blpapi_module=FixtureAPI).connect()
    tickers = [t for t, _ in US_ECO_TICKERS]
    check('universe is larger than one chunk', len(tickers) > CHUNK, str(len(tickers)))
    got = bbg.ref(tickers, ['PX_LAST'])
    check('every ticker comes back across chunks', len(got) == len(tickers), f'{len(got)}/{len(tickers)}')
    bbg.close()


# ------------------------------------------------------------- rendering ---
def test_rendering():
    print('rendering')
    sections, _ = build()
    now = dt.datetime(2026, 9, 4, 7, 30)
    for cfg in EMAILS:
        subject, frag = render_email(cfg, sections, TEST_TODAY, now, 'test')
        check(f'[{cfg["key"]}] subject filled in', '{' not in subject, subject)
        check(f'[{cfg["key"]}] tags balance',
              frag.count('<table') == frag.count('</table>') and frag.count('<tr') == frag.count('</tr>'))
        check(f'[{cfg["key"]}] inline styles only (Outlook strips <style>)', '<style' not in frag)
        check(f'[{cfg["key"]}] recipients not in the body', 'nomura.com' not in frag)
        check(f'[{cfg["key"]}] greeting present', GREETING in frag)
    subject, frag = render_email(EMAILS[1], sections, TEST_TODAY, now, 'test')
    check('earnings count reaches the subject', str(len(sections['hsi_earnings'].rows)) in subject, subject)

    row = dict(date=TEST_TODAY, time='08:30', country='US', bank='', event='Fish & Chips <PMI>',
               bbg_name=None, survey=1.0, prior=2.0, relevance=99, ticker='X Index')
    frag = render_calendar(Section('us_eco', [row]), TEST_TODAY)
    check('markup in an event name is escaped', '&lt;PMI&gt;' in frag and '&amp;' in frag)

    empty = render_email(EMAILS[0], {}, TEST_TODAY, now, 'test')[1]
    check('no data still renders an email', 'Nothing returned' in empty)

    doc = full_document('s', frag)
    check('full document wraps the fragment', doc.startswith('<html>') and doc.rstrip().endswith('</html>'))
    spliced = splice_into_signature('<html><body><p>-- <br>Sig</p></body></html>', '<b>BODY</b>')
    check('body goes above the signature', spliced.index('BODY') < spliced.index('Sig'))
    check('no <body> means the fragment is still kept',
          'BODY' in splice_into_signature('', '<b>BODY</b>'))


def test_eml_written():
    print('files written')
    sections, _ = build()
    now = dt.datetime(2026, 9, 4, 7, 30)
    subject, frag = render_email(EMAILS[0], sections, TEST_TODAY, now, 'test')
    html_path, eml_path = write_files('unittest', subject, frag, TEST_TODAY, now)
    raw = open(eml_path, encoding='utf-8').read()
    check('.eml opens in compose mode', 'X-Unsent: 1' in raw)
    check('.eml is addressed', 'nomura.com' in raw)
    check('.eml carries the HTML part', 'text/html' in raw)
    check('.html written', os.path.getsize(html_path) > 500)
    for p in (html_path, eml_path):
        os.remove(p)


# ----------------------------------------------------------- excel path ----
def test_excel_path():
    print('Excel path')
    try:
        import openpyxl
    except ImportError:
        print('  skip  openpyxl not installed')
        return
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'BQL'
    # A BQL calendar block, exactly the shape the add-in spills: header then rows.
    ws.append(['ID', 'RELEASE_DATE', 'RELEASE_TIME', 'EVENT_NAME', 'SURVEY_MEDIAN', 'PRIOR'])
    rows = [
        ('US Country', dt.datetime(2026, 9, 8), 0.3541666667, 'Change in Nonfarm Payrolls', 75, 22),
        ('US Country', dt.datetime(2026, 9, 8), 0.3541666667, 'Building Permits', 1.4, 1.5),   # excluded
        ('US Country', dt.datetime(2026, 9, 10), 0.3541666667, 'CPI MoM', 0.2, 0.3),
        ('US Country', dt.datetime(2026, 8, 1), 0.3541666667, 'CPI YoY', 2.9, 3.0),            # past
        ('US Country', dt.datetime(2027, 6, 1), 0.3541666667, 'GDP Annualized QoQ', 2.0, 1.8), # beyond
    ]
    for r in rows:
        ws.append(list(r))
    ws2 = wb.create_sheet('EARN')
    ws2.append(['ID', 'NAME', 'EXPECTED_REPORT_DT', 'EXPECTED_REPORT_TIME',
                'EARNINGS_CONF_CALL_DT', 'EARNINGS_CONF_CALL_TIME'])
    ws2.append(['700 HK Equity', 'TENCENT', dt.datetime(2026, 11, 12), 'Aft-mkt',
                dt.datetime(2026, 11, 12), 0.6875])
    ws2.append(['939 HK Equity', 'CCB', dt.datetime(2026, 10, 29), 'Bef-mkt', None, None])
    ws2.append(['1 HK Equity', 'CKH', dt.datetime(2025, 3, 20), 'Bef-mkt', None, None])   # past
    path = os.path.join(HERE, '_tmp_excel_test.xlsx')
    wb.save(path)
    try:
        sections = read_excel_sections(path, TEST_TODAY)
        eco = sections.get('us_eco')
        check('calendar block found', eco is not None)
        events = [r['event'] for r in eco.rows]
        check('two rows kept', len(eco.rows) == 2, str(events))
        check('excluded name dropped', 'Building Permits' not in events)
        check('past row dropped', 'CPI YoY' not in events)
        check('beyond-horizon row dropped', 'GDP Annualized QoQ' not in events)
        check('Excel time fraction decoded', eco.rows[0]['time'] == '08:30', str(eco.rows[0]['time']))
        check('survey read', eco.rows[0]['survey'] == 75, str(eco.rows[0]['survey']))
        earn = sections.get('hsi_earnings')
        check('earnings block found', earn is not None)
        check('two names kept', len(earn.rows) == 2, str([r['ticker'] for r in earn.rows]))
        check('earliest first', earn.rows[0]['ticker'] == '939 HK', str(earn.rows[0]))
        check('Equity suffix stripped', all(' Equity' not in r['ticker'] for r in earn.rows))
        check('call time decoded', earn.rows[1]['call_time'] == '16:30', str(earn.rows[1]['call_time']))
        subject, frag = render_email(EMAILS[1], sections, TEST_TODAY, dt.datetime(2026, 9, 4), 'xl')
        check('an Excel-sourced email renders', 'TENCENT' in frag)
    finally:
        os.remove(path)


def test_column_mapping():
    print('column mapping')
    m = map_columns(['ID', 'RELEASE_DATE', 'RELEASE_TIME', 'EVENT_NAME', 'SURVEY_MEDIAN', 'PRIOR'],
                       CAL_SPEC)
    check('date column', m['date'] == 1, str(m))
    check('time not confused with date', m['time'] == 2, str(m))
    check('event column', m['event'] == 3, str(m))
    check('survey column', m['survey'] == 4, str(m))
    check('prior column', m['prior'] == 5, str(m))
    m = map_columns(['#DATES', '#TIMES', '#EVENTS', '#SURVEY', '#PRIOR'], CAL_SPEC)
    check('BQL # headers still map', m.get('event') == 2 and m.get('survey') == 3, str(m))
    check('calendar header classified', classify_header(['RELEASE_DATE', 'EVENT_NAME']) == 'calendar')
    check('earnings header classified',
          classify_header(['EXPECTED_REPORT_DT', 'NAME']) == 'hsi_earnings')
    check('an unrelated header is ignored', classify_header(['Price', 'Qty']) is None)


def test_config_sanity():
    print('config')
    tickers = [t for t, _ in US_ECO_TICKERS]
    check('no duplicate US tickers', len(tickers) == len(set(tickers)))
    cb = [t for _, _, t, _, _ in CENTRAL_BANKS]
    check('no duplicate CB tickers', len(cb) == len(set(cb)))
    check('every ticker carries a yellow key',
          all(t.split()[-1] in ('Index', 'Equity', 'Curncy', 'Comdty') for t in tickers + cb))
    keys = [k for cfg in EMAILS for k in cfg['sections']]
    check('every email section is produced', set(keys) <= set(SECTION_TITLES))
    check('every section has a title', set(SECTION_TITLES) >= set(keys))
    check('recipients are the two names asked for',
          RECIPIENTS.count('@') == 2 and 'oscar.chan1@nomura.com' in RECIPIENTS)


def run_tests():
    for fn in [test_dates_and_times, test_sort_and_format, test_exclusions, test_pick_date_skips_stale,
               test_timezone_shift, test_api_path, test_bulk_fields, test_bad_security_and_field,
               test_chunking,
               test_rendering, test_eml_written, test_excel_path, test_column_mapping,
               test_config_sanity]:
        fn()
    print()
    if FAILURES:
        print(f'{len(FAILURES)} FAILED: ' + ', '.join(FAILURES))
        return 1
    print('all tests passed')
    return 0


# ============================================================== JUPYTER ===
# Six functions, meant to be called from a notebook cell.  None of them raises:
# a problem is printed in plain language, with Bloomberg's own wording kept, and
# the call returns None.  Nothing here is hidden - what failed is always named.


def _say(problem, hint=''):
    print(f'\n  Could not finish: {problem}')
    if hint:
        for line in hint.strip().split('\n'):
            print(f'  {line.strip()}')
    print()


def _explain(exc):
    """Turn the usual failures into a sentence that says what to do next."""
    text = str(exc)
    if isinstance(exc, ImportError) or 'blpapi' in text.lower() and 'install' in text.lower():
        missing = 'blpapi' if 'blpapi' in text else ('pywin32' if 'win32' in text else
                                                     ('openpyxl' if 'openpyxl' in text else text))
        return (f'{missing} is not available in this kernel.',
                'Run demo() to see the emails without Bloomberg, or start the notebook\n'
                'from the environment that has blpapi installed.')
    if 'session' in text.lower() or 'refdata' in text.lower():
        return (text, 'Is the Bloomberg terminal running and logged in on this machine?')
    if isinstance(exc, FileNotFoundError):
        return (f'file not found: {text}', 'Check the path, and use forward slashes.')
    return (text, '')


def _show(fragment, subject):
    """Draw the email inside the notebook, or say where the file is."""
    if IN_JUPYTER:
        try:
            from IPython.display import display, HTML
            display(HTML(f'<div style="margin:14px 0 4px 0;{FONT}font-size:11pt;color:{MUTED};">'
                         f'<b style="color:{NAVY};">{esc(subject)}</b><br>'
                         f'To: {esc(RECIPIENTS)}</div>{fragment}'))
            return True
        except Exception:
            pass
    return False


def _deliver(sections, today, source_label, to_outlook, display_them):
    out = {}
    for cfg in EMAILS:
        subject, fragment = render_email(cfg, sections, today, dt.datetime.now(), source_label)
        html_path, eml_path = write_files(cfg['key'], subject, fragment, today, dt.datetime.now())
        status = 'not sent'
        if to_outlook:
            try:
                _ok, status = outlook_draft(subject, fragment, display_them)
            except Exception as e:
                status = f'Outlook step failed: {e}'
        print(f'{cfg["key"]:9s} {status:26s} {os.path.basename(html_path)}')
        if not _show(fragment, subject):
            print(f'          open this to read it: {html_path}')
        out[cfg['key']] = dict(subject=subject, html=fragment,
                               html_path=html_path, eml_path=eml_path)
    return out


def demo(asof=None):
    """Both emails from sample data.  No Bloomberg, no Outlook, works anywhere."""
    try:
        today = dt.date.fromisoformat(asof) if asof else dt.date.today()
        fixture_configure(today)
        print('Sample data - this is the layout, not real numbers.\n')
        sections, _ = build_sections_blpapi(today, blpapi_module=FixtureAPI)
        print('')
        return _deliver(sections, today, 'SAMPLE DATA - not Bloomberg', False, False)
    except Exception as e:
        _say(*_explain(e))


def test():
    """The self-tests.  Prints a line per check and a verdict at the end."""
    try:
        return run_tests() == 0
    except Exception as e:
        _say(*_explain(e))


def probe(asof=None):
    """Check every field and ticker against the live terminal.  Do this first.

    Writes out/probe_YYYY-MM-DD.txt.  Send that file back and the config gets
    corrected in one pass."""
    try:
        today = dt.date.fromisoformat(asof) if asof else dt.date.today()
        _probe_report(today)
    except Exception as e:
        _say(*_explain(e))


def run(excel=None, display=False, outlook=True, asof=None):
    """The real thing: two drafts into Outlook, and shown in the notebook.

        run()                          from Bloomberg
        run(excel='C:/path/cal.xlsx')  from the workbook instead, no field risk
        run(outlook=False)             build them, do not touch Outlook
    """
    try:
        today = dt.date.fromisoformat(asof) if asof else dt.date.today()
        if excel:
            print('Reading the workbook ...')
            sections = read_excel_sections(excel, today)
            label = f'Bloomberg BQL via {os.path.basename(excel)}'
        else:
            sections = build_sections_blpapi(today)[0]
            label = 'Bloomberg Desktop API (//blp/refdata)'
        print('')
        return _deliver(sections, today, label, outlook, display)
    except Exception as e:
        _say(*_explain(e))


def compare(excel, asof=None):
    """Row-by-row diff: what Bloomberg gives against what the workbook holds."""
    try:
        today = dt.date.fromisoformat(asof) if asof else dt.date.today()
        api = build_sections_blpapi(today)[0]
        print('Reading the workbook ...')
        _parity_report(api, read_excel_sections(excel, today))
    except Exception as e:
        _say(*_explain(e))


MENU = """
  Daily calendar email - ready.  Call one of these in the next cell:

      demo()                          both emails from sample data, no Bloomberg
      test()                          run the self-tests
      probe()                         check fields and tickers on the terminal
      run()                           the real thing: two Outlook drafts
      run(excel='C:/path/cal.xlsx')   from the workbook instead
      compare('C:/path/cal.xlsx')     Bloomberg against the workbook

  Emails are drawn in the notebook and saved to the out folder.
  Nothing is ever sent - both drafts wait in Outlook for you to press Send.
"""


# ================================================================= MAIN ===
LAUNCHER = """@echo off
REM Double-click on the Bloomberg terminal PC, with Outlook open.
title Daily calendar email
cd /d "%~dp0"
where python >nul 2>nul && (python daily_email.py %* & goto done)
where py     >nul 2>nul && (py -3 daily_email.py %* & goto done)
echo   No Python on the PATH - open the prompt the desk notebooks use.
:done
echo.
pause
"""


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--demo', action='store_true',
                    help='sample data, no Bloomberg and no Outlook - see the layout')
    ap.add_argument('--test', action='store_true', help='run the self-tests and exit')
    ap.add_argument('--probe', action='store_true',
                    help='verify every field and ticker against the terminal')
    ap.add_argument('--display', action='store_true',
                    help='pop the drafts open instead of saving them')
    ap.add_argument('--no-outlook', action='store_true', help='write the files only')
    ap.add_argument('--xlsx', help='workbook holding the BQL output, for --excel / --parity')
    ap.add_argument('--excel', action='store_true',
                    help='read --xlsx instead of the API (no mnemonic risk)')
    ap.add_argument('--parity', action='store_true', help='diff the API against --xlsx')
    ap.add_argument('--bql', action='store_true', help='run the BQL strings - BQuant only')
    ap.add_argument('--asof', help='pretend it is another day, YYYY-MM-DD')
    ap.add_argument('--make-launcher', action='store_true',
                    help='write RUN_DAILY_EMAIL.bat next to this script and exit')
    args = ap.parse_args(argv)

    if args.test:
        return run_tests()

    if args.make_launcher:
        path = os.path.join(HERE, 'RUN_DAILY_EMAIL.bat')
        with open(path, 'w', encoding='utf-8') as fh:
            fh.write(LAUNCHER)
        print(f'Wrote {path} - double-click it to run the real thing.')
        return 0

    today = dt.date.fromisoformat(args.asof) if args.asof else dt.date.today()
    now = dt.datetime.now()
    fake = None
    if args.demo:
        fixture_configure(today)
        fake = FixtureAPI

    if args.probe:
        _probe_report(today, blpapi_module=fake)
        return 0

    if args.parity:
        if not args.xlsx:
            ap.error('--parity needs --xlsx')
        api_sections, _ = build_sections_blpapi(today, blpapi_module=fake)
        print('Reading workbook ...')
        _parity_report(api_sections, read_excel_sections(args.xlsx, today))
        return 0

    if args.excel:
        if not args.xlsx:
            ap.error('--excel needs --xlsx')
        print('Reading workbook ...')
        sections = read_excel_sections(args.xlsx, today)
        source_label = f'Bloomberg BQL via {os.path.basename(args.xlsx)}'
    elif args.bql:
        sections = build_sections_bql(today)
        source_label = 'Bloomberg BQL (BQuant)'
    else:
        sections, _ = build_sections_blpapi(today, blpapi_module=fake)
        source_label = ('SAMPLE DATA - not Bloomberg' if fake
                        else 'Bloomberg Desktop API (//blp/refdata)')

    print('')
    for cfg in EMAILS:
        subject, fragment = render_email(cfg, sections, today, now, source_label)
        html_path, eml_path = write_files(cfg['key'], subject, fragment, today, now)
        status = 'skipped (--demo)' if args.demo else 'not sent (--no-outlook)'
        if not args.no_outlook and not args.demo:
            ok, status = outlook_draft(subject, fragment, args.display)
        print(f'Draft [{cfg["key"]}] "{subject}"\n   Outlook: {status}\n'
              f'   files:   {html_path}\n            {eml_path}')
    return 0


if __name__ == '__main__':
    if IN_JUPYTER:
        print(MENU)                     # a notebook gets the menu, never argparse
    else:
        sys.exit(main())
else:
    if IN_JUPYTER:
        print(MENU)
