#!/usr/bin/env python3
"""Prove every formula points at the column it CLAIMS to.

The screener once ranked 510 of 511 rows on the public-subscription column
while its header said "subsector": a loop variable had been reused, so the gate
formula for row 5 was right and every row after it was wrong. Parity testing
missed it because the top comp still won on the sector fallback.

This audit reads the BUILT workbook and, for every generated formula, resolves
each `Database!$X$n` back to a field name and compares it against what that cell
is supposed to be pulling:

  1. comp table   — header text -> COMP_PULL[header] -> expected Database column
  2. Calc engine  — each scoring helper column -> its documented input field
  3. row drift    — a formula on row r must reference row r (not a frozen row)
  4. screener card — the target cells the engine reads must hold the labels the
     card shows next to them

Exit 1 on any mismatch. Cheap enough to run on every build.
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import build_xlsx as B

ROOT = Path(__file__).resolve().parent.parent
WB = ROOT / "out" / "HK_IPO_Database_v1.xlsx"
DBREF = re.compile(r"Database!\$?([A-Z]{1,2})\$?(\d+)")

# Calc scoring columns -> the Database field each one is defined to read.
CALC_EXPECT = {
    B.H_GATE: "subsector",
    B.H_SEC: "sector",
    B.H_SIZE: "deal_size_hkdm",
    B.H_PROF: "profitable_at_ipo",
    B.H_AH: "is_h_share",
    B.H_REC: "ipo_date",
    B.H_CS: "cornerstone_keys",
    B.H_PE: "pe_ipo",
    B.H_DEM: "oversub_public_mult",
}
# The screener card: engine cell -> the label that must sit beside it.
# the Screener's live panel sits at a COMPUTED row, so it cannot be pinned by
# address — assert the labels exist at all. Two rows were silently overwritten
# by blocks written underneath them (force-include, then H P/S now).
LIVE_PANEL = ["A price now", "A P/E (TTM)", "H P/E now", "H P/S now",
              "A-line company mkt cap",
              "A premium vs H"]

CARD_EXPECT = {"A7": "Sector", "A8": "Subsector", "A9": "Size",
               "A13": "Target P/E", "A14": "Public sub",
               # a control written into the comp table's band row (19) is
               # silently overwritten — assert where this one actually lives
               "C17": "Force-include"}


def field_of(letter):
    idx = openpyxl.utils.column_index_from_string(letter)
    if 1 <= idx <= len(B.DB_COLS):
        return B.DB_COLS[idx - 1][2]
    return f"<col {letter} out of range>"


def main():
    if not WB.exists():
        sys.exit(f"no workbook at {WB} — run build_xlsx.py first")
    wb = openpyxl.load_workbook(WB)
    ws, calc = wb["Screener"], wb["Calc (scoring engine)"]
    bad, checked = [], 0

    # ---- 1. comp table: every pulled column matches its own header ----------
    hdr_row = 20
    first = 21
    headers = {get_column_letter(c): ws.cell(hdr_row, c).value
               for c in range(1, len(B.COMP_COLS) + 1)}
    for col, head in headers.items():
        want = B.COMP_PULL.get(head)
        if not want:
            continue
        f = ws[f"{col}{first}"].value
        if not isinstance(f, str):
            continue
        refs = {field_of(L) for L, _r in DBREF.findall(f)}
        checked += 1
        # the cell may legitimately touch a second column (the n/m and N/A
        # guards read profitable_at_ipo / a_share_code / code); the column it
        # PULLS must still be present
        if want not in refs:
            bad.append(f"comp col {col} header {head!r} should pull {want!r} "
                       f"but references {sorted(refs)}")

    # ---- 2. Calc engine: each helper reads its documented field -------------
    scr = wb["Screener"]
    colA = [str(scr.cell(r, 1).value or "") for r in range(1, scr.max_row + 1)]
    for lab in LIVE_PANEL:
        if not any(v.startswith(lab) for v in colA):
            bad += 1
            print(f"  FAIL Screener: live-panel row {lab!r} is missing — "
                  f"a block written below it overwrote the row")

    for hcol, want in CALC_EXPECT.items():
        seen = set()
        for r in (B.DB_R0, B.DB_R0 + 1, B.DB_R0 + 7, B.DB_R0 + 200):
            f = calc[f"{hcol}{r}"].value
            if not isinstance(f, str):
                continue
            checked += 1
            refs = {field_of(L) for L, _rr in DBREF.findall(f)}
            seen |= refs
            if want not in refs:
                bad.append(f"Calc {hcol}{r} should read {want!r} but references "
                           f"{sorted(refs)}")
            # 3. row drift: a per-deal formula must reference its OWN row
            rows = {int(rr) for _L, rr in DBREF.findall(f)}
            if rows and rows != {r}:
                bad.append(f"Calc {hcol}{r} references rows {sorted(rows)} — "
                           f"expected only {r}")
        if len(seen) > 2:
            bad.append(f"Calc {hcol} reads too many fields across rows: {sorted(seen)}")

    # ---- 4. screener card labels sit beside the cells the engine reads ------
    for cell, label in CARD_EXPECT.items():
        got = str(ws[cell].value or "")
        checked += 1
        if not got.startswith(label):
            bad.append(f"card {cell} should be labelled {label!r}, found {got!r}")

    # ---- 5. no formula may point outside the data block ---------------------
    for sheet in (ws, calc):
        for row in sheet.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    for L, rr in DBREF.findall(c.value):
                        if int(rr) < B.DB_R0:
                            bad.append(f"{sheet.title}!{c.coordinate} references "
                                       f"Database!{L}{rr} — above the first data row")

    # ---- 6. the two deliverables must carry the same field set -------------
    import build_dashboard as D
    xl = {f for _b, _h, f, _fm, _w in B.DB_COLS if not f.startswith("_")}
    html = set(D.KEEP)
    # deliberate exceptions: HTML-only helpers with no column of their own
    HTML_ONLY = {"a_share_proxy", "sponsors_cn", "sponsors_display"}
    missing_html = xl - html
    missing_xl = html - xl - HTML_ONLY
    checked += len(xl)
    for f in sorted(missing_html):
        bad.append(f"field {f!r} is an Excel Database column but never reaches the HTML")
    for f in sorted(missing_xl):
        bad.append(f"field {f!r} ships to the HTML but has no Excel Database column")

    # ---- 7. the desk bundle must run the SAME stages as the Mac ------------
    import re as _re
    ipo_src = (ROOT / "ipo.py").read_text()
    bun_src = (ROOT / "ipo_lib" / "make_bundle.py").read_text()
    def stages(src, key):
        blk = _re.search(key + r"\s*=\s*\[(.*?)\n\]", src, _re.S)
        return {m for m in _re.findall(r'\("([a-z0-9-]+)",', blk.group(1))} if blk else set()
    mac = stages(ipo_src, "STAGES")
    desk = stages(bun_src, "STAGES")
    DESK_CANNOT = {"aastocks-deal"}          # playwright only
    checked += len(mac)
    for st in sorted(mac - desk - DESK_CANNOT):
        bad.append(f"stage {st!r} runs on the Mac but is MISSING from the desk bundle")
    for st in sorted(desk - mac):
        bad.append(f"stage {st!r} is in the desk bundle but not in ipo.py")

    print(f"audit_formulas: {checked} formula bindings + field-set + stage-parity checked")
    for b in bad:
        print("  MISMATCH:", b)
    print("  RESULT:", "CLEAN" if not bad else f"{len(bad)} PROBLEMS")
    sys.exit(1 if bad else 0)


if __name__ == "__main__":
    main()
