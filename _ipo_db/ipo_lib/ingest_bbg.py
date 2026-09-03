#!/usr/bin/env python3
"""Ingest the desk's pasted-values Bloomberg workbook (bbg.xlsx).

The user ran the BBG Verify tab on the terminal and pasted the RESULTS as
values into bbg.xlsx (one flat sheet, headers on row 4, one row per Database
row). This turns those values into a proper batch so merge can treat Bloomberg
as a source with provenance, instead of the workbook being the only place the
numbers live.

Per the desk's instructions:
- retail (CP036) and institutional (CP037) subscription OVERRIDE the scraped
  values wherever Bloomberg printed a number;
- "Mkt cap at listing (BBG)" is NOT ingested — the desk formula behind it is
  wrong (values arrive 1e-6 off in mixed units), and market cap must stay on
  the filing-derived basis;
- P/E at listing (BBG) is kept BESIDE the scraped P/E, not over it — the two
  measure different things (see the investigation note in MAINTENANCE);
- the A/H conversion ticker is used to cross-check (and fill) a_share_code.
"""
import json, re
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = Path("/Users/lemon/Desktop/REPO/code/bbg.xlsx")
OUT = ROOT / "data" / "batches" / "bbg_desk.json"

COLS = {3: "retail_sub", 5: "instl_sub", 8: "shoe_exercised", 10: "pe_now",
        11: "ah_ticker", 15: "pe_ipo", 17: "a_pe_at_hipo", 18: "hsahp_at_ipo"}


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def a_code_of(tick):
    """'603341 CH Equity' -> '603341.SS' (Bloomberg CH = the A-line composite)."""
    m = re.match(r"\s*(\d{6})\s+C[HSZ]", str(tick or ""))
    if not m:
        return None
    c = m.group(1)
    suf = ".SS" if c[0] == "6" else ".BJ" if c[0] in "48" else ".SZ"
    return c + suf


def main():
    ws = openpyxl.load_workbook(SRC, data_only=True)["Sheet1"]
    recs = []
    for r in range(5, ws.max_row + 1):
        m = re.match(r"\s*(\d{3,5})\b", str(ws.cell(r, 1).value or ""))
        if not m:
            continue
        rec = {"code": m.group(1).zfill(4)}
        for c, key in COLS.items():
            v = ws.cell(r, c).value
            if key in ("shoe_exercised", "ah_ticker"):
                sv = str(v or "").strip()
                if sv and not sv.startswith("#N/A"):
                    rec[key] = sv
            else:
                nv = num(v)
                if nv is not None:
                    rec[key] = round(nv, 4)
        if rec.get("ah_ticker"):
            rec["a_share_code_bbg"] = a_code_of(rec["ah_ticker"])
        if len(rec) > 1:
            recs.append(rec)
    OUT.write_text(json.dumps({"batch": "bbg_desk", "source": str(SRC),
                               "pasted_asof": "2026-08-20",
                               "ingested": date.today().isoformat(),
                               "deals": recs}, ensure_ascii=False, indent=1))
    def cov(k):
        return sum(1 for x in recs if x.get(k) is not None)
    print(f"wrote {OUT}: {len(recs)} rows | retail {cov('retail_sub')} | "
          f"instl {cov('instl_sub')} | P/E now {cov('pe_now')} | "
          f"P/E@IPO {cov('pe_ipo')} | A-P/E@H-IPO {cov('a_pe_at_hipo')} | "
          f"A-ticker {cov('a_share_code_bbg')}")


if __name__ == "__main__":
    main()
