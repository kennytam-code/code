#!/usr/bin/env python3
"""Prove the Excel screener COMPUTES, not just that it contains formulas.

openpyxl writes formulas but never evaluates them, so a workbook can look
perfect and still return #REF! on the desk. This builds a miniature workbook
from the same generator and evaluates the real formula chain with `formulas`,
then asserts the top comp matches the Python reference implementation.

The Bloomberg BDP cells on the AH tab cannot resolve off-terminal; only the
Screener/Database chain is evaluated, which is exactly the part that must work
with no add-in and no macros.
"""
import json, math, sys
from datetime import date, timedelta
from pathlib import Path

import formulas

ROOT = Path(__file__).resolve().parent.parent
MINI = ROOT / "out" / "_test_mini.xlsx"
DB_R0_TEST = 5
N = 40


def python_reference(target, deals, cfg):
    W = cfg["weights"]

    def score(d):
        if d.get("name") == target.get("name"):
            return -1e9
        gate = 1 if target.get("subsector") and d.get("subsector") == target["subsector"] else 0
        sec = 1 if target.get("sector") and d.get("sector") == target["sector"] else 0
        size = 0.0
        # the sheet guards with N(), so a blank/text size scores 0 rather than
        # blowing up — mirror that instead of assuming a number arrives
        tsize = target.get("size")
        tsize = tsize if isinstance(tsize, (int, float)) and tsize > 0 else None
        if tsize and d.get("deal_size_hkdm"):
            size = max(0.0, 1 - abs(math.log10(d["deal_size_hkdm"] / tsize))
                       / cfg["size_proximity_log10_halfwidth"])
        # unknown on either side scores 0 — never a free "match"
        dp, tp = d.get("profitable_at_ipo"), target.get("profitable")
        prof = 1 if (dp is not None and tp is not None and (dp == "Y") == bool(tp)) else 0
        dh, th = d.get("is_h_share"), target.get("is_h")
        ah = 1 if (dh is not None and th is not None and bool(dh) == bool(th)) else 0
        rec = 0.0
        if d.get("ipo_date") and target.get("ref_date"):
            days = (date.fromisoformat(target["ref_date"])
                    - date.fromisoformat(d["ipo_date"][:10])).days
            rec = max(0.0, 1 - days / cfg["recency_horizon_days"])
        pe = 0.0
        tpe = target.get("pe")
        tpe = tpe if isinstance(tpe, (int, float)) and tpe > 0 else None
        if tpe and d.get("pe_ipo") and d["pe_ipo"] > 0:
            pe = max(0.0, 1 - abs(math.log10(d["pe_ipo"] / tpe))
                     / cfg.get("pe_proximity_log10_halfwidth", 0.6))
        return (W["subsector_match"] * gate + W["sector_match_fallback"] * sec * (1 - gate)
                + W["size_proximity"] * size + W["profitability_match"] * prof
                + W["h_share_match"] * ah + W["recency"] * rec
                + W.get("pe_proximity", 0) * pe)
    return sorted(deals, key=lambda d: -score(d))


def main():
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    import build_xlsx
    build_xlsx.main(limit=N, out=str(MINI))

    import openpyxl as _px
    mini_wb = _px.load_workbook(MINI)
    xl = formulas.ExcelModel().loads(str(MINI)).finish()
    sol = xl.calculate()
    book = MINI.name.upper()

    def cell(sheet, ref):
        for k, v in sol.items():
            if k.upper().endswith(f"'[{book}]{sheet.upper()}'!{ref}"):
                try:
                    return v.value[0, 0]
                except Exception:
                    return v
        return None

    # v6 layout: B5 = picked deal, D8..D13 = effective target attributes,
    # B15 rank mode, B17 A-share filter, comps start at row 21, Match in X.
    target_name = cell("SCREENER", "B5")
    # Find each control by its LABEL, never by a hardcoded address: this test
    # silently read the size cell and called it "subsector" once, after the card
    # was re-laid out. Labels are the contract; row numbers are an accident.
    def row_of(label):
        for r in range(4, 19):
            v = str(mini_wb["Screener"].cell(r, 1).value or "")
            if v.lower().startswith(label.lower()):
                return r
        raise SystemExit(f"card label {label!r} not found — layout changed")
    R = {k: row_of(k) for k in ("Sector", "Subsector", "Size", "Profitable",
                                "H-share", "IPO /", "Target P/E", "Public sub",
                                "Rank by", "A-share filter")}
    sub = cell("SCREENER", f"D{R['Subsector']}")
    comp1 = cell("SCREENER", "B21")
    # letters COMPUTED from the shared contract — hardcoded "AG" broke every
    # time a comp column was added
    from openpyxl.utils import get_column_letter as _gl
    L_MATCH = _gl(build_xlsx.COMP_COLS.index("Match") + 1)
    L_CS = _gl(build_xlsx.COMP_COLS.index("Shared CS") + 1)
    comp1_match = cell("SCREENER", f"{L_MATCH}21")
    bucket = cell("SCREENER", "H9")
    guidance = cell("SCREENER", "G5")
    print(f"  rank mode   : {cell('SCREENER', f'B{R[chr(82)+chr(97)+chr(110)+chr(107)+chr(32)+chr(98)+chr(121)]}')}"
          f"  | A-share filter: {cell('SCREENER', f'B{R[chr(65)+chr(45)+chr(115)+chr(104)+chr(97)+chr(114)+chr(101)+chr(32)+chr(102)+chr(105)+chr(108)+chr(116)+chr(101)+chr(114)]}')}")
    print(f"  shared-CS #1: {cell('SCREENER', f'{L_CS}21')}  (target top-5 split U1: "
          f"{str(cell('CALC (SCORING ENGINE)', 'U1'))[:40]})")
    print(f"  target      : {target_name}")
    print(f"  subsector   : {sub}")
    print(f"  size bucket : {bucket}")
    print(f"  guidance    : {str(guidance)[:96]}")
    print(f"  comp #1     : {comp1}  [{comp1_match}]")

    deals = json.loads((ROOT / "data" / "deals.json").read_text())["deals"][:N]
    cfg = json.loads((ROOT / "data" / "screener_config.json").read_text())
    # Score from the attributes Excel itself resolved, so this compares the
    # SCORING FORMULA rather than how each side derived the target's inputs.
    serial = cell("SCREENER", f"D{R['IPO /']}")
    ref_date = (date(1899, 12, 30) + timedelta(days=int(serial))).isoformat() \
        if isinstance(serial, (int, float)) else date.today().isoformat()
    d11, d12 = cell("SCREENER", f"D{R['Profitable']}"), cell("SCREENER", f"D{R['H-share']}")
    tgt = {"name": target_name, "sector": cell("SCREENER", f"D{R['Sector']}"), "subsector": sub,
           "size": cell("SCREENER", f"D{R['Size']}"), "pe": cell("SCREENER", f"D{R['Target P/E']}"),
           "profitable": (d11 == "Y") if d11 in ("Y", "N") else None,
           "is_h": (d12 == "Y") if d12 in ("Y", "N") else None,
           "ref_date": ref_date}
    print(f"  ref date    : {ref_date}  size {tgt['size']}")
    ref = python_reference(tgt, deals, cfg)
    expect = ref[0].get("name")
    print(f"  python ref  : {expect}")
    ok = (str(comp1).strip() == str(expect).strip())
    print("  PARITY      :", "MATCH" if ok else "MISMATCH")

    # The gate itself, not just the top name: same-subsector comps must carry
    # the "same subsector" label. A variable clobber once wrote the correct
    # gate formula for ROW 5 ONLY and pointed every later row at the public-
    # subscription column — parity still MATCHED because the top comp won on
    # the sector fallback. Assert the gate column on real rows instead.
    if sub:
        same = [i for i, d in enumerate(deals) if d.get("subsector") == sub
                and d.get("name") != target_name]
        gates = [cell("CALC (SCORING ENGINE)", f"B{DB_R0_TEST + i}") for i in same[:6]]
        gate_ok = bool(gates) and all(g == 1 for g in gates)
        print(f"  subsec gate : {len(gates)} same-subsector rows -> {gates} "
              f"[{'OK' if gate_ok else 'FAIL — gate not reading the subsector column'}]")
        if not gate_ok:
            ok = False

    # every visible cell must evaluate — no #DIV/0!, #REF!, #VALUE! anywhere,
    # which is the failure the desk actually sees
    errs = []
    for k, v in sol.items():
        if f"[{book}]" not in k.upper():
            continue
        try:
            val = v.value[0, 0]
        except Exception:
            continue
        if isinstance(val, str) and val.startswith("#") and val.endswith("!"):
            # Bloomberg-chain cells (BDP/BDH and the cells that read them) are
            # terminal-only BY DESIGN; the offline evaluator computes both IF
            # branches eagerly, so their guards cannot protect them here.
            f0 = ""
            try:
                cellref = k.split("!")[-1]
                sheetn = k.split("]")[-1].split("'")[0]
                f0 = str(mini_wb[sheetn][cellref].value or "")
            except Exception:
                pass
            if "BDP(" in f0 or "BDH(" in f0 or "A_PX" in f0 or "A_HKD" in f0:
                continue
            errs.append(f"{k.split(']')[-1]} = {val}")
    print(f"  error cells : {len(errs)}" + (f"  e.g. {errs[:4]}" if errs else ""))
    if errs:
        ok = False
    if isinstance(comp1, str) and comp1.startswith("#"):
        print("  FAIL: formula returned an Excel error")
        ok = False

    # --- cornerstone-overlap fixture: prove the rank-by-cornerstone machinery
    # end-to-end. Find two deals in the mini that SHARE an investor, re-solve
    # the workbook with the first as the picked target, and assert the split
    # cell fills and the other deal's overlap count is >= 1.
    listed = [d for d in deals if d.get("cornerstone_investors")]
    pair = None
    for i, a in enumerate(listed):
        for b in listed[i + 1:]:
            shared = set(map(str.lower, a["cornerstone_investors"])) \
                & set(map(str.lower, b["cornerstone_investors"]))
            if shared:
                pair = (a, b, sorted(shared)[0])
                break
        if pair:
            break
    if pair:
        a, b, inv = pair
        # input keys are CASE-SENSITIVE on the book name — the display keys are
        # upper-cased but overrides must use the file's real name
        sol2 = xl.calculate(inputs={
            f"'[{MINI.name}]SCREENER'!B5": a["name"]})

        def cell2(sheet, ref):
            for k, v in sol2.items():
                if k.upper().endswith(f"'[{book}]{sheet.upper()}'!{ref}"):
                    try:
                        return v.value[0, 0]
                    except Exception:
                        return v
            return None
        u1 = cell2("CALC (SCORING ENGINE)", "U1")
        row_b = 5 + deals.index(b)          # DB_R0 + index
        overlap = cell2("CALC (SCORING ENGINE)", f"H{row_b}")
        cs_ok = bool(u1) and isinstance(overlap, (int, float)) and overlap >= 1
        print(f"  CS fixture  : target={a['name'][:20]} shares '{inv[:24]}' with "
              f"{b['name'][:20]} -> split U1={'set' if u1 else 'EMPTY'}, "
              f"overlap={overlap} [{'OK' if cs_ok else 'FAIL'}]")
        if not cs_ok:
            ok = False
    else:
        print("  CS fixture  : no overlapping pair inside the mini slice — skipped")
    # --- P/E-target probe: the comp table must survive a target WITH a P/E.
    # pe_ipo blanks hold Bloomberg-fallback TEXT; LOG10(text/x) is #VALUE!, and
    # that only fires when D13 carries a number — the default fixture (no P/E)
    # sailed past it while every pipeline pick with an expected multiple broke.
    sol3 = xl.calculate(inputs={f"'[{MINI.name}]SCREENER'!C13": 30})

    def cell3(sheet, ref):
        for k, v in sol3.items():
            if k.upper().endswith(f"'[{book}]{sheet.upper()}'!{ref}"):
                try:
                    return v.value[0, 0]
                except Exception:
                    return v
        return None
    pe_errs = 0
    for k, v in sol3.items():
        if f"[{book}]" not in k.upper():
            continue
        try:
            val = v.value[0, 0]
        except Exception:
            continue
        if isinstance(val, str) and val.startswith("#") and val.endswith("!"):
            cellref = k.split("!")[-1]
            sheetn = k.split("]")[-1].split("'")[0]
            f0 = ""
            try:
                f0 = str(mini_wb[sheetn][cellref].value or "")
            except Exception:
                pass
            if "BDP(" in f0 or "BDH(" in f0 or "A_PX" in f0 or "A_HKD" in f0:
                continue
            pe_errs += 1
    c1pe = cell3("SCREENER", "B21")
    pe_ok = pe_errs == 0 and isinstance(c1pe, str) and not c1pe.startswith("#")
    print(f"  P/E target  : override C13=30 -> comp #1 {str(c1pe)[:20]!r}, "
          f"{pe_errs} error cells [{'OK' if pe_ok else 'FAIL'}]")
    if not pe_ok:
        ok = False

    # --- A-share filter probe: with the hard filter ON, every RENDERED comp
    # row must be an A/H deal, and slots past the last valid comp must be
    # BLANK. The gated -999999 scores once leaked into the tail as phantom
    # non-A comps ("i set with a share then why those without still showing").
    r_filter = R["A-share filter"]
    sol4 = xl.calculate(inputs={f"'[{MINI.name}]SCREENER'!B{r_filter}": "With A-share"})

    def cell4(sheet, ref):
        for k, v in sol4.items():
            if k.upper().endswith(f"'[{book}]{sheet.upper()}'!{ref}"):
                try:
                    return v.value[0, 0]
                except Exception:
                    return v
        return None
    a_names = {d["name"] for d in deals if d.get("a_share_code")}
    tgt = cell4("SCREENER", "B5")
    shown = []
    for k in range(15):
        nm = cell4("SCREENER", f"B{21 + k}")
        if isinstance(nm, str) and nm and nm != "—" and not nm.startswith("#"):
            shown.append(nm)
    leaks = [nm for nm in shown if nm not in a_names and nm != tgt]
    n_a_avail = len([d for d in deals if d.get("a_share_code")
                     and d["name"] != tgt])
    af_ok = not leaks and (len(shown) <= max(1, n_a_avail))
    print(f"  A-filter    : With A-share -> {len(shown)} rows shown, "
          f"{n_a_avail} A/H comps available, leaks={leaks[:3]} "
          f"[{'OK' if af_ok else 'FAIL'}]")
    if not af_ok:
        ok = False

    # --- force-include probe: a typed code must pin that deal to comp #1,
    # even THROUGH the A-share filter it would otherwise fail
    tgt0 = cell("SCREENER", "B5")
    victim = next((d for d in reversed(deals)
                   if d["name"] != tgt0 and not d.get("a_share_code")), None)
    if victim:
        sol5 = xl.calculate(inputs={
            f"'[{MINI.name}]SCREENER'!D17": victim["code"],
            f"'[{MINI.name}]SCREENER'!B{r_filter}": "With A-share"})

        def cell5(sheet, ref):
            for k, v in sol5.items():
                if k.upper().endswith(f"'[{book}]{sheet.upper()}'!{ref}"):
                    try:
                        return v.value[0, 0]
                    except Exception:
                        return v
            return None
        top1 = cell5("SCREENER", "B21")
        fi_ok = top1 == victim["name"]
        print(f"  force-incl  : D17={victim['code']} (non-A deal, filter ON) -> "
              f"comp #1 {str(top1)[:22]!r} [{'OK' if fi_ok else 'FAIL'}]")
        if not fi_ok:
            ok = False
    else:
        print("  force-incl  : no non-A deal in the mini slice — skipped")

    MINI.unlink(missing_ok=True)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
