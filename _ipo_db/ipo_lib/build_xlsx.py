#!/usr/bin/env python3
"""Build out/HK_IPO_Database_v1.xlsx from data/deals.json (+ taxonomy, config,
pipeline). House style: module-level Arial fonts + semantic fills, put/hdr/inp/
calc helpers, DataValidation dropdowns, FormulaRule conditional formats.

Formula policy: Excel-2016-safe only (INDEX/MATCH/LARGE/COUNTIF/SUMPRODUCT,
no XLOOKUP/LET/dynamic arrays). Bloomberg BDP formulas are written next to
manual-override input cells; effective value = IF(override<>"",override,live).
BDP mnemonics are AWAIT-TERMINAL-VERIFY until confirmed on the desk.
"""
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "out" / "HK_IPO_Database_v1.xlsx"

ARIAL = "Arial"
TITLE = Font(name=ARIAL, size=13, bold=True)
SUB   = Font(name=ARIAL, size=9, italic=True, color="00808080")
SECT  = Font(name=ARIAL, size=10, bold=True)
WARN  = Font(name=ARIAL, size=10, bold=True, color="00C00000")
BODY  = Font(name=ARIAL, size=10)
BLUE  = Font(name=ARIAL, size=10, color="000000FF")
NOTE  = Font(name=ARIAL, size=9, italic=True, color="00808080")
HDRF  = Font(name=ARIAL, size=10, bold=True, color="00FFFFFF")
BOLD  = Font(name=ARIAL, size=10, bold=True)

F_HDR  = PatternFill("solid", fgColor="001F3864")   # navy
F_IN   = PatternFill("solid", fgColor="00DDEBF7")   # input blue
F_CALC = PatternFill("solid", fgColor="00F2F2F2")   # calc grey
F_ZEBRA = PatternFill("solid", fgColor="00F7F9FB")  # even-row wash, display only
F_GRN  = PatternFill("solid", fgColor="00C6EFCE")   # xchecked
F_AMB  = PatternFill("solid", fgColor="00FFE699")   # judgment / estimated
F_OVR  = PatternFill("solid", fgColor="00FFC000")   # conflict
_t = Side(style="thin")
BOX = Border(left=_t, right=_t, top=_t, bottom=_t)
C_HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_MID = Alignment(horizontal="center")

PCT = '0.0"%"'
MONEY = "#,##0"
PX = "0.000"
DATEF = "yyyy-mm-dd"

# ONE data fill only: orange = the sources disagreed (see conflicts.json).
# Everything else is plain — the desk asked for a clean sheet, and provenance
# detail lives in data/deals.json, not in cell paint.
STATUS_FILL = {"conflict": F_OVR}


def put(ws, coord, value, font=BODY, fill=None, fmt=None, border=None, align=None):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill:   c.fill = fill
    if fmt:    c.number_format = fmt
    if border: c.border = border
    if align:  c.alignment = align
    return c


def hdr(ws, coord, text, fill=F_HDR):
    put(ws, coord, text, font=HDRF, fill=fill, border=BOX, align=C_HDR)


def inp(ws, coord, value, fmt=None):
    put(ws, coord, value, font=BLUE, fill=F_IN, fmt=fmt, border=BOX)


def calc(ws, coord, value, fmt=None):
    put(ws, coord, value, font=BODY, fill=F_CALC, fmt=fmt, border=BOX)



# Chinese names for pipeline applicants, matched against the AAStocks
# planned-IPO roll (English feed names carry no CN name of their own)
_CN_HINTS = {"mech-mind": "梅卡曼德", "shein": "希音", "ingenic": "君正",
             "tanboer": "坦博爾", "direct drive": "本末", "ekh": "EKH"}


def _press_size(root, applicant):
    """Press-reported expected size for a PHIP name — an ESTIMATE, labelled."""
    import json as _j
    p = root / "data" / "batches" / "press_sizes.json"
    if not p.exists():
        return {}
    low = (applicant or "").lower()
    for r in _j.loads(p.read_text())["results"]:
        if r["match"] in low and r.get("expected_size_hkdm"):
            return {"expected_size_hkdm": r["expected_size_hkdm"],
                    "basis": f"estimated ({r['confidence']}: {r['source']})",
                    "a_share_code": r.get("a_share_code")}
    return {}


def _planned_cn(root, applicant):
    import json as _j
    p = root / "data" / "batches" / "aastocks_planned.json"
    if not p.exists():
        return None, None
    low = (applicant or "").lower()
    hint = next((v for k, v in _CN_HINTS.items() if k in low), None)
    if not hint:
        return None, None
    for r in _j.loads(p.read_text())["rows"]:
        if hint in r["name"]:
            return r["name"], r.get("industry")
    return None, None

def _offering_rows(root):
    """Deals in their OFFERING WINDOW (www2 New Listings) — real terms."""
    import json as _j
    p = root / "data" / "batches" / "newlistings.json"
    if not p.exists():
        return []
    out = []
    for r in _j.loads(p.read_text())["deals"]:
        cn, industry = _planned_cn(root, r.get("name"))
        press = _press_size(root, r.get("name"))
        ni, rev = r.get("ni_latest"), r.get("rev_latest")
        pe_mid = None
        if r.get("pe_expected_lo") and r.get("pe_expected_hi"):
            pe_mid = round((r["pe_expected_lo"] + r["pe_expected_hi"]) / 2, 1)
        elif r.get("pe_at_h_cap"):
            pe_mid = r["pe_at_h_cap"]
        out.append({
            "name": r.get("name"), "expected_code": r.get("code"),
            "name_cn": cn, "industry_en": r.get("industry_en") or industry,
            "subsector": r.get("subsector"), "sector": r.get("sector"),
            "status": r.get("status") or "OFFERING NOW",
            "expected_timing": r.get("listing_date") or "",
            "range_lo": r.get("range_lo"), "range_hi": r.get("range_hi"),
            "pe_expected_lo": r.get("pe_expected_lo"),
            "pe_expected_hi": r.get("pe_expected_hi"),
            "pe_expected_mid": pe_mid,
            "ps_expected_lo": r.get("ps_expected_lo"),
            "cornerstone_pct": r.get("cornerstone_pct"),
            "lot_size": r.get("lot_size"),
            "offer_period": r.get("offer_period"),
            "rev_latest": rev, "ni_latest": ni,
            "profitable_at_ipo": ("Y" if ni and ni > 0 else "N") if ni is not None else None,
            "sponsors": "; ".join(r.get("sponsors") or []) or None,
            "expected_size_hkdm": r.get("expected_net_hkdm") or press.get("expected_size_hkdm"),
            "expected_size_basis": ("prospectus: net proceeds at the maximum price"
                                    if r.get("expected_net_hkdm") else press.get("basis")),
            "business_desc": r.get("business_overview"),
            "doc_link": (r.get("prospectus_links") or [None])[0],
            "a_share_code": r.get("a_share_code") or press.get("a_share_code"),
            "is_h_share": "Y" if (r.get("a_share_code") or press.get("a_share_code")) else None,
            "a_price_now": r.get("a_price_now"), "fx_now": r.get("fx_now"),
            "a_pe_ttm": r.get("a_pe_ttm"), "a_mktcap_bn_cny": r.get("a_mktcap_bn_cny"),
            "h_cap_vs_a_pct": r.get("h_cap_vs_a_pct"), "pe_at_h_cap": r.get("pe_at_h_cap"),
            # same fact in the book's direction (A over H), so the pipeline row
            # is comparable with the Database's "A prem vs H at IPO" column
            "a_prem_vs_hcap_pct": (round((1 / (1 + r["h_cap_vs_a_pct"] / 100) - 1) * 100, 2)
                                   if r.get("h_cap_vs_a_pct") not in (None,)
                                   and r["h_cap_vs_a_pct"] > -100 else None),
            "use_of_proceeds": r.get("use_of_proceeds"),
            "code": r.get("code"),
        })
    return out


def _phip_as_pipeline(root, existing_names):
    """PHIP-stage applicants ARE the active pipeline — hearing cleared, weeks
    from pricing. Convert their parsed records into pipeline rows; AP-only
    applicants stay in the watch queue."""
    import json as _j
    try:
        apps = _j.loads((root / "data" / "batches" / "phip_pipeline.json").read_text())["applications"]
    except Exception:
        return []
    out = _offering_rows(root)
    low = [n.lower().split()[0] for n in existing_names if n]
    low += [(x.get("name") or "").lower().split()[0] for x in out if x.get("name")]
    for a in apps:
        if not a.get("has_phip") or a.get("sponsor_terminated"):
            continue
        nm = (a.get("applicant") or "").split("(formerly")[0].strip().rstrip(",")
        first = nm.lower().split()[0] if nm else ""
        if first and first in low:
            continue                      # already covered by a curated entry (Shein)
        p = a.get("parsed") or {}
        prof = p.get("profitable_at_ipo")
        cn, industry = _planned_cn(root, nm)
        press = _press_size(root, nm)
        out.append({
            "name": nm,
            "name_cn": cn,
            "sector": p.get("sector"), "subsector": p.get("subsector"),
            "status": "PHIP — hearing cleared",
            "expected_timing": f"PHIP {a.get('latest_submission') or ''}",
            "rev_latest": p.get("rev_latest"), "ni_latest": p.get("ni_latest"),
            "profitable_at_ipo": ("Y" if prof else "N") if prof is not None else None,
            "business_desc": p.get("business_overview"),
            "doc_link": a.get("doc_link"),
            # an A-to-H applicant prices off a live quote — say so, and say what
            # is still unknown rather than leaving the cell blank
            "is_h_share": "Y" if p.get("is_ah_applicant") else None,
            "a_share_code": p.get("a_share_code"),
            "valuation_notes": "; ".join(filter(None, [
                p.get("a_share_note"),
                f"subsector via {p['subsector_src']}" if p.get("subsector_src") else None,
                "expected size not yet public (PHIP stage) — type it in the blue cell "
                "to rank comps on size" if not p.get("expected_shares") else None])) or None,
            "sponsors": "; ".join((p.get("sponsors") or [])
                                   or (p.get("coordinators") or [])) or None,
            "expected_size_hkdm": press.get("expected_size_hkdm"),
            "expected_size_basis": press.get("basis"),
        })
    return out


def load():
    deals = json.loads((ROOT / "data" / "deals.json").read_text())["deals"]
    tax = json.loads((ROOT / "data" / "taxonomy.json").read_text())
    cfg = json.loads((ROOT / "data" / "screener_config.json").read_text())
    pipe_p = ROOT / "data" / "batches" / "pipeline.json"
    pipe = json.loads(pipe_p.read_text())["deals"] if pipe_p.exists() else []
    pipe = pipe + _phip_as_pipeline(ROOT, [str(x.get("name")) for x in pipe])
    from pipeline_dedupe import merge_pipeline
    pipe = merge_pipeline(pipe)      # one row per company (the SHEIN split)
    counts_p = ROOT / "data" / "official_counts.json"
    counts = json.loads(counts_p.read_text()) if counts_p.exists() else {}
    return deals, tax, cfg, pipe, counts


# ---------------------------------------------------------------- Database ---
# (band, header, field, fmt, width) — bands render as a merged section row so
# related columns read as one group instead of "a bunch of stuff"
DB_COLS = [
    ("IDENTITY", "Code", "code", "@", 7),
    ("IDENTITY", "Name", "name", None, 22),
    ("IDENTITY", "Name (CN)", "name_cn", None, 13),
    ("IDENTITY", "Sector", "sector", None, 12),
    ("IDENTITY", "Subsector", "subsector", None, 22),
    ("IDENTITY", "Industry (AAStocks)", "industry_en", None, 20),
    ("IDENTITY", "Regime", "listing_regime", None, 9),
    ("IDENTITY", "IPO date", "ipo_date", DATEF, 11),
    ("DEAL TERMS", "Deal size (HK$m)", "deal_size_hkdm", MONEY, 12),
    ("DEAL TERMS", "Size basis", "size_basis", None, 15),
    ("DEAL TERMS", "Range low (HK$)", "price_range_lo", PX, 9),
    ("DEAL TERMS", "Max/cap (HK$)", "price_range_hi", PX, 9),
    ("DEAL TERMS", "Final px (HK$)", "final_price", PX, 9),
    ("DEAL TERMS", "Priced at cap", "priced_at_cap", None, 8),
    ("DEAL TERMS", "% of cap", "pct_of_cap", '0.0"%"', 8),
    ("DEAL TERMS", "Priced in range %", "pct_in_range", '0.0"%"', 11),
    ("DEAL TERMS", "Mkt cap at IPO (HK$m)", "mktcap_ipo_hkdm", MONEY, 12),
    ("DEAL TERMS", "Mkt cap basis", "mktcap_basis", None, 26),
    ("DEMAND", "Public sub (x)", "oversub_public_mult", "#,##0.0", 10),
    ("DEMAND", "Intl sub (x)", "oversub_intl_mult", "0.00", 9),
    ("DEMAND", "Cornerstone (% of offer)", "cornerstone_pct", '0.0"%";;"none"', 11),
    # The float THREE ways, all off one identity: offer less the locked-up
    # cornerstone take. The money and share columns need no market cap, so they
    # survive the two deals whose cap is not derivable; the % is the same
    # number over the cap.
    ("DEMAND", "Eff. free float (HK$m)", "eff_free_float_hkdm", MONEY, 13),
    ("DEMAND", "Eff. free float (shares)", "eff_free_float_shares", "#,##0", 15),
    ("DEMAND", "Eff. free float (% of cap)", "eff_free_float_pct", '0.0"%"', 11),
    ("DEMAND", "Cornerstone investors", "cornerstone_investors", None, 46),
    ("DEMAND", "CS keys", "cornerstone_keys", None, 16),
    ("DEMAND", "Greenshoe size (%)", "greenshoe_pct", '0.0"%"', 9),
    ("DEMAND", "Shoe outcome (final)", "greenshoe_exercised_final", None, 17),
    # the date the stabilisation mandate dies, from the filing's own sentence —
    # the aftermarket desk plans around this, so it belongs on the row
    ("DEMAND", "Shoe ends (filed)", "stabilization_end_date", None, 13),
    # the bank that held the shoe and the after-market bid — a different
    # question from who sponsored the deal, and the key the SM League groups on
    ("DEMAND", "Stabilising manager", "stabilizing_manager", None, 34),
    # the bank-family key the SM League groups on — same treatment as CS keys,
    # visible so the grouping is never a black box
    ("DEMAND", "SM key", "stabilizing_manager_key", None, 14),
    ("PERFORMANCE", "Day-1", "first_day_return_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "Day-1 open pop", "day1_open_pop_pct", '+0.0"%";-0.0"%"', 10),
    ("PERFORMANCE", "Day-1 open→close", "day1_open_close_pct", '+0.0"%";-0.0"%"', 11),
    ("PERFORMANCE", "1-week", "ret_1w_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "1-month", "ret_1m_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "3-month", "ret_3m_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "1w ex-pop", "aftermkt_1w_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "1m ex-pop", "aftermkt_1m_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "3m ex-pop", "aftermkt_3m_pct", '+0.0"%";-0.0"%"', 9),
    ("PERFORMANCE", "Alpha 1w vs index", "alpha_1w_pct", '+0.0"%";-0.0"%"', 10),
    ("PERFORMANCE", "Alpha 1m vs index", "alpha_1m_pct", '+0.0"%";-0.0"%"', 10),
    ("PERFORMANCE", "Alpha 3m vs index", "alpha_3m_pct", '+0.0"%";-0.0"%"', 10),
    ("PERFORMANCE", "Alpha 1m ex-pop", "alpha_1m_expop_pct", '+0.0"%";-0.0"%"', 10),
    ("PERFORMANCE", "Index, IPO→1m", "bench_1m_pct", '+0.0"%";-0.0"%"', 11),
    ("PERFORMANCE", "Index, d1→1m", "bench_1m_expop_pct", '+0.0"%";-0.0"%"', 11),
    ("PERFORMANCE", "Benchmark", "benchmark", None, 22),
    ("PERFORMANCE", "Since IPO", "since_ipo_pct", '+0.0"%";-0.0"%"', 10),
    ("PERFORMANCE", "Prices as of", "price_asof", DATEF, 11),
    ("FUNDAMENTALS", "Latest FY rev (HK$m)", "rev_latest", MONEY, 11),
    ("FUNDAMENTALS", "Latest FY NI (HK$m)", "ni_latest", MONEY, 11),
    ("FUNDAMENTALS", "Profitable", "profitable_at_ipo", None, 8),
    ("FUNDAMENTALS", "P/E at IPO", "pe_ipo", "0.0", 8),
    ("FUNDAMENTALS", "P/E at IPO (BBG)", "pe_ipo_bbg", "0.0", 10),
    ("FUNDAMENTALS", "P/E today (BBG)", "pe_now", "0.0", 10),
    ("FUNDAMENTALS", "P/S at IPO", "ps_ipo", "0.0", 8),
    # P/S TODAY has no public source — a filing states revenue and a cap at the
    # OFFER, never a live multiple — so every row of this column resolves off
    # BBG Verify col S on the terminal and says so off it.
    ("FUNDAMENTALS", "P/S today (BBG)", "ps_now", "0.0", 11),
    ("A / H", "A/H pair", "is_h_share", None, 8),
    ("A / H", "A-share code", "a_share_code", None, 11),
    ("A / H", "A prem vs H at IPO", "a_premium_ipo_pct", '+0.0"%";-0.0"%"', 11),
    ("A / H", "A close used (HK$)", "a_close_hkd", PX, 10),
    ("A / H", "A premium (today)", "a_premium_now", "+0.0%;-0.0%", 11),
    # company-level cap from the A line: ALL share classes x A price (Tencent
    # field-45 basis, verified vs ICBC), CNY->HKD at the snapshot FX. This is
    # the "Luxshare is not small" number — the H tranche alone says HK$24bn
    # while the company is ~HK$450bn.
    ("A / H", "A-line mkt cap now (HK$m)", "a_mktcap_now_hkdm", MONEY, 14),
    ("A / H", "A P/E @H-IPO (BBG)", "_a_pe_at_ipo_bbg", "0.0", 11),
    ("A / H", "HSAHP @IPO (BBG)", "_hsahp_bbg", "0.0", 11),
    ("BANKS & DOCS", "Sponsors", "sponsors", None, 30),
    ("BANKS & DOCS", "Sponsors (AAStocks)", "sponsors_en", None, 30),
    ("BANKS & DOCS", "Bookrunners", "bookrunners_display", None, 34),
    ("BANKS & DOCS", "Underwriters (AAStocks)", "underwriters_en", None, 34),
    ("BANKS & DOCS", "Prospectus", "prospectus_link", None, 11),
    ("BANKS & DOCS", "Allotment result", "allotment_link", None, 13),
    ("BANKS & DOCS", "Stabilisation notice", "stabilization_link", None, 15),
    ("NOTES", "Valuation notes", "valuation_notes", None, 40),
    # replaces the old always-empty "Valuation notes": every explained absence
    # collected into one readable cell, so a blank on this row is never a mystery
    ("NOTES", "Why anything is blank", "_blank_notes", None, 64),
]
# the *_note fields merge_batches writes, in the order they read best
NOTE_FIELDS = [("intl_note", "intl sub"), ("pe_note", "P/E"), ("shoe_note", "greenshoe"),
               ("range_note", "price range"), ("range_lo_note", "range low"),
               ("ah_note", "A/H"),
               ("ret_note", "returns"), ("ret_1w_note", "1w"), ("ret_1m_note", "1m"),
               ("ret_3m_note", "3m"), ("sponsor_note", "sponsor"),
               ("mktcap_note", "market cap"), ("cornerstone_note", "cornerstone"),
               ("cornerstone_list_note", "cornerstone list"),
               ("cornerstone_pct_note", "cornerstone %"),
               ("greenshoe_note", "greenshoe size"), ("ps_note", "P/S"),
               ("stabmgr_note", "stabilising manager"),
               ("fin_note", "financials"), ("size_note", "deal size"),
               # v13: reasons that existed in the data but never reached the
               # "Why anything is blank" column — a blank with an unread reason
               # is indistinguishable from an unexplained blank
               ("fin_check", "financials rejected"), ("day1_oc_note", "day-1 open"),
               # v26 sweep: every reason the merge writes must appear here — an
               # unread reason is indistinguishable from an unexplained blank
               ("price_note", "price series"),
               ("doc_note", "filing links"),
               ("eff_ff_note", "eff free float"),
               ("eff_ff_shares_note", "eff float shares"),
               ("ipo_date_note", "listing date"),
               ("stabilization_note", "stabilisation notice"),
               ("alpha_note", "alpha/index"), ("oversub_public_mult_note", "public sub"),
               ("aastocks_note", "AAStocks fields"), ("size_basis_note", "size basis"),
               ("pct_of_cap_note", "% of cap")]
LINK_FIELDS = {"prospectus_link": "prospectus",
               "allotment_link": "allotment",
               "stabilization_link": "stabilisation"}
# Fields the public filings sometimes never state, mapped to the BBG Verify
# column that answers them. A cell with no scraped value becomes a live pull on
# the terminal instead of a blank; everywhere else it explains itself.
BBG_FALLBACK = {"oversub_intl_mult": "E", "oversub_public_mult": "C",
                "a_share_code": "K", "mktcap_ipo_hkdm": "M", "pe_ipo": "O",
                # no public source exists for either: a live P/S, and the A
                # line's cap TODAY where the Tencent snapshot had none
                "ps_now": "S", "a_mktcap_now_hkdm": "T"}
BBG_BLANK_TEXT = "not filed — run on terminal"
# A/H columns that read N/A (not blank, never 0) when there is no A line
AH_FIELDS = {"a_premium_ipo_pct", "a_close_hkd", "a_premium_now",
             "a_mktcap_now_hkdm"}
DB_R0 = 5  # first data row
# scoring engine lives on its own, clearly-titled sheet — nothing hidden and
# nothing overlapping the data (v4 bug: helpers at AK-AR were overwritten by
# new data columns, which painted Sponsors/links with 0/1 formulas)
CALC = "'Calc (scoring engine)'"
H_GATE, H_SEC, H_SIZE, H_PROF, H_AH, H_REC, H_CS, H_PE, H_DEM, H_SCORE = (
    "B", "C", "D", "E", "F", "G", "H", "I", "J", "K")


def listify(v):
    return "; ".join(v) if isinstance(v, list) else v


def sheet_database(wb, deals, n):
    ws = wb.create_sheet("Database")
    put(ws, "A1", "HK IPO DATABASE — Main Board IPOs 2021–2026", TITLE)
    put(ws, "A2", "one row per deal · money HK$m · % in percent units · sign: Day-1 + = closed above offer  ·  "
                  "fill legend: green=cross-checked  amber=judgment/estimated  orange=source conflict  grey=derived", SUB)
    # band row (row 3): merged section titles so groups read as one unit
    from itertools import groupby
    BAND_TINT = {"IDENTITY": "00203864", "DEAL TERMS": "00274E13", "DEMAND": "007A5C00",
                 "PERFORMANCE": "00203864", "FUNDAMENTALS": "00434343",
                 "A / H": "00742323", "BANKS & DOCS": "00274E63", "NOTES": "00434343"}
    col = 1
    for band, grp in groupby(DB_COLS, key=lambda c: c[0]):
        width = len(list(grp))
        put(ws, f"{get_column_letter(col)}3", band, HDRF,
            PatternFill("solid", fgColor=BAND_TINT.get(band, "001F3864")),
            border=BOX, align=C_HDR)
        if width > 1:
            ws.merge_cells(start_row=3, start_column=col, end_row=3,
                           end_column=col + width - 1)
        col += width
    band_starts = set()
    seen_band, colx = None, 1
    for b_, *_r in DB_COLS:
        if b_ != seen_band:
            band_starts.add(colx)
            seen_band = b_
        colx += 1
    F_EXPOP_HDR = PatternFill("solid", fgColor="001F7A6D")   # teal = ex-pop family
    for j, (_b, h, *_rest) in enumerate(DB_COLS, 1):
        expop = "ex-pop" in h or "open→close" in h
        hdr(ws, f"{get_column_letter(j)}4", h,
            fill=F_EXPOP_HDR if expop else PatternFill("solid", fgColor="00305496"))
        ws.column_dimensions[get_column_letter(j)].width = DB_COLS[j - 1][4]
    for i, d in enumerate(deals):
        r = DB_R0 + i
        pv = d.get("_prov", {})
        for j, (_b, _h, f, fmt, _w) in enumerate(DB_COLS, 1):
            if f in ("_a_pe_at_ipo_bbg", "_hsahp_bbg"):
                # Bloomberg-only history: lives on the terminal, explains itself
                # here. _hsahp_bbg = the Hang Seng AH Premium index on the
                # listing day (BBG Verify col R); the A-P/E sits in col Q.
                vcol = "Q" if f == "_a_pe_at_ipo_bbg" else "R"
                c2 = f"{get_column_letter(j)}{r}"
                if d.get("a_share_code"):
                    put(ws, c2, f"=IF('BBG Verify'!{vcol}{r}=\"\",\"run on terminal\","
                                f"'BBG Verify'!{vcol}{r})", fmt=fmt, fill=F_CALC)
                else:
                    put(ws, c2, "N/A")
                continue
            if f == "_blank_notes":
                v = " · ".join(f"{lbl}: {d[k]}" for k, lbl in NOTE_FIELDS if d.get(k))
            else:
                v = listify(d.get(f))
            # the flag is set only where an A line was found, so an empty cell
            # is the answer "no pair" — write it rather than leaving a blank
            if f == "is_h_share" and v in (None, ""):
                v = "N"
            # a blank multiple is an ANSWER: n/m = loss-maker (no earnings to
            # divide by), pre-rev = the filed P&L has no revenue line at all
            if f == "pe_ipo" and v in (None, "") and d.get("profitable_at_ipo") == "N":
                v, fmt = "n/m", None
            if f == "ps_ipo" and v in (None, "") and d.get("rev_latest") == 0:
                v, fmt = "pre-rev", None
            # an issuer with no A line has no A/H numbers to show — say N/A
            # rather than leaving cells that read as a zero premium
            if (v in (None, "") and f in AH_FIELDS and not d.get("a_share_code")):
                v = "N/A"
                fmt = None
            if isinstance(v, bool):
                v = "Y" if v else "N"
            if f == "ipo_date" and v:
                v = date.fromisoformat(v[:10])
            st = (pv.get(f) or {}).get("status")
            fill = STATUS_FILL.get(st)
            cell = f"{get_column_letter(j)}{r}"
            if f in LINK_FIELDS and v:
                # clickable, so "did you actually read the filing?" is checkable
                put(ws, cell, f'=HYPERLINK("{v}","{LINK_FIELDS[f]}")',
                    font=Font(name=ARIAL, size=10, color="000000FF", underline="single"))
            elif v in (None, "") and f in BBG_FALLBACK:
                # Nothing public filled this cell. Rather than leave it empty,
                # point it at the matching BBG Verify column: on the terminal it
                # fills itself, off the terminal it says so in words.
                src = BBG_FALLBACK[f]
                put(ws, cell, f"=IF('BBG Verify'!{src}{r}=\"\",\"{BBG_BLANK_TEXT}\","
                              f"'BBG Verify'!{src}{r})", fmt=fmt, fill=F_CALC)
            else:
                put(ws, cell, v, fmt=fmt, fill=fill)
            if j in band_starts and j > 1:
                cc = ws[cell]
                cc.border = Border(left=Side(style="medium", color="00305496"),
                                   right=cc.border.right, top=cc.border.top,
                                   bottom=cc.border.bottom)
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(DB_COLS))}{DB_R0 + n - 1}"
    _prettify(ws, f"A{DB_R0}:{get_column_letter(len(DB_COLS))}{DB_R0 + n - 1}")
    ws.row_dimensions[4].height = 30
    # --- heat: 511 rows of signed percentages are unreadable as bare numbers.
    # A three-colour scale anchored at ZERO (not at the min/max, which would
    # paint the least-bad loss green) turns the performance band into a picture
    # you can scan. Display-only: no value or formula is touched.
    last = DB_R0 + n - 1
    for fld in ("first_day_return_pct", "day1_open_pop_pct", "day1_open_close_pct",
                "ret_1w_pct", "ret_1m_pct", "ret_3m_pct",
                "aftermkt_1w_pct", "aftermkt_1m_pct", "aftermkt_3m_pct",
                "alpha_1m_pct", "alpha_1m_expop_pct", "since_ipo_pct",
                "a_premium_ipo_pct", "a_premium_now"):
        try:
            L = db_col(fld)
        except StopIteration:
            continue
        ws.conditional_formatting.add(
            f"{L}{DB_R0}:{L}{last}",
            ColorScaleRule(start_type="num", start_value=-60, start_color="00F8696B",
                           mid_type="num", mid_value=0, mid_color="00FFFFFF",
                           end_type="num", end_value=60, end_color="0063BE7B"))
    # magnitude bars where the question is "how big", not "how good"
    for fld, colr in (("deal_size_hkdm", "004472C4"), ("mktcap_ipo_hkdm", "008FAADC"),
                      ("oversub_public_mult", "00ED7D31"), ("oversub_intl_mult", "00F4B183"),
                      ("cornerstone_pct", "00A9D08E")):
        try:
            L = db_col(fld)
        except StopIteration:
            continue
        ws.conditional_formatting.add(
            f"{L}{DB_R0}:{L}{last}",
            DataBarRule(start_type="min", end_type="max", color=colr, showValue=True))
    for r_ in range(DB_R0, last + 1):
        ws.row_dimensions[r_].height = 15
    _heat_first(ws)
    return ws


def _prettify(ws, data_ref):
    """The 'terminal grid' look: no worksheet gridlines, zebra striping on the
    data rows. Both are display-only — no value or formula is touched."""
    ws.sheet_view.showGridLines = False
    ws.conditional_formatting.add(
        data_ref, FormulaRule(formula=["MOD(ROW(),2)=0"], fill=F_ZEBRA,
                              stopIfTrue=False))


def _heat_first(ws):
    """Excel resolves overlapping conditional formats by priority, LOWEST first.

    The comp table paints whole rows by match quality, and those rules are
    written before the colour scales, so without this the row tint would win in
    the return columns and the heat would silently never appear. Renumber so
    scales and bars rank ahead of the row fills; the row tint still shows in
    every column a scale does not cover.
    """
    rules = [(rng, r) for rng in ws.conditional_formatting for r in rng.rules]
    heat = [x for x in rules if x[1].type in ("colorScale", "dataBar")]
    rest = [x for x in rules if x[1].type not in ("colorScale", "dataBar")]
    for i, (_rng, r) in enumerate(heat + rest, 1):
        r.priority = i


def db_col(col_field):
    """Column letter for a Database field — every formula resolves through this so
    inserting a column can never silently repoint a reference."""
    return get_column_letter(next(i for i, c in enumerate(DB_COLS, 1) if c[2] == col_field))


def db_range(col_field, n):
    L = db_col(col_field)
    return f"Database!${L}${DB_R0}:${L}${DB_R0 + n - 1}"


def define(wb, name, ref):
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def sheet_scores(wb, n, npipe):
    """The scoring engine, on its own visibly-named sheet — fully inspectable."""
    ws_db = wb.create_sheet("Calc (scoring engine)")
    put(ws_db, "A1", "SCORING ENGINE", TITLE)
    put(ws_db, "A2", "One row per Database deal. The Screener ranks by column H "
                     "(comp score vs the Screener target). "
                     "Weights are named cells on the Notes tab.", SUB)
    for j, h in enumerate(["Deal", "Subsector =", "Sector =", "Size pts", "Profit =",
                           "A/H =", "Recency", "Cornerstone", "P/E prox",
                           "Demand prox", "SCORE"], 1):
        hdr(ws_db, f"{get_column_letter(j)}4", h)
        ws_db.column_dimensions[get_column_letter(j)].width = 12 if j > 1 else 24
    # U1:Y1 = the target's first five cornerstone investors, split off its
    # Database list (T1 holds the list, T2:T5 the successive remainders). The
    # scratch block lives in T:Y — clear of BOTH the score columns (B:J) and
    # the pick-list mirror (K:R). A J1:J5 scratch collided with the H_SCORE
    # column and silently killed every comp formula; never share columns.
    inv_rng = db_range("cornerstone_keys", n)
    nm_rng = db_range("name", n)
    put(ws_db, "S1", "target cornerstone split →", NOTE)
    ws_db["T1"] = f'=IFERROR(INDEX({inv_rng},MATCH(Screener!$B$5,{nm_rng},0))&"","")'
    for k, first in enumerate("UVWXY"):
        row_src = f"$T${k+1}"
        ws_db[f"{first}1"] = (f'=TRIM(IFERROR(LEFT({row_src},FIND(";",{row_src}&";")-1),'
                              f'{row_src}))')
        if k < 4:
            ws_db[f"T{k+2}"] = f'=IFERROR(MID({row_src},FIND(";",{row_src})+1,9999),"")'
    t = "Screener!"  # target attribute cells
    c_sub, c_sec = db_col("subsector"), db_col("sector")
    c_size, c_prof = db_col("deal_size_hkdm"), db_col("profitable_at_ipo")
    c_ah, c_date, c_name = db_col("is_h_share"), db_col("ipo_date"), db_col("name")
    cname = db_col("name")
    for i in range(n):
        r = DB_R0 + i
        ws_db[f"A{r}"] = f"=Database!{cname}{r}"
        D = "Database!"
        # Target attributes come from the Screener's EFFECTIVE column D — what
        # was typed in the override column, else what the picked deal supplies.
        # Reading column B instead would silently ignore every manual entry.
        ws_db[f"{H_GATE}{r}"] = f"=IF(AND({t}$D$8<>\"\",{D}${c_sub}{r}={t}$D$8),1,0)"
        ws_db[f"{H_SEC}{r}"] = f"=IF(AND({t}$D$7<>\"\",{D}${c_sec}{r}={t}$D$7),1,0)"
        ws_db[f"{H_SIZE}{r}"] = (f"=IF(OR(N({D}${c_size}{r})<=0,N({t}$D$9)<=0),0,"
                                 f"MAX(0,1-ABS(LOG10(N({D}${c_size}{r})/{t}$D$9))/W_SIZEHW))")
        ws_db[f"{H_PROF}{r}"] = (f"=IF(OR({D}${c_prof}{r}=\"\",{t}$D$10=\"\"),0,"
                                 f"IF({D}${c_prof}{r}={t}$D$10,1,0))")
        ws_db[f"{H_AH}{r}"] = (f"=IF(OR({D}${c_ah}{r}=\"\",{t}$D$11=\"\"),0,"
                               f"IF({D}${c_ah}{r}={t}$D$11,1,0))")
        # N() forces a number: an unmatched pick leaves D13 as text and the
        # subtraction then poisons every score with #VALUE!
        ws_db[f"{H_REC}{r}"] = (f"=IF(OR({D}${c_date}{r}=\"\",N({t}$D$12)<=0),0,"
                                f"MAX(0,1-(N({t}$D$12)-{D}${c_date}{r})/W_RECD))")
        # HOW MANY cornerstone names this deal shares with the target: the
        # target's first five investors sit in K1:O1 (split from its Database
        # list), plus whatever was typed in the Screener's investor box
        c_inv = db_col("cornerstone_keys")
        parts = [f"IF($U$1=\"\",0,N(ISNUMBER(SEARCH($U$1,{D}${c_inv}{r}&\"\"))))"]
        for col in ("V", "W", "X", "Y"):
            parts.append(f"IF(${col}$1=\"\",0,N(ISNUMBER(SEARCH(${col}$1,{D}${c_inv}{r}&\"\"))))")
        parts.append(f"IF({t}$B$17=\"\",0,N(ISNUMBER(SEARCH(LOWER({t}$B$17),{D}${c_inv}{r}&\"\"))))")
        ws_db[f"{H_CS}{r}"] = "=" + "+".join(parts)
        c_pe = db_col("pe_ipo")
        # valuation proximity: same log-distance shape as size; loss-makers
        # (blank P/E) score 0 — never a free match
        # N() on the DATABASE cell too: a blank P/E holds the Bloomberg-fallback
        # TEXT, and LOG10(text/x) is #VALUE! — which took out the whole comp
        # table whenever the target carried a P/E (the "pipeline pick breaks
        # the screener" bug). Text scores 0, exactly like a true blank.
        ws_db[f"{H_PE}{r}"] = (f"=IF(OR(N({D}${c_pe}{r})<=0,N({t}$D$13)<=0),0,"
                               f"MAX(0,1-ABS(LOG10(N({D}${c_pe}{r})/{t}$D$13))/W_PEHW))")
        # demand proximity: how close this deal's public subscription was to the
        # target's — the factor that actually tracks a similar debut.
        # NB: named c_dem, NOT c_sub — reusing c_sub here clobbered the
        # SUBSECTOR column for every row after the first, so the gate compared
        # subsector names against the public-subscription column and the
        # screener silently lost subsector-first ranking beyond row 5.
        c_dem = db_col("oversub_public_mult")
        ws_db[f"{H_DEM}{r}"] = (
            f"=IF(OR(N({D}${c_dem}{r})<=0,N({t}$D$14)<=0),0,"
            f"MAX(0,1-ABS(LOG10({D}${c_dem}{r}/{t}$D$14))/0.6))")
        # base score; then the A-share filter is a hard gate, and "rank by
        # cornerstone overlap" makes the shared count dominate
        base = (f"W_SUB*{H_GATE}{r}+W_SEC*{H_SEC}{r}*(1-{H_GATE}{r})+W_SIZE*{H_SIZE}{r}"
                f"+W_PROF*{H_PROF}{r}+W_AH*{H_AH}{r}+W_REC*{H_REC}{r}"
                f"+W_CS*{H_CS}{r}+W_PE*{H_PE}{r}")
        # test the FLAG, never the code cell: a_share_code is a BBG-fallback
        # FORMULA whose off-terminal text ("not filed — run on terminal") made
        # every deal read as an A/H pair once already (the v14 lesson)
        has_a = f"({D}${db_col('is_h_share')}{r}=\"Y\")"
        gate = (f"IF(OR({t}$B$18=\"All\",AND({t}$B$18=\"With A-share\",{has_a}),"
                f"AND({t}$B$18=\"Without A-share\",NOT({has_a}))),1,0)")
        # FORCE-INCLUDE (B19): codes typed comma-separated pin those deals to
        # the top of the comp table, past every filter — the user's judgment
        # outranks the score. The target itself still never appears.
        c_code = db_col("code")
        forced = (f"ISNUMBER(SEARCH(\",\"&{D}${c_code}{r}&\",\","
                  f"\",\"&SUBSTITUTE({t}$D$17,\" \",\"\")&\",\"))")
        ws_db[f"{H_SCORE}{r}"] = (
            f"=IF({D}${c_name}{r}={t}$B$5,-999999,"
            f"IF({forced},2000000,"
            f"IF({gate}=0,-999999,"
            f"IF({t}$B$16=\"cornerstone overlap first\",{H_CS}{r}*100000,"
            f"IF({t}$B$16=\"demand-similar first\",{H_DEM}{r}*100000,0))"
            f"+{base})))+ROW()/1000000")



# ---------------------------------------------------------------- Screener ---
# --- comp-table contract: ONE definition, shared by the builder and
# ipo_lib/audit_formulas.py. The screener once ranked on the wrong column for
# 510 of 511 rows because nothing cross-checked the header against the field
# its formula actually pulled; the auditor now does, off these constants.
# --- comp-table contract, GENERATED from DB_COLS (v15) -----------------------
# The user's rule: "the screener should have whatever we got in database."
# Hand-curated lists drifted every time the Database gained a column, so the
# comp table now mirrors DB_COLS mechanically: same headers, same order, same
# number formats, same band names — minus identity/plumbing columns that make
# no sense per-comp — then appends the scoring block. audit_formulas checks the
# result against the same source, so a Database column can no longer be absent
# from the screener by accident.
COMP_SKIP = {
    # "code" used to be skipped ("the Name cell carries it") — but the Excel
    # Name cell carries the NAME ONLY, and the desk asked for the code
    "name", "name_cn",           # Name is the row label already
    "cornerstone_keys",          # internal matching key, not analysis
    "prospectus_link", "allotment_link", "stabilization_link",   # links
    "_blank_notes", "valuation_notes",                           # prose
    "benchmark", "price_asof", "mktcap_basis", "size_basis",     # provenance prose
    "sponsors_en", "underwriters_en",     # the sponsor/bookrunner pair suffices
}
COMP_COLS = ["Rank", "Name"]
COMP_PULL = {"Name": "name"}
COMP_FMT = {}
_bands = [("IDENTITY", 2)]
for _b, _h, _f, _fmt, _w in DB_COLS:
    if _f in COMP_SKIP or _f.startswith("_blank"):
        continue
    _hdr = {"Cornerstone (% of offer)": "Cornerstone %"}.get(_h, _h)
    COMP_COLS.append(_hdr)
    COMP_PULL[_hdr] = _f
    COMP_FMT[_hdr] = _fmt
    if _bands[-1][0] == _b:
        _bands[-1] = (_b, _bands[-1][1] + 1)
    else:
        _bands.append((_b, 1))
# The two LIVE Bloomberg multiples are valuation, so they belong INSIDE the
# FUNDAMENTALS band next to P/E at IPO — appending them to the end stranded
# them beside Match/Score, where they read as scoring inputs rather than as
# what they are (today's multiple against the one the deal priced at).
_LIVE_MULT = [("P/E now (x)", "pe_now", "0.0"),
              ("P/S now (x)", "ps_ipo", "0.0")]     # live BDP leads; at-IPO is the fallback
_ins, _pos = None, 0
for _i, (_bn, _cnt) in enumerate(_bands):
    _pos += _cnt
    if _bn == "FUNDAMENTALS":
        _ins = _pos                                  # right after the last fundamentals column
        _bands[_i] = (_bn, _cnt + len(_LIVE_MULT))
        break
if _ins is None:                                     # no fundamentals band — keep old behaviour
    _ins = len(COMP_COLS)
    _bands[-1] = (_bands[-1][0], _bands[-1][1] + len(_LIVE_MULT))
for _k, (_h, _f, _fmt) in enumerate(_LIVE_MULT):
    COMP_COLS.insert(_ins + _k, _h)
    COMP_PULL[_h] = _f
    COMP_FMT[_h] = _fmt
COMP_COLS += ["Shared CS", "Match", "Score", "why"]
COMP_BANDS_GEN = _bands + [("SCORING", 4)]
# signed-percent columns get the return heat; magnitude bars stay hand-picked
COMP_RET_COLS = {h for h in COMP_COLS
                 if isinstance(COMP_FMT.get(h), str) and COMP_FMT[h].startswith('+0.0')}

def sheet_screener(wb, deals, pipe, cfg, n):
    ws = wb.create_sheet("Screener")
    put(ws, "A1", "COMPS SCREENER", TITLE)
    put(ws, "A2", "Pick a deal (blue). Comps are ranked SUBSECTOR FIRST — a same-subsector deal always "
                  "beats a same-sector one. The score for every comp is shown in full on the right, so "
                  "nothing is hidden.", SUB)

    # Combined pick-list (pipeline + every listed deal) lives on the VISIBLE
    # Calc sheet, columns K:Q — not on the Screener. v4 died when data columns
    # grew into hidden helpers, and v7's wider comp table (28 columns, ending
    # at AB) did exactly the same to a mirror parked at AA:AG: the 'why' column
    # overwrote the pick-list and every dependent formula died. Helper ranges
    # never share a sheet with a growing table again.
    wsc = wb["Calc (scoring engine)"]
    MIR0 = 3            # data starts at K3 (row 1 holds the target-split cells)
    for j, h in enumerate(["name", "sector", "subsector", "size HK$m",
                           "profitable", "H-share", "date"]):
        put(wsc, f"{get_column_letter(27 + j)}2", h, NOTE)
    npipe = len(pipe)
    for k in range(npipe):  # pipeline first: they're the usual targets
        r = MIR0 + k
        pr = 6 + k  # Pipeline data rows start at 6
        # Pipeline letters after the Size-basis insert: F=size, K=timing,
        # L=profitable, M=H-share — derive nothing by hand next time, read them
        # from PIPE_COLS so an inserted column cannot silently skew the mirror
        pletter = {f: get_column_letter(i + 1) for i, (_h, f, *_x) in enumerate(PIPE_COLS)}
        for col, pfield in (("AA", "name"), ("AB", "sector"), ("AC", "subsector"),
                            ("AD", "expected_size_hkdm"), ("AE", "profitable_at_ipo"),
                            ("AF", "is_h_share"), ("AG", "expected_timing"),
                            # a pipeline target's P/E = the prospectus expected
                            # multiple, so P/E-proximity scoring works pre-listing
                            ("AH", "pe_expected_mid")):
            wsc[f"{col}{r}"] = f"=Pipeline!{pletter[pfield]}{pr}"
    # THREE DISJOINT BLOCKS on Calc, and they must stay that way:
    #   B:K  score components   ·   T:Y  cornerstone split   ·   AA:AH  pick-list
    # The score block grew into a K-based mirror once and silently blanked every
    # comp; keep new helpers to their own letters.
    mirror = [("AA", "name"), ("AB", "sector"), ("AC", "subsector"),
              ("AD", "deal_size_hkdm"), ("AE", "profitable_at_ipo"),
              ("AF", "is_h_share"), ("AG", "ipo_date"), ("AH", "pe_ipo")]
    for i in range(n):
        r = MIR0 + npipe + i
        dr = DB_R0 + i
        for col, field in mirror:
            wsc[f"{col}{r}"] = f"=Database!{db_col(field)}{dr}"
    last = MIR0 - 1 + npipe + n
    define(wb, "AllDeals", f"{CALC}!$AA${MIR0}:$AA${last}")

    # ------------------------------------------------ 1. pick, or just type ---
    # One tab does both jobs: pick something already in the book, or describe a
    # deal that is not in it yet. Anything typed in the OVERRIDE column wins, so
    # a pipeline name with no terms becomes a full target by typing two cells.
    #
    # LAYOUT CONTRACT (v13) — column B means ONE thing per block, which is why
    # the card used to read as a jumble:
    #   rows 7-14  THE TARGET GRID   B = from the pick · C = you type · D = engine
    #   rows 15-18 HOW TO RANK       B = you type (there is no "from the pick")
    # Row numbers are named below; every formula elsewhere refers to them, so a
    # future re-layout is a one-line change, not a hunt through 45 addresses.
    put(ws, "A4", "TARGET — pick it, or type your own terms", HDRF, F_HDR,
        border=BOX, align=C_MID)
    ws.merge_cells("A4:E4")
    put(ws, "A5", "Pick a deal (pipeline or listed)", BODY, border=BOX)
    inp(ws, "B5", pipe[0].get("name") if pipe else deals[-1].get("name"))
    dv = DataValidation(type="list", formula1="=AllDeals", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B5"])

    hdr(ws, "A6", "TARGET ATTRIBUTE"); hdr(ws, "B6", "From the pick")
    hdr(ws, "C6", "TYPE TO OVERRIDE"); hdr(ws, "D6", "Engine uses"); hdr(ws, "E6", "what it does")
    # (label, Calc mirror column, number format, hint)
    attrs = [("Sector", "AB", None, ""),
             ("Subsector", "AC", None, "THE GATE — same-subsector comps rank first"),
             ("Size (HK$m)", "AD", MONEY, ""),
             ("Profitable (Y/N)", "AE", None, ""),
             ("H-share (Y/N)", "AF", None, ""),
             ]
    for k, (label, col, fmt, hint) in enumerate(attrs):
        r = 7 + k
        put(ws, f"A{r}", label, BODY, border=BOX)
        calc(ws, f"B{r}", f'=IFERROR(INDEX({CALC}!${col}${MIR0}:${col}${last},'
                          f'MATCH($B$5,AllDeals,0)),"")', fmt)
        inp(ws, f"C{r}", None, fmt)
        # the effective value: typed beats picked, always
        calc(ws, f"D{r}", f'=IF($C{r}<>"",$C{r},$B{r})', fmt)
        if hint:
            put(ws, f"E{r}", hint, NOTE)
    put(ws, "A12", "IPO / expected date", BODY, border=BOX)
    calc(ws, "B12", f'=IFERROR(INDEX({CALC}!$AG${MIR0}:$AG${last},'
                    f'MATCH($B$5,AllDeals,0)),"")', DATEF)
    inp(ws, "C12", None, DATEF)
    calc(ws, "D12", '=IF($C12<>"",$C12,IF(ISNUMBER($B12),$B12,TODAY()))', DATEF)
    put(ws, "A13", "Target P/E (x)", BODY, border=BOX)
    calc(ws, "B13", f'=IFERROR(INDEX({CALC}!$AH${MIR0}:$AH${last},'
                    f'MATCH($B$5,AllDeals,0)),"")', "0.0")
    inp(ws, "C13", None, "0.0")
    calc(ws, "D13", '=IF($C13<>"",$C13,$B13)', "0.0")
    put(ws, "E13", "comps near this multiple rank higher", NOTE)
    put(ws, "A14", "Public sub (x), expected", BODY, border=BOX)
    calc(ws, "B14", f'=IFERROR(INDEX({db_range("oversub_public_mult", n)},'
                    f'MATCH($B$5,{db_range("name", n)},0)),"")', "#,##0.0")
    inp(ws, "C14", None, "#,##0.0")
    calc(ws, "D14", '=IF($C14<>"",$C14,$B14)', "#,##0.0")
    put(ws, "E14", "drives 'demand-similar first'", NOTE)

    # ---- rows 15-18: pure controls. B is ALWAYS yours to type here ----------
    hdr(ws, "A15", "HOW TO RANK & FILTER"); hdr(ws, "B15", "You choose")
    hdr(ws, "C15", ""); hdr(ws, "D15", ""); hdr(ws, "E15", "what it does")
    put(ws, "A16", "Rank by", BODY, border=BOX)
    inp(ws, "B16", "standard")
    dvm = DataValidation(type="list",
                         formula1='"standard,cornerstone overlap first,demand-similar first"',
                         allow_blank=False)
    ws.add_data_validation(dvm)
    dvm.add(ws["B16"])
    put(ws, "E16", "standard = subsector-first · the other two are explained below", NOTE)
    put(ws, "A17", "Cornerstone investor to match", BODY, border=BOX)
    inp(ws, "B17", None)
    put(ws, "E17", "type e.g. hillhouse — adds one key to 'Shared CS'", NOTE)
    put(ws, "A18", "A-share filter", BODY, border=BOX)
    inp(ws, "B18", "All")
    dva = DataValidation(type="list", formula1='"All,With A-share,Without A-share"',
                         allow_blank=False)
    ws.add_data_validation(dva)
    dva.add(ws["B18"])
    put(ws, "C18", "Comps to show", BODY, border=BOX)
    inp(ws, "D18", 10, "0")
    put(ws, "E18", "screen A+H only / non-A only · 3-15 comps", NOTE)
    # NB: row 19 is the comp table's BAND row — a control written there is
    # silently overwritten. The card's second control column (C/D) is free on
    # row 17, and audit_formulas now asserts this label so it cannot drift.
    put(ws, "C17", "Force-include", BODY, border=BOX)
    inp(ws, "D17", None)
    put(ws, "F17", "codes as in the Database, comma-separated (9888, 2015) — pinned "
                   "to the top of the comps, past every filter", NOTE)

    # what the engine is actually using, in words
    calc(ws, "G5", '=IF(AND($D$8="",N($D$9)<=0),'
                   '"Ranking on profile and recency only — no subsector, no size. '
                   'Type either one in the TYPE TO OVERRIDE column and the comps re-rank as you type.",'
                   'IF($D$8="","No subsector — ranking on size, profile and recency. '
                   'Pick one in the override column to rank subsector-first.",'
                   'IF(N($D$9)<=0,"No deal size yet (normal for a PHIP-stage applicant) — '
                   'ranking on subsector, profile and recency. Type the expected HK$m size in the '
                   'override column to add size proximity.",'
                   '"Ranking subsector-first on "&$D$8&", size HK$"&TEXT($D$9,"#,##0")&"m.")))')

    put(ws, "G7", "SIZE BENCHMARK", SECT)
    put(ws, "G8", "Percentile vs 2021-26 deals", BODY, border=BOX)
    calc(ws, "H8", f'=IF(N($D$9)<=0,"n/a",COUNTIF({db_range("deal_size_hkdm", n)},"<"&$D$9)'
                   f'/COUNT({db_range("deal_size_hkdm", n)}))', "0%")
    put(ws, "G9", "Bucket", BODY, border=BOX)
    bkts = cfg["size_buckets_hkdm"]
    calc(ws, "H9", f"=IF(N($D$9)<=0,\"size not set\",IF($D$9>={bkts[0]['min']},\"{bkts[0]['label']}\","
                   f"IF($D$9>={bkts[1]['min']},\"{bkts[1]['label']}\","
                   f"IF($D$9>={bkts[2]['min']},\"{bkts[2]['label']}\",\"{bkts[3]['label']}\"))))")

    # ------------------------------------- 2. what the book knows about it ----
    put(ws, "G11", "WHAT THE DATABASE HOLDS ON THIS DEAL", SECT)
    known = [("Day-1", "first_day_return_pct", '+0.0"%";-0.0"%"'),
             ("1-month", "ret_1m_pct", '+0.0"%";-0.0"%"'),
             ("Public sub (x)", "oversub_public_mult", "#,##0.0"),
             ("Cornerstone %", "cornerstone_pct", PCT),
             ("Sponsors", "sponsors", None)]
    for k, (label, field, fmt) in enumerate(known):
        r = 12 + k
        put(ws, f"G{r}", label, BODY, border=BOX)
        calc(ws, f"H{r}",
             f'=IFERROR(IF(INDEX({db_range(field, n)},MATCH($B$5,{db_range("name", n)},0))="",'
             f'"—",INDEX({db_range(field, n)},MATCH($B$5,{db_range("name", n)},0))),'
             f'"not listed yet")', fmt)

    # -------------------------------------------------------- 3. the comps ---
    comp_cols = COMP_COLS
    # band row generated with the columns — same bands as the Database itself
    col0 = 1
    for band, width in COMP_BANDS_GEN:
        put(ws, f"{get_column_letter(col0)}19", band, HDRF, F_HDR, border=BOX, align=C_HDR)
        if width > 1:
            ws.merge_cells(start_row=19, start_column=col0, end_row=19,
                           end_column=col0 + width - 1)
        col0 += width
    ws.row_dimensions[19].height = 18
    ws.row_dimensions[20].height = 26
    for j, h in enumerate(comp_cols):
        expop = "ex-pop" in h
        hdr(ws, f"{get_column_letter(1 + j)}20", h,
            fill=PatternFill("solid", fgColor="001F7A6D" if expop else "00305496"))
    score_rng = f"{CALC}!${H_SCORE}${DB_R0}:${H_SCORE}${DB_R0 + n - 1}"
    pull = COMP_PULL
    RET_COLS = COMP_RET_COLS
    MAXCOMP = 15
    R_COMP = 21                                   # first comp row
    gate_rng = f"{CALC}!${H_GATE}${DB_R0}:${H_GATE}${DB_R0 + n - 1}"
    cs_rng = f"{CALC}!${H_CS}${DB_R0}:${H_CS}${DB_R0 + n - 1}"
    C_CS = get_column_letter(comp_cols.index("Shared CS") + 1)      # W
    C_MTCH = get_column_letter(comp_cols.index("Match") + 1)        # X
    C_SCORE = get_column_letter(comp_cols.index("Score") + 1)       # Y
    C_WHY = get_column_letter(comp_cols.index("why") + 1)           # Z
    LASTC = C_WHY
    for k in range(MAXCOMP):
        r = R_COMP + k
        # respects the N selector AND stops at the last un-gated deal — with a
        # hard filter on (e.g. "With A-share") the -999999 gated scores would
        # otherwise leak into the tail slots as phantom comps
        shown = (f'IF(OR({k + 1}>MAX(3,MIN(15,N($D$18))),'
                 f'IFERROR(LARGE({score_rng},{k + 1}),-999999)<=-900000),"",')
        put(ws, f"A{r}", None, BODY, border=BOX, align=C_MID)
        calc(ws, f"A{r}", f'={shown}{k + 1})')
        mtch = f"MATCH(LARGE({score_rng},{k+1}),{score_rng},0)"
        for j, h in enumerate(comp_cols[1:-4], 2):   # Name .. P/S now
            fld = pull[h]
            rng = db_range(fld, n)
            # number format travels WITH the Database column definition
            fmt = COMP_FMT.get(h)
            if h == "P/E now (x)":
                cno = db_range("code", n)
                body = (f'IFERROR(BDP(TEXT(INDEX({cno},{mtch}),"0")&" HK Equity",'
                        f'"PE_RATIO"),'
                        f'IF(INDEX({rng},{mtch})="","—",INDEX({rng},{mtch})))')
            elif h == "P/S now (x)":
                # live P/S: the desk's PX_TO_SALES_RATIO formula, quoted
                # properly; off-terminal it falls back to the at-IPO P/S
                cno = db_range("code", n)
                body = (f'IFERROR(BDP(TEXT(INDEX({cno},{mtch}),"0")&" HK Equity",'
                        f'"PX_TO_SALES_RATIO"),'
                        f'IF(INDEX({rng},{mtch})="","—",INDEX({rng},{mtch})))')
            elif fld == "pe_ipo":
                # a loss-maker's P/E reads n/m, never a dash that looks unpulled
                nm = db_range("profitable_at_ipo", n)
                body = (f'IF(N(INDEX({rng},{mtch}))<=0,'
                        f'IF(INDEX({nm},{mtch})="N","n/m","—"),INDEX({rng},{mtch}))')
            else:
                body = f'IF(INDEX({rng},{mtch})="","—",INDEX({rng},{mtch}))'
            calc(ws, f"{get_column_letter(j)}{r}", f'={shown}IFERROR({body},""))', fmt)
        calc(ws, f"{C_CS}{r}", f'={shown}IFERROR(IF(INDEX({cs_rng},{mtch})>0,'
                               f'INDEX({cs_rng},{mtch}),""),""))', "0")
        calc(ws, f"{C_MTCH}{r}", f'={shown}IF(IFERROR(INDEX({gate_rng},{mtch}),0)=1,'
                                 f'"same subsector","sector only"))')
        calc(ws, f"{C_SCORE}{r}", f'={shown}ROUND(IFERROR(LARGE({score_rng},{k+1}),0),0))', "0")
        # plain-English score breakdown, so the ranking is never a black box
        calc(ws, f"{C_WHY}{r}", f'={shown}IFERROR("sub "&INDEX({gate_rng},{mtch})*W_SUB'
                          f'&" + size "&ROUND(INDEX({CALC}!${H_SIZE}${DB_R0}:${H_SIZE}${DB_R0+n-1},{mtch})*W_SIZE,0)'
                          f'&" + prof "&INDEX({CALC}!${H_PROF}${DB_R0}:${H_PROF}${DB_R0+n-1},{mtch})*W_PROF'
                          f'&" + recency "&ROUND(INDEX({CALC}!${H_REC}${DB_R0}:${H_REC}${DB_R0+n-1},{mtch})*W_REC,0)'
                          f'&IF(INDEX({cs_rng},{mtch})>0," + shared CS x"&INDEX({cs_rng},{mtch}),""),""))')
        ws.conditional_formatting.add(
            f"A{r}:{LASTC}{r}", FormulaRule(formula=[f'${C_MTCH}{r}="same subsector"'], fill=F_GRN))
        ws.conditional_formatting.add(
            f"A{r}:{LASTC}{r}", FormulaRule(formula=[f'${C_MTCH}{r}="sector only"'], fill=F_AMB))
    # The match tint paints the row; the return columns get the same zero-anchored
    # heat as the Database so a comp table reads the same way the book does.
    for h in RET_COLS:
        if h not in comp_cols:
            continue
        L = get_column_letter(comp_cols.index(h) + 1)
        ws.conditional_formatting.add(
            f"{L}{R_COMP}:{L}{R_COMP + MAXCOMP - 1}",
            ColorScaleRule(start_type="num", start_value=-60, start_color="00F8696B",
                           mid_type="num", mid_value=0, mid_color="00FFFFFF",
                           end_type="num", end_value=60, end_color="0063BE7B"))
    for h in ("Deal size (HK$m)", "Public sub (x)", "Cornerstone %"):
        L = get_column_letter(comp_cols.index(h) + 1)
        ws.conditional_formatting.add(
            f"{L}{R_COMP}:{L}{R_COMP + MAXCOMP - 1}",
            DataBarRule(start_type="min", end_type="max", color="004472C4", showValue=True))
    mr = R_COMP + MAXCOMP + 1
    put(ws, f"A{mr}", "MEDIAN of shown comps", BOLD, border=BOX)
    # MEDIAN over every numeric column, derived from the shared contract —
    # dates and text columns are excluded by their own formats
    for h in comp_cols[2:-4]:
        fmtv = COMP_FMT.get(h)
        if not fmtv or fmtv == DATEF or fmtv == "@":
            continue
        col = get_column_letter(comp_cols.index(h) + 1)
        calc(ws, f"{col}{mr}",
             f'=IFERROR(MEDIAN({col}{R_COMP}:{col}{R_COMP+MAXCOMP-1}),"")', fmtv)
    put(ws, f"A{mr+1}", "P/E shows n/m for loss-makers (use P/S). P/E now + P/S now = live BDP on the "
                        "terminal, else the desk's pasted Bloomberg value. A-premium = A over H "
                        "(+ = A trades above H). Ex-pop columns start at the day-1 close. "
                        "A/H columns show N/A when the issuer has no A line; '—' = not on file. "
                        "Shared CS = cornerstone investors in common with the target.", NOTE)

    # ---------------------------------------- 4. target vs comp #1, in words --
    r0 = mr + 3
    put(ws, f"A{r0}", "TARGET vs CLOSEST COMP", SECT)
    hdr(ws, f"B{r0+1}", "Target"); hdr(ws, f"C{r0+1}", "Comp #1"); hdr(ws, f"D{r0+1}", "Median of shown")

    def _cc(h):
        return get_column_letter(comp_cols.index(h) + 1)
    card = [("Name", "$B$5", f"$B${R_COMP}", ""),
            ("Subsector", "$D$8", f"${_cc('Subsector')}${R_COMP}", ""),
            ("Size HK$m", "$D$9", f"${_cc('Deal size (HK$m)')}${R_COMP}", f"${_cc('Deal size (HK$m)')}${mr}"),
            ("P/E (x)", "", f"${_cc('P/E at IPO')}${R_COMP}", f"${_cc('P/E at IPO')}${mr}"),
            ("Public sub (x)", "", f"${_cc('Public sub (x)')}${R_COMP}", f"${_cc('Public sub (x)')}${mr}"),
            ("Day-1", "", f"${_cc('Day-1')}${R_COMP}", f"${_cc('Day-1')}${mr}"),
            ("1-month", "", f"${_cc('1-month')}${R_COMP}", f"${_cc('1-month')}${mr}"),
            ("3-month", "", f"${_cc('3-month')}${R_COMP}", f"${_cc('3-month')}${mr}")]
    for k, (label, tgt, c1, med) in enumerate(card):
        r = r0 + 2 + k
        put(ws, f"A{r}", label, BODY, border=BOX)
        calc(ws, f"B{r}", f"={tgt}" if tgt else '="—"')
        calc(ws, f"C{r}", f'=IFERROR({c1},"")')
        if med:
            calc(ws, f"D{r}", f'=IFERROR({med},"")')

    # ------------------------- 5. the A-share month, straight off Bloomberg --
    # The four native charts are gone (the HTML dashboard owns comparative
    # charting). What Excel alone can do is LIVE Bloomberg: when the picked
    # deal has an A line, a BDH spill pulls its last month of closes and the
    # line chart below draws it. Off the terminal the block explains itself.
    R_BDH = r0 + 2 + len(card) + 2
    put(ws, f"F{R_BDH}", "A-SHARE, LAST MONTH — live Bloomberg (fills on the terminal)", SECT)
    ahc_rng = db_range("a_share_code", n)
    nm_rng2 = db_range("name", n)
    # pipeline lookups for the same pick: a NEW deal (Ingenic) is not a
    # Database row, so the A-share chain must fall through to the Pipeline tab
    pl = {f: get_column_letter(i + 1) for i, (_h, f, *_x) in enumerate(PIPE_COLS)}
    p_rng = lambda f: f"Pipeline!${pl[f]}$6:${pl[f]}${5 + max(1, npipe)}"
    p_name = p_rng("name")
    calc(ws, f"F{R_BDH+1}",
         f'=IFERROR(INDEX({ahc_rng},MATCH($B$5,{nm_rng2},0)),'
         f'IFERROR(INDEX({p_rng("a_share_code")},MATCH($B$5,{p_name},0)),""))')
    put(ws, f"G{R_BDH+1}", "← A-share code of the pick — Database first, then the "
                           "Pipeline tab (so an offering-window deal resolves too)", NOTE)
    # Bloomberg ticker: 300223.SZ -> "300223 CH Equity"
    # IFERROR wrap: FIND on a code with no venue suffix must fail to "" —
    # and the offline evaluator computes both IF branches eagerly
    calc(ws, f"F{R_BDH+2}",
         f'=IFERROR(IF(F{R_BDH+1}="","",LEFT(F{R_BDH+1},FIND(".",F{R_BDH+1})-1)&" CH Equity"),"")')
    hdr(ws, f"F{R_BDH+3}", "Date"); hdr(ws, f"G{R_BDH+3}", "Close (CNY)")
    put(ws, f"F{R_BDH+4}",
        f'=IF($F${R_BDH+2}="","no A-share for this pick",'
        f'BDH($F${R_BDH+2},"PX_LAST",TEXT(TODAY()-31,"yyyymmdd"),TEXT(TODAY(),"yyyymmdd")))',
        BODY, F_CALC, border=BOX)
    # ---- A-SHARE LIVE PANEL (columns A:D, same rows as the BDH block) ------
    # Live BDP on the terminal; off the terminal every row falls back to the
    # desk-scraped Tencent quote captured in the Pipeline tab, so the panel is
    # never blank and never fake. A-premium convention: + = A above H.
    put(ws, f"A{R_BDH}", "A-SHARE LIVE PANEL — live BDP on the terminal, scraped quote off it", SECT)
    tick = f"$F${R_BDH+2}"
    h_terms = (f'IFERROR(INDEX({db_range("final_price", n)},MATCH($B$5,{nm_rng2},0)),'
               f'IFERROR(INDEX({p_rng("range_hi")},MATCH($B$5,{p_name},0)),""))')
    # HK ticker of the pick, built WITHOUT leading zeros (the terminal
    # rejects "0606 HK Equity"); Database first, then the Pipeline tab
    h_code = (f'IFERROR(INDEX({db_range("code", n)},MATCH($B$5,{nm_rng2},0)),'
              f'IFERROR(INDEX({p_rng("expected_code")},MATCH($B$5,{p_name},0)),""))')
    h_tick = f'IFERROR(TEXT(N({h_code}),"0")&" HK Equity","")'
    rows_live = [
        ("A price now (CNY)",
         f'=IF({tick}="","no A line",IFERROR(BDP({tick},"PX_LAST"),'
         f'IFERROR(INDEX({p_rng("a_price_now")},MATCH($B$5,{p_name},0)),"run on terminal")))',
         "0.00", ""),
        ("CNY→HKD", '=IFERROR(BDP("CNYHKD Curncy","PX_LAST"),1.10)', "0.0000", ""),
        ("A price (HK$)",
         '=IF(OR(NOT(ISNUMBER(A_PX)),N(A_FX)<=0),"n/a",A_PX*A_FX)', "0.00", ""),
        ("A P/E (TTM)",
         f'=IF({tick}="","n/a",IFERROR(BDP({tick},"PE_RATIO"),'
         f'IFERROR(INDEX({p_rng("a_pe_ttm")},MATCH($B$5,{p_name},0)),"run on terminal")))',
         "0.0", ""),
        ("H offer / cap (HK$)", f'={h_terms}', "0.00", ""),
        ("A premium vs H", '=IFERROR(IF(OR(N(A_HKD)<=0,N(H_TRM)<=0),"n/a",A_HKD/H_TRM-1),"n/a")',
         "+0.0%;-0.0%", "+ = A above the H terms"),
        # the PICK's own H line, live: the comp table carries P/E now and P/S
        # now per comp, but the desk reads the target here — so it needs both
        ("H P/E now (TTM)",
         f'=IF({h_tick}="","n/a",IFERROR(BDP({h_tick},"PE_RATIO"),'
         f'IFERROR(INDEX({db_range("pe_now", n)},MATCH($B$5,{nm_rng2},0)),"run on terminal")))',
         "0.0", "live BDP; falls back to the desk paste"),
        ("H P/S now",
         f'=IF({h_tick}="","n/a",IFERROR(BDP({h_tick},"PX_TO_SALES_RATIO"),'
         f'IFERROR(INDEX({db_range("ps_ipo", n)},MATCH($B$5,{nm_rng2},0)),"run on terminal")))',
         "0.0", "live BDP; falls back to P/S at IPO"),
        # the WHOLE COMPANY on the A line, not just the H tranche: Luxshare's
        # H slice is HK$24bn while the issuer is ~HK$490bn, and reading the
        # tranche as the company is the mistake this row exists to stop
        ("A-line company mkt cap (HK$m)",
         f'=IF({tick}="","no A line",IFERROR(BDP({tick},"CUR_MKT_CAP")*'
         f'IF(N(A_FX)>0,A_FX,1.10)/1000000,'
         f'IFERROR(INDEX({db_range("a_mktcap_now_hkdm", n)},MATCH($B$5,{nm_rng2},0)),'
         f'"run on terminal")))',
         "#,##0", "all share classes x the A price; scraped fallback"),
    ]
    for k, (lab, f_, fmt_, note_) in enumerate(rows_live):
        rr_ = R_BDH + 1 + k
        put(ws, f"A{rr_}", lab, BODY, border=BOX)
        calc(ws, f"B{rr_}", f_, fmt_)
        if note_:
            put(ws, f"C{rr_}", note_, NOTE)
    define(wb, "A_PX", f"Screener!$B${R_BDH+1}")
    define(wb, "A_FX", f"Screener!$B${R_BDH+2}")
    define(wb, "A_HKD", f"Screener!$B${R_BDH+3}")
    define(wb, "H_TRM", f"Screener!$B${R_BDH+5}")
    ch = LineChart()
    ch.title = "A-share close, last month (Bloomberg)"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.y_axis.title = "CNY"
    ch.x_axis.title = "session"
    ch.display_blanks = "gap"
    ch.height, ch.width = 8.5, 17
    ch.add_data(Reference(ws, min_col=7, min_row=R_BDH + 3, max_row=R_BDH + 27),
                titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=6, min_row=R_BDH + 4, max_row=R_BDH + 27))
    ws.add_chart(ch, f"J{R_BDH}")

    def col_of(h):
        return comp_cols.index(h) + 1

    # widths travel with the Database column definitions; scoring block by hand
    dbw = {h2: w2 for _b2, h2, _f2, _fm2, w2 in DB_COLS}
    dbw = {({"Cornerstone (% of offer)": "Cornerstone %"}.get(k, k)): v
           for k, v in dbw.items()}
    widths = dict({"Rank": 5, "Name": 24, "Code": 7, "P/E now (x)": 10, "P/S now (x)": 10, "Shared CS": 9,
                   "Match": 14, "Score": 8, "why": 46}, **{
        h2: dbw[h2] for h2 in comp_cols if h2 in dbw})
    for h, w in widths.items():
        if h in comp_cols:
            ws.column_dimensions[get_column_letter(col_of(h))].width = w
    ws.column_dimensions["A"].width = 27            # label column up top
    # control-card geometry: C is INPUT-only, E holds SHORT hints that overflow
    # into empty F — long explanations live in the HOW-TO card below, so no
    # text ever collides with the G-column info panels
    # boxed control card, soft fill — the eye finds the inputs immediately.
    # v12 reskin: zebra on the attribute rows, every override cell input-blue,
    # every effective cell calc-grey, so the three columns read as three roles.
    thin = Side(style="thin", color="00B8C4D0")
    zebra = PatternFill("solid", fgColor="00F2F6FB")
    for rr in range(4, 19):
        for cc in range(1, 6):
            c0 = ws.cell(row=rr, column=cc)
            if c0.fill.start_color.rgb in (None, "00000000") and not c0.value:
                c0.fill = zebra if (rr in (8, 10, 12, 14) and cc <= 5)                     else PatternFill("solid", fgColor="00FBFCFE")
            c0.border = Border(left=thin if cc == 1 else c0.border.left,
                               right=thin if cc == 5 else c0.border.right,
                               top=thin if rr == 4 else c0.border.top,
                               bottom=thin if rr == 18 else c0.border.bottom)
    ws.row_dimensions[4].height = 20
    for rr in range(5, 19):
        ws.row_dimensions[rr].height = 16

    # ---- HOW CORNERSTONE MATCHING WORKS — live, not a black box (pt 2) -----
    # Shows the ACTUAL keys the engine is comparing for the current pick, so
    # the mechanism explains itself with the user's own target. Sits under the
    # A-share live panel (rows 19+ belong to the comp table).
    # DERIVED, never a literal: this block used to start at a hardcoded
    # R_BDH+8 and silently ate the last live row when the panel grew (first
    # H P/S now, and it would have eaten the A-line cap row too) — the same
    # collision class that ate the force-include control. One row of air
    # after however many rows the panel actually has.
    csr = R_BDH + 1 + len(rows_live) + 1
    put(ws, f"A{csr}", "HOW CORNERSTONE MATCHING WORKS", SECT)
    put(ws, f"A{csr+1}", "Investor names are normalised to a key ('GIC Private Limited' → gic), "
                         "a comp scores +1 per key it shares, and 'Shared CS' shows the count.", NOTE)
    put(ws, f"A{csr+2}", "This pick's top-5 keys:", NOTE)
    calc(ws, f"D{csr+2}",
         f'=TRIM({CALC}!$U$1&" "&{CALC}!$V$1&" "&{CALC}!$W$1&" "&'
         f'{CALC}!$X$1&" "&{CALC}!$Y$1)')
    put(ws, f"A{csr+3}", "RANK MODES — standard: subsector first · cornerstone overlap first: "
                         "shared investors dominate · demand-similar first: closest "
                         "public-subscription level (the best debut predictor).", NOTE)
    for rr in range(csr, csr + 4):
        for cc in range(1, 6):
            c0 = ws.cell(row=rr, column=cc)
            c0.border = Border(left=thin if cc == 1 else c0.border.left,
                               right=thin if cc == 5 else c0.border.right,
                               top=thin if rr == csr else c0.border.top,
                               bottom=thin if rr == csr + 3 else c0.border.bottom)
    ws.row_dimensions[1].height = 26
    _heat_first(ws)
    ws.freeze_panes = "C21"
    _prettify(ws, f"A{R_COMP}:{LASTC}{R_COMP + MAXCOMP - 1}")
    return ws


# ---------------------------------------------------------------- Pipeline ---
PIPE_COLS = [("Code", "expected_code", None, 8), ("Name", "name", None, 24),
             ("Name (CN)", "name_cn", None, 12), ("Sector", "sector", None, 12),
             ("Subsector", "subsector", None, 22),
             ("Expected size (HK$m)", "expected_size_hkdm", MONEY, 13),
             ("Size lo (US$m)", "expected_size_lo_usdm", MONEY, 10),
             ("Size hi (US$m)", "expected_size_hi_usdm", MONEY, 10),
             ("Size basis", "expected_size_basis", None, 26),
             ("Range lo (HK$)", "range_lo", PX, 9),
             ("Max/cap (HK$)", "range_hi", PX, 9),
             ("Expected P/E", "pe_expected_mid", "0.0", 10),
             ("Cornerstone %", "cornerstone_pct", '0.0"%"', 11),
             ("Offer period", "offer_period", None, 20),
             ("Status", "status", None, 11), ("Expected timing", "expected_timing", None, 11),
             ("Profitable", "profitable_at_ipo", None, 8), ("H-share", "is_h_share", None, 7),
             ("A-share code", "a_share_code", None, 11),
             ("A px now", "a_price_now", PX, 9),
             ("A P/E (TTM)", "a_pe_ttm", "0.0", 9),
             ("H cap vs A", "h_cap_vs_a_pct", '+0.0"%";-0.0"%"', 10),
             ("A prem vs H cap", "a_prem_vs_hcap_pct", '+0.0"%";-0.0"%"', 12),
             ("Sponsors", "sponsors", None, 26),
             ("Prospectus", "doc_link", None, 11),
             ("Business", "business_desc", None, 46), ("Valuation notes", "valuation_notes", None, 40)]


def sheet_pipeline(wb, pipe):
    ws = wb.create_sheet("Pipeline")
    put(ws, "A1", "ACTIVE PIPELINE", TITLE)
    put(ws, "A2", "expected size/timing are INPUT-BLUE — update as terms firm up; screener reads columns D/E/F/J/K/L", SUB)
    put(ws, "A4", "", BODY)
    for j, (h, *_rest) in enumerate(PIPE_COLS, 1):
        hdr(ws, f"{get_column_letter(j)}5", h)
        ws.column_dimensions[get_column_letter(j)].width = PIPE_COLS[j - 1][3]
    fx = 7.8
    for i, d in enumerate(pipe):
        r = 6 + i
        if not d.get("expected_size_hkdm") and d.get("expected_size_hi_usdm"):
            mid = (d.get("expected_size_lo_usdm", d["expected_size_hi_usdm"]) +
                   d["expected_size_hi_usdm"]) / 2
            d["expected_size_hkdm"] = round(mid * fx)
        for j, (_h, f, fmt, _w) in enumerate(PIPE_COLS, 1):
            v = listify(d.get(f))
            if isinstance(v, bool):
                v = "Y" if v else "N"
            cell = f"{get_column_letter(j)}{r}"
            if f == "doc_link" and v:
                put(ws, cell, f'=HYPERLINK("{v}","filing")',
                    font=Font(name=ARIAL, size=10, color="000000FF", underline="single"))
            elif f in ("expected_size_hkdm", "expected_size_lo_usdm",
                       "expected_size_hi_usdm", "expected_timing", "status",
                       "subsector"):
                inp(ws, cell, v, fmt)
            else:
                put(ws, cell, v, fmt=fmt)
    # zebra + live-offering tint, same language as the Database
    _prettify(ws, f"A6:{get_column_letter(len(PIPE_COLS))}{5 + len(pipe)}")
    F_HOT = PatternFill("solid", fgColor="00FDE9D9")
    st_col = next(i for i, (_h, f, *_x) in enumerate(PIPE_COLS, 1) if f == "status")
    for i, d in enumerate(pipe):
        if "OFFERING" in str(d.get("status") or ""):
            for cc in range(1, len(PIPE_COLS) + 1):
                c0 = ws.cell(row=6 + i, column=cc)
                if c0.fill.start_color.rgb in (None, "00000000"):
                    c0.fill = F_HOT
        ws.row_dimensions[6 + i].height = 15
    dvp = DataValidation(type="list", formula1="=SubsectorList", allow_blank=True)
    ws.add_data_validation(dvp)
    dvp.add(f"E6:E{5 + len(pipe)}")
    put(ws, f"A{7 + len(pipe)}",
        f"BLUE = yours to edit. Type/adjust EXPECTED SIZE (HK$m) and SUBSECTOR here — "
        f"the Screener reads these cells LIVE, so comps re-rank the moment you type. "
        f"HKD size = midpoint of US$ range × {fx} where not explicitly reported.", NOTE)

    # ---- the LIVE HKEX application queue (auto-refreshed, not hand-picked) ----
    import json as _json
    r0 = 9 + len(pipe)
    put(ws, f"A{r0}", "LIVE HKEX APPLICATION QUEUE (from the exchange's own AP & PHIP feed)", SECT)
    put(ws, f"A{r0+1}", "PHIP posted = listing hearing CLEARED, typically days-to-weeks from launch. "
                        "Refreshed every time you run refresh.", SUB)
    try:
        phip = _json.loads((ROOT / "data" / "batches" / "phip_pipeline.json").read_text())
        apps = phip.get("applications", [])
    except Exception:
        apps = []
    for j, h in enumerate(["Applicant", "Stage", "First filed", "Latest filing",
                           "Days in process", "Document"], 1):
        hdr(ws, f"{get_column_letter(j)}{r0+2}", h)
    shown = [a for a in apps if a.get("has_phip")] +             [a for a in apps if not a.get("has_phip")][:30]
    for i, a in enumerate(shown):
        r = r0 + 3 + i
        stage = a.get("stage", "")
        if a.get("sponsor_terminated"):
            stage += "  [SPONSOR TERMINATED]"
        put(ws, f"A{r}", a.get("applicant"), border=BOX,
            font=BOLD if a.get("has_phip") else BODY,
            fill=F_GRN if a.get("has_phip") else None)
        put(ws, f"B{r}", stage, border=BOX,
            font=WARN if a.get("sponsor_terminated") else BODY)
        put(ws, f"C{r}", a.get("first_filing"), border=BOX)
        put(ws, f"D{r}", a.get("latest_submission"), border=BOX)
        put(ws, f"E{r}", a.get("days_in_process"), fmt="0", border=BOX)
        if a.get("doc_link"):
            put(ws, f"F{r}", f'=HYPERLINK("{a["doc_link"]}","open filing")',
                font=Font(name=ARIAL, size=10, color="000000FF", underline="single"))
    put(ws, f"A{r0 + 4 + len(shown)}",
        f"{len([a for a in apps if a.get('has_phip')])} PHIP-stage of {len(apps)} live applications; "
        f"showing PHIPs plus the 30 most recent Application Proofs.", NOTE)
    return ws


# ---------------------------------------------------------------- AH tab -----
def sheet_ah(wb, deals, snap_date):
    ws = wb.create_sheet("AH")
    pairs = [d for d in deals if d.get("a_share_code")]
    put(ws, "A1", "A/H SPREAD MONITOR", TITLE)
    put(ws, "A2", "A premium = (A×CNY→HKD) ÷ H − 1 · + = A trades ABOVE H · LIVE columns are Bloomberg BDP "
                  "(work on the terminal only — mnemonics await terminal verify) · BLUE overrides beat live · "
                  f"snapshot column as of {snap_date}", SUB)
    put(ws, "A4", "CNYHKD", BOLD, border=BOX)
    put(ws, "B4", '=IFERROR(BDP("CNYHKD Curncy","PX_LAST"),"")', BODY, fill=F_CALC, border=BOX)
    inp(ws, "C4", None)
    calc(ws, "D4", '=IF($C$4<>"",$C$4,$B$4)', "0.0000")
    put(ws, "E4", "live / override / effective", NOTE)
    headers = ["H code", "Name", "Subsector", "H live (BDP)", "H override", "H eff",
               "A code", "A live (BDP)", "A override", "A eff", "A premium (live)",
               "Snapshot A prem", "Note"]
    for j, h in enumerate(headers, 1):
        hdr(ws, f"{get_column_letter(j)}6", h)
    for i, d in enumerate(pairs):
        r = 7 + i
        acode_num = str(d["a_share_code"]).split(".")[0]
        put(ws, f"A{r}", d["code"], fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", d.get("subsector"), border=BOX)
        put(ws, f"D{r}", f'=IFERROR(BDP("{int(d["code"])} HK Equity","PX_LAST"),"")',
            BODY, F_CALC, PX, BOX)
        inp(ws, f"E{r}", None, PX)
        calc(ws, f"F{r}", f'=IF($E{r}<>"",$E{r},$D{r})', PX)
        put(ws, f"G{r}", d["a_share_code"], border=BOX)
        put(ws, f"H{r}", f'=IFERROR(BDP("{acode_num} CH Equity","PX_LAST"),"")',
            BODY, F_CALC, PX, BOX)
        inp(ws, f"I{r}", None, PX)
        calc(ws, f"J{r}", f'=IF($I{r}<>"",$I{r},$H{r})', PX)
        calc(ws, f"K{r}", f'=IF(OR(N($F{r})<=0,N($J{r})<=0,N($D$4)<=0),"",($J{r}*$D$4)/$F{r}-1)', "+0.0%;-0.0%")
        put(ws, f"L{r}", d.get("a_premium_now"), fmt="+0.0%;-0.0%", border=BOX)
        put(ws, f"M{r}", d.get("ah_note"), NOTE)
    if pairs:
        rng = f"K7:K{6 + len(pairs)}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=["ABS(K7)>0.3"], fill=F_AMB))
    r0 = 9 + len(pairs)
    put(ws, f"A{r0}", "NO A-SHARE LINE — closest pure-A proxies", SECT)
    proxies = [d for d in deals if d.get("a_share_proxy")]
    for j, h in enumerate(["H code", "Name", "Proxy A code", "Proxy name", "Rationale"], 1):
        hdr(ws, f"{get_column_letter(j)}{r0 + 1}", h)
    for i, d in enumerate(proxies):
        r = r0 + 2 + i
        px = d["a_share_proxy"]
        put(ws, f"A{r}", d["code"], fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", px.get("code") if isinstance(px, dict) else px, border=BOX)
        put(ws, f"D{r}", px.get("name") if isinstance(px, dict) else "", border=BOX)
        put(ws, f"E{r}", px.get("rationale") if isinstance(px, dict) else d.get("ah_note"), NOTE)
    for c, w in (("A", 8), ("B", 24), ("C", 20), ("D", 11), ("E", 10), ("F", 9),
                 ("G", 11), ("H", 10), ("I", 10), ("J", 9), ("K", 11), ("L", 11), ("M", 34)):
        ws.column_dimensions[c].width = w
    return ws


# ------------------------------------------------------------ Verify (BBG) ---
def sheet_verify_changes(wb, deals, n):
    """Everything the v20/v21 audits CHANGED, with a Bloomberg call beside it.

    The desk asked to verify on the terminal. This tab is the whole list on
    one screen: our value, the independent Bloomberg value, and a verdict
    cell that reads MATCH or CHECK by itself when the formulas resolve. Every
    row states what we relied on, so a disagreement can be argued rather than
    guessed at.
    """
    import json as _j
    ws = wb.create_sheet("Verify (BBG)")
    by_code = {d["code"]: d for d in deals}

    from corp_actions import load_actions, load_entitlements, bbg_factor
    acts0 = load_actions(ROOT)
    ent0 = load_entitlements(ROOT)

    def cum_factor(code):
        # what separates OUR raw print from BLOOMBERG's adjusted one: the
        # price-scale actions AND the entitlement (TERP) adjustments
        return bbg_factor(code, acts0, ent0)

    put(ws, "A1", "VERIFY ON THE TERMINAL — every value the v20/v21 audits changed", TITLE)
    put(ws, "A2", "Open on Bloomberg and let the BDP/BDH columns resolve. VERDICT says MATCH when "
                  "Bloomberg agrees, CHECK when it does not — with both numbers side by side. "
                  "Off-terminal every Bloomberg cell reads 'run on terminal', never an error. "
                  "Grey rows = mnemonic still AWAITING first-use verification.", SUB)
    r = 4

    # ---- A. listing date + day-1, measured on the local session list -------
    est = {}
    p = ROOT / "data" / "batches" / "hkex_allotments.json"
    if p.exists():
        est = {x["code"]: (x.get("ipo_date_est") or "")[:10]
               for x in _j.loads(p.read_text())["deals"] if isinstance(x, dict)}
    changed_date = [d for d in deals
                    if est.get(d["code"]) and est[d["code"]] != (d.get("ipo_date") or "")[:10]]
    put(ws, f"A{r}", f"A · LISTING DATE + DAY-1 CLOSE  ({len(changed_date)} deals — the stored date "
                     f"was the allotment-announcement day, usually a Saturday)", SECT)
    r += 1
    for j, h in enumerate(["Code", "Name", "Our listing date", "Our day-1 close",
                           "Our day-1 %", "BBG close that day", "VERDICT",
                           "What we relied on"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    r += 1
    for d in changed_date:
        c = d["code"]
        put(ws, f"A{r}", c, fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", d.get("ipo_date"), fmt=DATEF, border=BOX)
        d1c = (d.get("final_price") or 0) * (1 + (d.get("first_day_return_pct") or 0) / 100)
        cumf = cum_factor(c)
        put(ws, f"D{r}", round(d1c, 3) if d1c else None, fmt=PX, border=BOX)
        put(ws, f"E{r}", d.get("first_day_return_pct"), fmt='+0.0"%";-0.0"%"', border=BOX)
        put(ws, f"F{r}", f'=IFERROR(BDH(TEXT(N($A{r}),"0")&" HK Equity","PX_LAST",'
                         f'$C{r},$C{r}),"run on terminal")', BODY, F_CALC, PX, BOX)
        # Bloomberg's history is back-ADJUSTED for later corporate actions;
        # our value is the raw traded print. Compare on Bloomberg's basis.
        exp_ref = f"$D{r}/{cumf}" if abs(cumf - 1) > 1e-9 else f"$D{r}"
        calc(ws, f"G{r}", f'=IF(NOT(ISNUMBER($F{r})),"run on terminal",'
                          f'IF(ABS($F{r}/({exp_ref})-1)<0.005,"MATCH",'
                          f'"CHECK — BBG "&TEXT($F{r},"0.000")))')
        put(ws, f"H{r}", f"AAStocks listing date (was {est.get(c)}); close from the "
                         f"local kline session list"
                         + (f"; BBG basis = raw ÷ {cumf:g} (later corporate "
                            f"actions back-adjust its history)" if abs(cumf - 1) > 1e-9 else ""),
            NOTE, border=BOX)
        r += 1
    r += 1

    # ---- B. offer prices ruled against the filing parse --------------------
    ruled = []
    for fn, key in (("press_figures.json", "final_price"),
                    ("conflict_rulings.json", "value")):
        p2 = ROOT / "data" / "batches" / fn
        if not p2.exists():
            continue
        blob = _j.loads(p2.read_text())
        rows = blob.get("deals") or blob.get("rulings") or []
        for x in rows:
            if fn.startswith("conflict") and x.get("field") != "final_price":
                continue
            v = x.get(key)
            if v is not None and x.get("code") in by_code:
                ruled.append((x["code"], v, (x.get("src") or x.get("note") or "")[:150]))
    put(ws, f"A{r}", f"B · OFFER PRICE  ({len(ruled)} deals — the filing parse had returned the "
                     f"MAXIMUM offer price, or a figure that was not a price at all)", SECT)
    r += 1
    for j, h in enumerate(["Code", "Name", "Our offer price", "", "",
                           "BBG offer price", "VERDICT", "What we relied on"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    r += 1
    for c, v, src in ruled:
        d = by_code[c]
        put(ws, f"A{r}", c, fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", d.get("final_price"), fmt=PX, border=BOX)
        put(ws, f"F{r}", "no BDP field (EQY_INIT_PO_SH_PRC rejected on the "
                         "desk terminal 2026-08-26) — check the DES/IPO screen",
            NOTE, border=BOX)
        calc(ws, f"G{r}", f'=IF(ISNUMBER($F{r}),IF(ABS($F{r}/$C{r}-1)<0.005,'
                          f'"MATCH","CHECK"),"verified via section A: the day-1 '
                          f'close matched Bloomberg and the day-1 % is '
                          f'consistent only with this price")')
        put(ws, f"H{r}", src, NOTE, border=BOX)
        r += 1
    r += 1

    # ---- C. corporate actions the raw prints carried unadjusted ------------
    acts = dict(acts0)
    for c_ in ent0:                  # entitlement-only deals belong here too
        acts.setdefault(c_, [])
    put(ws, f"A{r}", f"C · CORPORATE ACTIONS  ({len(acts)} deals — Bloomberg's history is "
                     f"ADJUSTED, ours is raw-plus-correction, so our day-1 close divided by the "
                     f"cumulative factor must equal Bloomberg's print for that day)", SECT)
    r += 1
    for j, h in enumerate(["Code", "Name", "Listing date", "Our day-1 close",
                           "÷ factor = expected", "BBG close that day", "VERDICT",
                           "Action applied"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    r += 1
    for c, evs in sorted(acts.items()):
        d = by_code.get(c)
        if not d:
            continue
        cum = cum_factor(c)          # includes entitlement TERP factors
        d1c = (d.get("final_price") or 0) * (1 + (d.get("first_day_return_pct") or 0) / 100)
        put(ws, f"A{r}", c, fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", d.get("ipo_date"), fmt=DATEF, border=BOX)
        put(ws, f"D{r}", round(d1c, 3) if d1c else None, fmt=PX, border=BOX)
        calc(ws, f"E{r}", f"=IFERROR($D{r}/{cum},\"\")", PX)
        put(ws, f"F{r}", f'=IFERROR(BDH(TEXT(N($A{r}),"0")&" HK Equity","PX_LAST",'
                         f'$C{r},$C{r}),"run on terminal")', BODY, F_CALC, PX, BOX)
        calc(ws, f"G{r}", f'=IF(NOT(ISNUMBER($F{r})),"run on terminal",'
                          f'IF(ABS($F{r}/$E{r}-1)<0.02,"MATCH",'
                          f'"CHECK — BBG "&TEXT($F{r},"0.000")))')
        parts = [f"{e['date']} x{e['ratio']:g}" for e in evs]
        parts += [f"{e['date']} {e.get('event','entitlement')} (BBG-only x{e['factor']:g})"
                  for e in ent0.get(c, [])]
        put(ws, f"H{r}", "; ".join(parts) + f"  (BBG basis x{cum:g})",
            NOTE, border=BOX)
        r += 1
    r += 1

    # ---- D. deal sizes rebuilt after an impossible gross -------------------
    sized = [d for d in deals if d.get("size_note")
             and "below the deal" in str(d.get("size_note"))]
    put(ws, f"A{r}", f"D · DEAL SIZE  ({len(sized)} deals — the stated gross was smaller than "
                     f"the deal's own net proceeds, which cannot happen)", SECT)
    r += 1
    for j, h in enumerate(["Code", "Name", "Our deal size (HK$m)", "Basis", "",
                           "BBG offering size", "VERDICT", "What we relied on"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    r += 1
    for d in sized:
        c = d["code"]
        put(ws, f"A{r}", c, fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", d.get("deal_size_hkdm"), fmt=MONEY, border=BOX)
        put(ws, f"D{r}", d.get("size_basis"), NOTE, border=BOX)
        put(ws, f"F{r}", "no BDP field for offering size (EQUITY_OFFERINGS "
                         "is a BDS order-book function) — check CACS / the "
                         "IPO screen",
            NOTE, border=BOX)
        calc(ws, f"G{r}", '"check on CACS — the stated gross was arithmetically '
                          'impossible (below net), so the rebuilt figure is '
                          'shares x price or net proceeds"')
        put(ws, f"H{r}", str(d.get("size_note"))[:150], NOTE, border=BOX)
        r += 1

    ws.freeze_panes = "A5"
    for col, w in (("A", 8), ("B", 24), ("C", 15), ("D", 15), ("E", 17),
                   ("F", 17), ("G", 26), ("H", 78)):
        ws.column_dimensions[col].width = w
    return ws


# ---------------------------------------------------------------- CS League --
def sheet_cs_league(wb, deals):
    """AAStocks-style cornerstone league — same aggregation the HTML tab uses
    (clean_names.cs_league), so the two deliverables cannot disagree."""
    from clean_names import cs_league
    rows = cs_league(deals)
    ws = wb.create_sheet("CS League")
    put(ws, "A1", "CORNERSTONE LEAGUE — how deals anchored by each investor traded", TITLE)
    put(ws, "A2", "One row per investor (grouped on the Screener's normalized key, so long/short "
                  "forms of one house count once). Averages are simple means across their deals: "
                  "day-1 pop = offer→close; ex-pop legs strip the pop (day-1 close→1w/1m/3m). "
                  "Hit = share of their deals closing day-1 above offer. Use the filter arrows "
                  "to cut by deal count.", SUB)
    # band row: WITH POP (vs offer) | EX-POP (from the day-1 close)
    put(ws, "D3", "WITH POP — vs offer", BOLD)
    put(ws, "H3", "EX-POP — from day-1 close", BOLD)
    heads = ["Investor", "Deals", "Day-1 hit",
             "Avg day-1 pop", "Avg 1w", "Avg 1m", "Avg 3m",
             "Avg 1w ex-pop", "Avg 1m ex-pop", "Avg 3m ex-pop",
             "Their deals (code · name)"]
    for j, h in enumerate(heads, 1):
        hdr(ws, f"{get_column_letter(j)}4", h)
    PCT = '+0.0"%";-0.0"%"'
    for i, rrow in enumerate(rows):
        r = 5 + i
        put(ws, f"A{r}", rrow["investor"], border=BOX)
        put(ws, f"B{r}", rrow["n"], fmt="0", border=BOX)
        put(ws, f"C{r}", rrow["hit"], fmt='0"%"', border=BOX)
        for col, key in (("D", "avg_d1"), ("E", "avg_1w_pop"), ("F", "avg_1m_pop"),
                         ("G", "avg_3m_pop"), ("H", "avg_1w"), ("I", "avg_1m"),
                         ("J", "avg_3m")):
            put(ws, f"{col}{r}", rrow[key], fmt=PCT, border=BOX)
        put(ws, f"K{r}", "; ".join(f"{d['code']} {d['name']}" for d in rrow["deals"]),
            NOTE, border=BOX)
    if rows:
        last = 4 + len(rows)
        for col in "DEFGHIJ":
            ws.conditional_formatting.add(
                f"{col}5:{col}{last}",
                ColorScaleRule(start_type="num", start_value=-40, start_color="00F8696B",
                               mid_type="num", mid_value=0, mid_color="00FFFFFF",
                               end_type="num", end_value=60, end_color="0063BE7B"))
        ws.auto_filter.ref = f"A4:K{last}"
    ws.freeze_panes = "D5"
    for c, w in (("A", 44), ("B", 7), ("C", 9), ("D", 13), ("E", 11), ("F", 11),
                 ("G", 11), ("H", 13), ("I", 13), ("J", 13), ("K", 70)):
        ws.column_dimensions[c].width = w
    return ws


# ---------------------------------------------------------------- SM League --
def sheet_stab_league(wb, deals):
    """Stabilising-manager league, built exactly like CS League and off the
    same shared aggregation (clean_names.stab_league) the HTML tab uses."""
    from clean_names import stab_league
    rows = stab_league(deals)
    ws = wb.create_sheet("SM League")
    put(ws, "A1", "STABILISING-MANAGER LEAGUE — how deals each bank defended traded", TITLE)
    put(ws, "A2", "One row per stabilising manager (grouped on the bank family, so "
                  "'Goldman Sachs (Asia) L.L.C.' and 'Goldman Sachs International' count once). "
                  "The stabilising manager holds the greenshoe and the after-market bid, so this "
                  "is a different question from the sponsor league: not who sold the deal, but "
                  "who defended it. Averages are simple means. THE DAY-1 BLOCK splits the session "
                  "the manager actually defended: 'open vs issue' is the pop (where it opened "
                  "against the price the bank sold at), 'close vs issue' is the day-1 return, and "
                  "'open→close' is whether that open was HELD or given back — a negative there "
                  "with a positive close means the bank spent the day supporting a fading stock. "
                  "Ex-pop legs strip the pop (day-1 close→1w/1m/3m). Hit = share closing day-1 above "
                  "offer. Shoe full/lapsed = of the deals whose outcome is known — a shoe "
                  "exercised in full never needed support; a lapsed one was bought back in. "
                  "READ THIS AS DEAL MIX, NOT A SKILL RANKING: the banks at the bottom lead the "
                  "large international deals — bigger books, institutionally priced, less left on "
                  "the table — while the top of the table sits on smaller HK retail-driven "
                  "offerings where pops are structurally larger. A low median means 'defended a "
                  "tightly-priced deal', not 'defended it badly'; compare a bank against deals of "
                  "its own size and regime.", SUB)
    put(ws, "D3", "WITH POP — vs offer", BOLD)
    put(ws, "H3", "EX-POP — from day-1 close", BOLD)
    put(ws, "K3", "SHOE OUTCOME", BOLD)
    put(ws, "D3", "DAY 1 — the session the manager actually defended", BOLD)
    put(ws, "H3", "WITH POP — vs offer", BOLD)
    put(ws, "K3", "EX-POP — from day-1 close", BOLD)
    put(ws, "N3", "SHOE OUTCOME", BOLD)
    heads = ["Stabilising manager", "Deals", "Day-1 hit",
             "Day-1 open vs issue", "Day-1 close vs issue", "Day-1 open→close",
             "Day-1 known",
             "Avg 1w", "Avg 1m", "Avg 3m",
             "Avg 1w ex-pop", "Avg 1m ex-pop", "Avg 3m ex-pop",
             "Shoe full %", "Shoe lapsed %", "Outcome known",
             "Avg deal size (HK$m)", "Their deals (code · name)"]
    for j, h in enumerate(heads, 1):
        hdr(ws, f"{get_column_letter(j)}4", h)
    PCT = '+0.0"%";-0.0"%"'
    for i, rrow in enumerate(rows):
        r = 5 + i
        put(ws, f"A{r}", rrow["manager"], border=BOX)
        put(ws, f"B{r}", rrow["n"], fmt="0", border=BOX)
        put(ws, f"C{r}", rrow["hit"], fmt='0"%"', border=BOX)
        # D/E/F are the day-1 session: where it opened against the price the
        # bank sold at, where it closed, and whether that open was held.
        put(ws, f"D{r}", rrow["avg_d1_open"], fmt=PCT, border=BOX)
        put(ws, f"E{r}", rrow["avg_d1"], fmt=PCT, border=BOX)
        put(ws, f"F{r}", rrow["avg_d1_open_close"], fmt=PCT, border=BOX)
        put(ws, f"G{r}", rrow["d1_open_known"], fmt="0", border=BOX)
        for col, key in (("H", "avg_1w_pop"), ("I", "avg_1m_pop"),
                         ("J", "avg_3m_pop"), ("K", "avg_1w"), ("L", "avg_1m"),
                         ("M", "avg_3m")):
            put(ws, f"{col}{r}", rrow[key], fmt=PCT, border=BOX)
        put(ws, f"N{r}", rrow["shoe_full_pct"], fmt='0"%"', border=BOX)
        put(ws, f"O{r}", rrow["shoe_lapsed_pct"], fmt='0"%"', border=BOX)
        put(ws, f"P{r}", rrow["shoe_known"], fmt="0", border=BOX)
        put(ws, f"Q{r}", rrow["avg_size"], fmt=MONEY, border=BOX)
        put(ws, f"R{r}", "; ".join(f"{d['code']} {d['name']}" for d in rrow["deals"]),
            NOTE, border=BOX)
    if rows:
        last = 4 + len(rows)
        for col in "DEFHIJKLM":
            ws.conditional_formatting.add(
                f"{col}5:{col}{last}",
                ColorScaleRule(start_type="num", start_value=-40, start_color="00F8696B",
                               mid_type="num", mid_value=0, mid_color="00FFFFFF",
                               end_type="num", end_value=60, end_color="0063BE7B"))
        ws.auto_filter.ref = f"A4:R{last}"
    ws.freeze_panes = "D5"
    for c, w in (("A", 40), ("B", 7), ("C", 9), ("D", 19), ("E", 20), ("F", 17),
                 ("G", 12), ("H", 11), ("I", 11), ("J", 11), ("K", 13), ("L", 13),
                 ("M", 13), ("N", 11), ("O", 12), ("P", 13), ("Q", 18), ("R", 70)):
        ws.column_dimensions[c].width = w
    return ws


# ---------------------------------------------------------------- SizeBench --
def sheet_sizebench(wb, deals, cfg, n):
    ws = wb.create_sheet("SizeBench")
    put(ws, "A1", "SIZE BENCHMARKS", TITLE)
    put(ws, "A2", "bucket thresholds are INPUT-BLUE (README weights section explains) — counts recompute", SUB)
    bkts = cfg["size_buckets_hkdm"]
    put(ws, "A4", "Bucket", BOLD); put(ws, "B4", "Min HK$m", BOLD)
    for i, b in enumerate(bkts):
        put(ws, f"A{5 + i}", b["label"], BODY, border=BOX)
        inp(ws, f"B{5 + i}", b["min"], MONEY)
    years = sorted({(d.get("ipo_date") or "")[:4] for d in deals if d.get("ipo_date")})
    hdr(ws, "D4", "Year")
    for j, b in enumerate(bkts):
        hdr(ws, f"{get_column_letter(5 + j)}4", b["label"])
    hdr(ws, f"{get_column_letter(5 + len(bkts))}4", "Sum HK$bn")
    size_rng = db_range("deal_size_hkdm", n)
    date_rng = db_range("ipo_date", n)
    for i, y in enumerate(years):
        r = 5 + i
        put(ws, f"D{r}", y, BODY, border=BOX, align=C_MID)
        for j in range(len(bkts)):
            lo = f"$B${5 + j}"
            hi_clause = "" if j == 0 else f",{size_rng},\"<\"&$B${4 + j}"
            calc(ws, f"{get_column_letter(5 + j)}{r}",
                 f"=COUNTIFS({date_rng},\">=\"&DATE({y},1,1),{date_rng},\"<=\"&DATE({y},12,31),"
                 f"{size_rng},\">=\"&{lo}{hi_clause})", "0")
        calc(ws, f"{get_column_letter(5 + len(bkts))}{r}",
             f"=SUMIFS({size_rng},{date_rng},\">=\"&DATE({y},1,1),{date_rng},\"<=\"&DATE({y},12,31))/1000",
             "#,##0.0")
    put(ws, "A12", "MEGA-DEAL PRECEDENTS (top 15 by proceeds)", SECT)
    top = sorted([d for d in deals if d.get("deal_size_hkdm")],
                 key=lambda d: -d["deal_size_hkdm"])[:15]
    for j, h in enumerate(["Code", "Name", "IPO date", "HK$m", "Subsector"], 1):
        hdr(ws, f"{get_column_letter(j)}13", h)
    for i, d in enumerate(top):
        r = 14 + i
        put(ws, f"A{r}", d["code"], fmt="@", border=BOX)
        put(ws, f"B{r}", d.get("name"), border=BOX)
        put(ws, f"C{r}", d.get("ipo_date"), fmt=DATEF, border=BOX)
        put(ws, f"D{r}", d.get("deal_size_hkdm"), fmt=MONEY, border=BOX)
        put(ws, f"E{r}", d.get("subsector"), border=BOX)
    for c, w in (("A", 8), ("B", 26), ("C", 11), ("D", 12), ("E", 22), ("D", 12)):
        ws.column_dimensions[c].width = w
    return ws


# ---------------------------------------------------------------- others -----
def sheet_taxonomy(wb, tax):
    ws = wb.create_sheet("Taxonomy")
    put(ws, "A1", "SECTOR / SUBSECTOR TAXONOMY", TITLE)
    put(ws, "A2", "single source: data/taxonomy.json — edit there and rebuild; screener gates on subsector first", SUB)
    # column E: flat label list for dropdowns. An INLINE DataValidation list is
    # silently truncated at 255 characters by Excel, so dropdowns MUST point at
    # a range, never a joined string.
    flat = [s["label"] for subs in tax["sectors"].values() for s in subs]
    for i, lbl in enumerate(flat):
        put(ws, f"E{4 + i}", lbl, NOTE)
    ws.column_dimensions["E"].hidden = True
    define(wb, "SubsectorList", f"Taxonomy!$E$4:$E${3 + len(flat)}")
    # every sector says which index its alpha is measured against, so an alpha
    # number is never against an unnamed benchmark
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent))
    try:
        from fetch_prices import BENCH, DEFAULT_BENCH
    except Exception:
        BENCH, DEFAULT_BENCH = {}, ("^HSI", "Hang Seng Index")
    hdr(ws, "A3", "Sector / subsector")
    hdr(ws, "B3", "")
    hdr(ws, "C3", "Comp note")
    hdr(ws, "D3", "Alpha benchmark (ticker — name)")
    r = 4
    for sec, subs in tax["sectors"].items():
        put(ws, f"A{r}", sec, SECT, F_CALC)
        bt, blabel = BENCH.get(sec, DEFAULT_BENCH)
        put(ws, f"D{r}", f"{bt} — {blabel}", BODY, F_CALC)
        r += 1
        for s in subs:
            put(ws, f"B{r}", s["label"], BODY)
            put(ws, f"C{r}", s.get("comp_note", ""), NOTE)
            r += 1
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70
    return ws


def sheet_verification(wb, deals, counts, n):
    ws = wb.create_sheet("Verification")
    put(ws, "A1", "VERIFICATION — book vs official HKEX statistics", TITLE)
    put(ws, "A2", "book counts are IPOs with a public offer (allotment results filed); official totals include "
                  "listings by introduction / transfers — the Delta column is expected to be small and positive", SUB)
    for j, h in enumerate(["Year", "Deals in book", "Official new listings", "Delta",
                           "Book proceeds HK$bn", "Official IPO funds HK$bn", "Diff %"], 1):
        hdr(ws, f"{get_column_letter(j)}4", h)
    years = sorted({(d.get("ipo_date") or "")[:4] for d in deals if d.get("ipo_date")})
    date_rng = db_range("ipo_date", n)
    size_rng = db_range("deal_size_hkdm", n)
    ydata = counts.get("years", {}) if isinstance(counts, dict) else {}
    for i, y in enumerate(years):
        r = 5 + i
        put(ws, f"A{r}", y, BODY, border=BOX, align=C_MID)
        calc(ws, f"B{r}", f"=COUNTIFS({date_rng},\">=\"&DATE({y},1,1),{date_rng},\"<=\"&DATE({y},12,31))", "0")
        off = ydata.get(y, {})
        put(ws, f"C{r}", off.get("new_listings_total"), fmt="0", border=BOX)
        calc(ws, f"D{r}", f"=IF(C{r}=\"\",\"n/a\",C{r}-B{r})", "0")
        calc(ws, f"E{r}", f"=SUMIFS({size_rng},{date_rng},\">=\"&DATE({y},1,1),"
                          f"{date_rng},\"<=\"&DATE({y},12,31))/1000", "#,##0.0")
        offm = off.get("equity_funds_raised_ipo_hkdm")
        put(ws, f"F{r}", round(offm / 1000, 1) if offm else None, fmt="#,##0.0", border=BOX)
        calc(ws, f"G{r}", f"=IF(OR(F{r}=\"\",E{r}=0),\"n/a\",F{r}/E{r}-1)", "+0.0%;-0.0%")
    r0 = 6 + len(years)
    put(ws, f"A{r0}", "LANDMARK SPOT-CHECKS", SECT)
    for j, h in enumerate(["Code", "Name", "Field", "Official", "In book", "Check"], 1):
        hdr(ws, f"{get_column_letter(j)}{r0 + 1}", h)
    refs = counts.get("reference_deals", []) if isinstance(counts, dict) else []
    r = r0 + 2
    code_rng = db_range("code", n)
    for ref in refs:
        for field, official in (("final_price", ref.get("final_price")),
                                ("deal_size_hkdm", ref.get("deal_size_hkdm"))):
            if official is None:
                continue
            put(ws, f"A{r}", ref.get("code"), fmt="@", border=BOX)
            put(ws, f"B{r}", ref.get("name"), border=BOX)
            put(ws, f"C{r}", field, border=BOX)
            put(ws, f"D{r}", official, fmt=PX if "price" in field else MONEY, border=BOX)
            calc(ws, f"E{r}", f"=INDEX({db_range(field, n)},MATCH(\"{ref.get('code')}\",{code_rng},0))",
                 PX if "price" in field else MONEY)
            calc(ws, f"F{r}", f"=IF(E{r}=\"\",\"MISSING\",IF(ABS(E{r}-D{r})/D{r}<0.02,\"OK\",\"CHECK\"))")
            ws.conditional_formatting.add(f"F{r}", FormulaRule(formula=[f'$F{r}="CHECK"'], fill=F_OVR))
            r += 1
    for c, w in (("A", 8), ("B", 24), ("C", 20), ("D", 12), ("E", 12), ("F", 10), ("G", 9)):
        ws.column_dimensions[c].width = w
    return ws


def sheet_verification_excl(ws, deals, r0):
    """Answer 'is anyone missing?' by SHOWING what was excluded and why."""
    import json as _json
    put(ws, f"A{r0}", "WHAT IS EXCLUDED FROM THE BOOK (and why)", SECT)
    put(ws, f"A{r0+1}", "Inclusion rule: a Main Board listing that filed an Allotment Results "
                        "announcement, i.e. an IPO with a public offering. Everything below "
                        "appeared in the aggregator's listing table but is NOT an IPO by that rule.", NOTE)
    have = {d["code"] for d in deals}
    try:
        roster = _json.loads((ROOT / "data" / "batches" / "bulk_roster.json").read_text())["deals"]
    except Exception:
        roster = []
    try:
        allot = _json.loads((ROOT / "data" / "batches" / "hkex_allotments.json").read_text())["deals"]
    except Exception:
        allot = []
    gem = [a for a in allot if a["board"] == "GEM"]
    extra = [r for r in roster if r["code"] not in have
             and not any(a["code"] == r["code"] for a in allot)]
    r = r0 + 2
    for j, h in enumerate(["Code", "Name", "Date", "Why excluded"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    for a in gem:
        r += 1
        for j, v in enumerate([a["code"], a["stock_name_short"], a["ipo_date_est"],
                               "GEM board — Main Board only by design"], 1):
            put(ws, f"{get_column_letter(j)}{r}", v, border=BOX)
    for x in extra:
        r += 1
        for j, v in enumerate([x["code"], x.get("name"), x.get("ipo_date"),
                               "no Allotment Results filed — listing by introduction / "
                               "transfer / not an IPO"], 1):
            put(ws, f"{get_column_letter(j)}{r}", v, border=BOX)
    put(ws, f"A{r+2}", f"{len(gem)} GEM listings + {len(extra)} non-IPO listings excluded. "
                       f"Anything not on this list and not in the Database is a genuine gap — "
                       f"tell the builder.", NOTE)
    return r + 3


def sheet_bbg(wb, deals, n):
    """Bloomberg cross-check sheet — the user's own formula pattern, verbatim:
    column B resolves the IPO order ID once per row, and the CP036/CP037/shoe
    columns reference THAT id (B{r}), exactly like =BDP(A1&" Action","CP036").
    Only computes on a terminal; inert and clearly labelled elsewhere."""
    ws = wb.create_sheet("BBG Verify")
    put(ws, "A1", "BLOOMBERG CROSS-CHECK — terminal only", TITLE)
    put(ws, "A2", "Column B pulls each deal's Bloomberg IPO order ID (EQUITY_OFFERINGS); "
                  "the columns after it feed off that ID. Compare against the scraped "
                  "columns pulled in from the Database. 511 BDS calls are heavy — let it "
                  "run once, then Paste-Special values if you want it static.", SUB)
    heads = ["Code / Name", "BBG IPO order ID", "Retail o/sub CP036", "Scraped public sub",
             "Instn o/sub CP037", "Scraped intl sub", "Greenshoe facility",
             "Shoe exercised (BBG)", "Shoe outcome (scraped)", "Current P/E (BBG)",
             "A/H conversion (BBG)", "A-share (scraped)", "Mkt cap at listing (BBG)",
             "Mkt cap (scraped)", "P/E at listing (BBG)", "P/E at IPO (scraped)",
             "A-share P/E at H-IPO (BBG)", "HSAHP Index at IPO (BBG)",
             "P/S today (BBG)", "A-share mkt cap now (BBG)"]
    for j, h in enumerate(heads, 1):
        hdr(ws, f"{get_column_letter(j)}4", h)
        ws.column_dimensions[get_column_letter(j)].width = 15 if j > 1 else 26
    ccode = db_col("code")
    for i in range(n):
        r = DB_R0 + i
        tk = f'Database!${ccode}{r}&" HK Equity"'
        put(ws, f"A{r}", f"=Database!{ccode}{r}&\"  \"&Database!{db_col('name')}{r}",
            BODY, border=BOX)
        put(ws, f"B{r}",
            f'=IFERROR(INDEX(BDS({tk},"EQUITY_OFFERINGS","array=t","endcol=1"),'
            f'MATCH("IPO",BDS({tk},"EQUITY_OFFERINGS","array=t","startcol=2","endcol=2"),0)),"")',
            BODY, F_CALC, border=BOX)
        put(ws, f"C{r}", f'=IF($B{r}="","",IFERROR(BDP($B{r}&" Action","CP036"),""))',
            BODY, F_CALC, fmt="#,##0.00", border=BOX)
        put(ws, f"D{r}", f"=Database!{db_col('oversub_public_mult')}{r}", BODY, fmt="#,##0.00", border=BOX)
        put(ws, f"E{r}", f'=IF($B{r}="","",IFERROR(BDP($B{r}&" Action","CP037"),""))',
            BODY, F_CALC, fmt="#,##0.00", border=BOX)
        put(ws, f"F{r}", f"=Database!{db_col('oversub_intl_mult')}{r}", BODY, fmt="#,##0.00", border=BOX)
        put(ws, f"G{r}", f'=IF($B{r}="","",IFERROR(BDP($B{r}&" Action","GREENSHOE_FACILITY"),""))',
            BODY, F_CALC, fmt="#,##0.00", border=BOX)
        put(ws, f"H{r}", f'=IF($B{r}="","",IFERROR(BDP($B{r}&" Action","OFFERING_GREENSHOE_SHARES_EX"),""))',
            BODY, F_CALC, fmt="#,##0.00", border=BOX)
        put(ws, f"I{r}", f"=Database!{db_col('greenshoe_exercised_final')}{r}", BODY, border=BOX)
        put(ws, f"J{r}", f'=IFERROR(BDP({tk},"PE_RATIO"),"")', BODY, F_CALC, fmt="0.0", border=BOX)
        # A/H_SHARE_CONVERSION returns the OTHER line's ticker, so a value here
        # where the scraped column says nothing means the issuer does have an A
        # line the filings never named — Bloomberg wins that disagreement.
        put(ws, f"K{r}", f'=IFERROR(BDP({tk},"A/H_SHARE_CONVERSION"),"")',
            BODY, F_CALC, border=BOX)
        put(ws, f"L{r}", f'=IF(Database!{db_col("a_share_code")}{r}="","(none found)",'
                         f'Database!{db_col("a_share_code")}{r})', BODY, border=BOX)
        ipo_c = db_col("ipo_date")
        # desk-verified: CUR_MKT_CAP already arrives in millions — the old
        # /1000000 made every value read as ~0.00
        put(ws, f"M{r}", f'=IFERROR(BDH({tk},"CUR_MKT_CAP",Database!${ipo_c}{r},'
                         f'Database!${ipo_c}{r}),"")',
            BODY, F_CALC, fmt=MONEY, border=BOX)
        put(ws, f"N{r}", f"=Database!{db_col('mktcap_ipo_hkdm')}{r}", BODY, fmt=MONEY, border=BOX)
        put(ws, f"O{r}", f'=IFERROR(BDH({tk},"PE_RATIO",Database!${ipo_c}{r},'
                         f'Database!${ipo_c}{r}),"")', BODY, F_CALC, fmt="0.0", border=BOX)
        put(ws, f"P{r}", f'=IF(Database!{db_col("pe_ipo")}{r}="","(n/m or not derivable)",'
                         f'Database!{db_col("pe_ipo")}{r})', BODY, fmt="0.0", border=BOX)
        # the A line's OWN multiple on the day the H priced — A/H deals only.
        # 300223.SZ -> "300223 CH Equity"; blank for issuers with no A line.
        ac = db_col("a_share_code")
        put(ws, f"Q{r}",
            f'=IF(Database!{ac}{r}="","",'
            f'IFERROR(BDH(LEFT(Database!{ac}{r},FIND(".",Database!{ac}{r})-1)'
            f'&" CH Equity","PE_RATIO",Database!${ipo_c}{r},Database!${ipo_c}{r}),""))',
            BODY, F_CALC, fmt="0.0", border=BOX)
        # the Hang Seng AH Premium index ON the listing day — where the whole
        # A/H complex stood when this deal priced. A/H rows only.
        put(ws, f"R{r}",
            f'=IF(Database!{ac}{r}="","",'
            f'IFERROR(BDH("HSAHP Index","PX_LAST",Database!${ipo_c}{r},'
            f'Database!${ipo_c}{r}),""))',
            BODY, F_CALC, fmt="0.0", border=BOX)
        # TODAY's P/S on the H line. Nothing public fills this — the filings
        # state revenue and a cap at the OFFER, never a live multiple — so the
        # Database's "P/S today" column reads straight off this cell.
        put(ws, f"S{r}", f'=IFERROR(BDP({tk},"PX_TO_SALES_RATIO"),"")',
            BODY, F_CALC, fmt="0.0", border=BOX)
        # the A LINE's market cap today, from Bloomberg rather than the Tencent
        # quote the snapshot uses. A/H rows only; the scraped a_mktcap_now
        # column falls back to this cell wherever the snapshot had no figure.
        put(ws, f"T{r}",
            f'=IF(Database!{ac}{r}="","",'
            f'IFERROR(BDP(LEFT(Database!{ac}{r},FIND(".",Database!{ac}{r})-1)'
            f'&" CH Equity","CUR_MKT_CAP"),""))',
            BODY, F_CALC, fmt=MONEY, border=BOX)
    # a disagreement should be visible without reading two columns
    for a, b in (("C", "D"), ("E", "F"), ("M", "N"), ("O", "P")):
        ws.conditional_formatting.add(
            f"{a}{DB_R0}:{a}{DB_R0 + n - 1}",
            FormulaRule(formula=[f'AND(ISNUMBER({a}{DB_R0}),ISNUMBER({b}{DB_R0}),'
                                 f'ABS({a}{DB_R0}-{b}{DB_R0})>0.25*ABS({b}{DB_R0}))'],
                        fill=F_OVR))
    put(ws, f"A{DB_R0 + n + 1}",
        "Orange = Bloomberg and the scraped value differ by more than 25%. "
        "A blank scraped cell beside a Bloomberg value is a gap this file could not "
        "fill from public filings — take Bloomberg's.", NOTE)
    ws.freeze_panes = "B5"
    return ws


def sheet_readme(wb, cfg, as_of, deals):
    """Deliberately short. The working guide is delivered separately, not buried
    in the file — the desk asked for fewer footnotes, not more."""
    ws = wb.create_sheet("Notes")
    w = cfg["weights"]
    put(ws, "A1", "HK IPO DATABASE", TITLE)
    put(ws, "A2", f"{len(deals)} Main Board IPOs 2021-2026 + pipeline · data as of {as_of}", SUB)
    lines = [
        ("", ""),
        ("Blue cells", "you type these. Everything else calculates."),
        ("Start here", "Screener tab — pick a deal OR type your own terms; comps rank instantly."),
        ("Screener", "pick any past or pipeline deal, get its closest comps."),
        ("CS League", "one row per cornerstone investor: every deal they anchored, average "
                      "day-1 pop and 1w/1m/3m ex-pop. Same grouping key as the Screener."),
        ("SM League", "the same table for STABILISING MANAGERS — the bank holding the "
                      "greenshoe and the after-market bid. Read it against CS League: one "
                      "says who anchored the deal, the other who defended it. The shoe "
                      "columns are the tell — exercised in full means the price never "
                      "needed support; lapsed means stock was bought back in."),
        ("Eff. free float", "deal size x (1 - cornerstone %) / market cap - the slice of the "
                            "company that can actually trade on day 1 (cornerstones are locked "
                            "6 months)."),
        ("Analogs", "the raw subscription-bucket history behind the Screener read-out."),
        ("Green / amber", "cross-checked / single-source or judgment."),
        ("Money", "HK$ millions. Day-1 + = closed above offer. A premium + = A trades ABOVE H."),
        ("Subscription", "filing basis: 10x = ten times the shares on offer. Bloomberg CP036/CP037 "
                         "override the scrape wherever the desk paste carries them."),
        ("Ex-pop columns", "TEAL headers. 1w/1m/3m ex-pop start at the day-1 CLOSE (standard "
                           "aftermarket basis); Alpha 1m ex-pop nets the index over the identical "
                           "window. The TRADEABLE-entry view — buy at the day-1 OPEN — is the "
                           "Day-1 open→close column here plus the open-rebased charts in the "
                           "dashboard."),
        ("P/E & P/S basis", "ALL at-IPO multiples are TRAILING, never forward: market cap at the "
                            "final offer price ÷ the last FULL pre-IPO fiscal year's net income "
                            "(P/E) or revenue (P/S), as filed in the prospectus. One basis for "
                            "every deal AND the pipeline's expected multiples, so they compare. "
                            "Loss-makers show n/m P/E — use P/S; an 18A pre-revenue biotech's "
                            "huge P/S is shown with its scale explained rather than left blank."),
        ("P/E, two readings — WHICH TO USE",
         "There are two AT-IPO P/E columns and they are both right, on different "
         "bases. 'P/E at IPO' = final mktcap ÷ last PRE-IPO fiscal-year net income, "
         "the prospectus basis — the SAME basis for all 514 rows, with both inputs "
         "visible as columns beside it. 'P/E at IPO (BBG)' = price ÷ trailing-12m "
         "EPS at listing, Bloomberg's own basis. They diverge most for 2021-23 "
         "vintages (median BBG/ours ~0.5) because earnings grew between the covered "
         "FY and listing and BBG uses weighted pre-deal shares. USE OURS to compare "
         "deals against each other — it is the only one computed identically for "
         "every row, and you can audit it. USE BLOOMBERG'S when quoting a number "
         "someone will check on a terminal, because that is what they will see. "
         "The book already treats BBG as the referee: where our derived multiple "
         "breaks the plausibility cap, a BBG print within 25% restores it as real "
         "(CALB at ~549x), and agreement at an absurd level is read as a shared "
         "data artifact rather than a confirmation."),
        ("P/E today / P/S today", "'P/E today (BBG)' is the desk paste; 'P/S today "
                                  "(BBG)' has no public source at all, so every row "
                                  "resolves off BBG Verify col S on the terminal and "
                                  "says 'run on terminal' off it. Both sit in "
                                  "FUNDAMENTALS beside the at-IPO multiples so the "
                                  "then-vs-now comparison is one glance."),
        ("P/S now (Screener)", "live BDP PX_TO_SALES_RATIO beside P/E now — resolves on the "
                               "terminal; off-terminal it falls back to the at-IPO P/S."),
        ("A-share mkt cap now", "the Tencent snapshot fills it for A/H pairs; where it "
                                "cannot, the cell resolves off BBG Verify col T "
                                "(CUR_MKT_CAP on the A ticker). Non-A/H rows read N/A."),
        ("Force-include (D17)", "type codes comma-separated (e.g. 9888, 2015) and those deals pin "
                                "to the top of the comps, past every filter. Same control exists "
                                "in the dashboard's Screener."),
        ("", ""),
        ("Not captured", "underwriting fee splits (never public per deal)."),
        ("Judgment call", "sector and subsector are analyst-assigned, shown amber."),
    ]
    r = 3
    for a, b in lines:
        r += 1
        put(ws, f"A{r}", a, BOLD if a else BODY)
        put(ws, f"B{r}", b, BODY)
    r += 2
    put(ws, f"A{r}", "SCREENER WEIGHTS (edit to re-tune)", SECT)
    for name, val, desc in [("W_SUB", w["subsector_match"], "same subsector — dominant by design"),
                            ("W_SEC", w["sector_match_fallback"], "same sector fallback"),
                            ("W_SIZE", w["size_proximity"], "size proximity"),
                            ("W_PROF", w["profitability_match"], "profitability match"),
                            ("W_AH", w["h_share_match"], "H-share match"),
                            ("W_REC", w["recency"], "recency"),
                            ("W_CS", w.get("shared_cornerstone", 12),
                             "shares a cornerstone investor with the target"),
                            ("W_PE", w.get("pe_proximity", 150),
                             "P/E proximity — closer multiples rank higher (beats profitability)"),
                            ("W_PEHW", cfg.get("pe_proximity_log10_halfwidth", 0.6),
                             "P/E log half-width"),
                            ("W_SIZEHW", cfg["size_proximity_log10_halfwidth"], "size log half-width"),
                            ("W_RECD", cfg["recency_horizon_days"], "recency horizon (days)")]:
        r += 1
        put(ws, f"A{r}", name, BOLD)
        inp(ws, f"B{r}", val)
        put(ws, f"C{r}", desc, NOTE)
        define(wb, name, f"Notes!$B${r}")

    r += 2
    put(ws, f"A{r}", "WHY THESE WEIGHTS (measured, not assumed)", SECT)
    for line in (
        "Tested on 24,090 deal pairs: how close were two deals' day-1 returns when they",
        "matched on each factor, versus two random deals (median gap 27.5pp)?",
        "  both A+H .................. 14.2pp  (+49% better than random)",
        "  subscription within 2x .... 18.6pp  (+32% better)",
        "  same sector ............... 26.3pp  (+4% better)",
        "  same subsector ............ 29.2pp  (no better than random)",
        "  size / P/E / cornerstone .. 29-33pp (no better than random)",
        "So subsector-first still decides WHICH companies are comparable (a chip maker is",
        "valued against chip makers), but it does NOT predict the debut. When the question",
        "is 'how will it trade', switch Rank by -> 'demand-similar first', and read the",
        "Analogs tab, where subscription buckets carry the real signal.",
    ):
        r += 1
        put(ws, f"A{r}", line, NOTE)

    # --- where every Database column comes from -----------------------------
    r += 2
    put(ws, f"A{r}", "WHERE EACH DATABASE COLUMN COMES FROM", SECT)
    r += 1
    for h, txt in (("Column", "Source and method"),):
        hdr(ws, f"A{r}", h)
        hdr(ws, f"B{r}", txt)
    for colname, src in COLUMN_SOURCES:
        r += 1
        put(ws, f"A{r}", colname, BOLD, border=BOX)
        put(ws, f"B{r}", src, BODY, border=BOX, align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 26
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 96
    ws.column_dimensions["C"].width = 40
    return ws


# Every column, and exactly what produced it. Written out because "where did
# this number come from" is the first question asked of any figure that goes
# into a trade, and the answer must not live only in the builder's head.
COLUMN_SOURCES = [
    ("Code / Name / date", "HKEX Allotment Results announcement — every IPO must file one, so the "
                           "roster is enumerated from the filings themselves, not from a list."),
    ("Name (CN)", "HKEX bilingual securities feed (activestock/inactivestock _c.json)."),
    ("Sector / Subsector", "Analyst-assigned against data/taxonomy.json. New deals get a "
                           "provisional keyword label (~49% exact) shown amber until relabelled."),
    ("Deal size", "Ladder, and the basis is printed beside it: stated gross proceeds > "
                  "final price x offer shares > net proceeds. From the final-price announcement "
                  "and the prospectus."),
    ("Range low / Max-cap", "Prospectus indicative range; the cap is the 'Maximum Offer Price'. "
                            "A fixed-price offer has no range and is labelled as such."),
    ("Final px", "Allotment Results announcement (the struck price)."),
    ("% of cap", "final price / cap. Fixed-price offers are 100% by definition."),
    ("Mkt cap at IPO", "Ladder, printed in 'Mkt cap basis': final price x shares on listing > "
                       "deal size / offer % of enlarged capital > issuer-stated expected market "
                       "cap > AAStocks 上市市值 scaled to the struck price. Cross-checked against "
                       "AAStocks; >25% disagreement is flagged orange."),
    ("Public / Intl sub", "Allotment Results — the subscription section, read per tranche. "
                          "Under-subscription (<1x) is captured as data, not dropped."),
    ("Cornerstone % / investors", "Prospectus cornerstone section (names cleaned of PDF table "
                                  "damage), filled from AAStocks 機構性投資者 where the filing "
                                  "text did not survive extraction."),
    ("Greenshoe size", "over-allocated shares / offer shares, from the prospectus."),
    ("Shoe outcome", "The END OF STABILISATION notice filed ~30 days after listing — the "
                     "allotment-day wording can only ever say 'not yet'."),
    ("Day-1", "Listing-day close vs the struck offer price (Yahoo). Where Yahoo has no session "
              "at the listing date the value comes from AAStocks instead — it is never measured "
              "off a later session."),
    ("1-week / 1-month / 3-month", "Close 5 / 21 / 63 trading bars after the debut, all vs the "
                                   "OFFER price."),
    ("1m ex-pop", "Same horizon measured from the day-1 CLOSE instead — strips the debut pop and "
                  "answers whether it held."),
    ("Alpha / Index", "The deal's own sector index over the identical window, anchored at the "
                      "index close BEFORE the listing (the moment the subscription money was "
                      "committed). Index named per row; map on the Taxonomy tab."),
    ("Since IPO", "Latest close vs the offer price."),
    ("Revenue / NI", "Prospectus Financial Information section, latest full year, converted to "
                     "HK$ at a fixed rate stated in the file."),
    ("P/E at IPO", "market cap / latest FY net income. Loss-makers read n/m — a negative P/E is "
                   "meaningless — and carry P/S instead."),
    ("H disc vs A at IPO", "H offer price vs the A-share close on the last session BEFORE the H "
                           "listing, converted at that day's CNYHKD. Negative = H struck below "
                           "the A line. Verified by hand on CATL (-6.7%)."),
    ("A/H premium (today)", "Latest A and H closes, same conversion."),
    ("Sponsors / Bookrunners", "Prospectus cover page. The AAStocks columns beside them are 保薦人 "
                               "and 包銷商 from the deal's AAStocks page — a second, independent "
                               "print of the same fact."),
    ("Prospectus / Allotment / Stabilisation", "Direct HKEX filing URLs, resolved per stock code."),
]


def main(limit=None, out=None):
    deals, tax, cfg, pipe, counts = load()
    if limit:                      # miniature build, used by the formula test
        keep = {d["code"] for d in deals[:limit]}
        deals = [d for d in deals if d["code"] in keep]
    n = len(deals)
    wb = Workbook()
    wb.remove(wb.active)
    as_of = date.today().isoformat()
    an = build_analogs(deals)
    ws_db = sheet_database(wb, deals, n)
    sheet_pipeline(wb, pipe)
    # Calc first: the Screener's pick-list and target-split cells live there
    sheet_scores(wb, n, len(pipe))
    sheet_screener(wb, deals, pipe, cfg, n)
    sheet_analogs(wb, deals, an)
    sheet_readme(wb, cfg, as_of, deals)
    sheet_bbg(wb, deals, n)
    sheet_ah(wb, deals, as_of)
    sheet_verify_changes(wb, deals, n)
    sheet_cs_league(wb, deals)
    sheet_stab_league(wb, deals)
    sheet_sizebench(wb, deals, cfg, n)
    sheet_taxonomy(wb, tax)
    sheet_verification(wb, deals, counts, n)
    # explicit tab order: the daily tool first, reference material last
    order = ["Screener", "Database", "Pipeline", "AH", "CS League", "SM League",
             "Verify (BBG)", "Analogs",
             "SizeBench", "BBG Verify", "Taxonomy", "Verification",
             "Calc (scoring engine)", "Notes"]
    wb._sheets = ([wb[t] for t in order if t in wb.sheetnames]
                  + [ws2 for ws2 in wb._sheets if ws2.title not in order])
    # tab colours: the three tabs the desk lives in stand out
    TABCOLOR = {"Screener": "2A78D6", "Database": "1F3864", "Pipeline": "1BAF7A",
                "BBG Verify": "ED7D31"}
    for ws2 in wb.worksheets:
        ws2.sheet_properties.tabColor = TABCOLOR.get(ws2.title, "BFBFBF")
    _assert_formulas_balanced(wb)
    dest = Path(out) if out else OUT
    dest.parent.mkdir(exist_ok=True, parents=True)
    wb.save(dest)
    # HK stock codes are TEXT ("0300" must keep its leading zero), which makes
    # Excel paint a green "number stored as text" triangle on every row. The
    # cure is the ignoredErrors element — openpyxl accepts the attribute but
    # never serialises it, so it is injected into the saved sheet XML.
    _suppress_text_number_warning(dest, {"Database": "A5:A600", "Pipeline": "A6:A40"})
    print(f"wrote {dest} ({n} deals, {len(pipe)} pipeline)")
    return dest


def _suppress_text_number_warning(path, sheet_ranges):
    """Add <ignoredErrors numberStoredAsText> to the given sheets' XML.

    Rewrites the .xlsx zip in place. The element must sit late in the sheet,
    after pageSetup and before any drawing, so it is inserted ahead of the
    first of those tags rather than blindly appended.
    """
    import re as _re
    import shutil
    import zipfile
    from openpyxl import load_workbook as _lw

    wb2 = _lw(path)
    order = wb2.sheetnames
    wb2.close()
    targets = {}
    for name, sqref in sheet_ranges.items():
        if name in order:
            targets[f"xl/worksheets/sheet{order.index(name) + 1}.xml"] = sqref
    if not targets:
        return
    tmp = Path(str(path) + ".tmp")
    with zipfile.ZipFile(path) as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            sq = targets.get(item.filename)
            if sq:
                xml = data.decode("utf-8")
                if "<ignoredErrors" not in xml:
                    tag = (f'<ignoredErrors><ignoredError sqref="{sq}" '
                           f'numberStoredAsText="1"/></ignoredErrors>')
                    m = _re.search(r"<(drawing|legacyDrawing|tableParts|extLst)\b", xml)
                    xml = (xml[:m.start()] + tag + xml[m.start():]) if m else \
                        xml.replace("</worksheet>", tag + "</worksheet>")
                    data = xml.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(str(tmp), str(path))


def _assert_formulas_balanced(wb):
    """Excel shows an unbalanced formula as a repair prompt on OPEN, which is the
    worst possible place to find out. One stray ')' in a guidance string shipped
    a workbook that would not load, so the build refuses to save one."""
    import re as _re
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                depth = 0
                for ch in _re.sub(r'"[^"]*"', "", v):     # ignore quoted parens
                    depth += (ch == "(") - (ch == ")")
                    if depth < 0:
                        break
                if depth != 0:
                    bad.append(f"{ws.title}!{cell.coordinate}: {v[:90]}")
    if bad:
        raise SystemExit("unbalanced formulas — refusing to write:\n  "
                         + "\n  ".join(bad[:10]))




# ------------------------------------------------------- analogs + new deal --
SUB_BUCKETS = [("<10x", 0, 10), ("10-100x", 10, 100), ("100-1000x", 100, 1000),
               (">=1000x", 1000, 1e12)]
SIZE_BUCKETS = [("<HK$500m", 0, 500), ("HK$0.5-2bn", 500, 2000),
                ("HK$2-10bn", 2000, 10000), (">=HK$10bn", 10000, 1e12)]


def _stats(vals):
    """n, median, hit-rate, p25, p75 for a set of day-1 returns."""
    if not vals:
        return None
    s = sorted(vals)
    n = len(s)

    def q(p):
        i = max(0, min(n - 1, int(round(p * (n - 1)))))
        return s[i]
    med = s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2
    return {"n": n, "median": round(med, 1), "hit": round(100 * sum(1 for v in s if v > 0) / n),
            "p25": round(q(0.25), 1), "p75": round(q(0.75), 1)}


def build_analogs(deals):
    """Historical day-1 outcomes, bucketed. Computed here so Excel needs only
    INDEX/MATCH — MEDIAN(IF()) would be an array formula and Excel-2016 hostile."""
    pool = [d for d in deals if d.get("first_day_return_pct") is not None]
    by_sub_bucket, by_bucket, by_sub, by_year_bucket = {}, {}, {}, {}
    for label, lo, hi in SUB_BUCKETS:
        grp = [d for d in pool if d.get("oversub_public_mult") is not None
               and lo <= d["oversub_public_mult"] < hi]
        st = _stats([d["first_day_return_pct"] for d in grp])
        if st:
            by_bucket[label] = st
        for d in grp:
            key = (d.get("subsector") or "?", label)
            by_sub_bucket.setdefault(key, []).append(d["first_day_return_pct"])
        for d in grp:
            y = d["ipo_date"][:4]
            by_year_bucket.setdefault((y, label), []).append(d["first_day_return_pct"])
    for d in pool:
        by_sub.setdefault(d.get("subsector") or "?", []).append(d["first_day_return_pct"])

    def med(vals):
        s2 = sorted(vals)
        return round(s2[len(s2) // 2] if len(s2) % 2
                     else (s2[len(s2) // 2 - 1] + s2[len(s2) // 2]) / 2, 1)

    val = {}
    for d in deals:
        sub = d.get("subsector")
        if not sub:
            continue
        v = val.setdefault(sub, {"pes": [], "pss": []})
        if d.get("pe_ipo") and 0 < d["pe_ipo"] <= 200:
            v["pes"].append(d["pe_ipo"])
        if d.get("ps_ipo") and 0 < d["ps_ipo"] <= 60:
            v["pss"].append(d["ps_ipo"])
    val = {k: {"pe": med(v["pes"]) if v["pes"] else None,
               "ps": med(v["pss"]) if v["pss"] else None,
               "n_pe": len(v["pes"]), "n_ps": len(v["pss"])} for k, v in val.items()}
    return {
        "val": val,
        "sub_bucket": {k: _stats(v) for k, v in by_sub_bucket.items()},
        "bucket": by_bucket,
        "subsector": {k: _stats(v) for k, v in by_sub.items()},
        "year_bucket": {k: _stats(v) for k, v in by_year_bucket.items()},
        "pool_n": len(pool),
        "years": sorted({d["ipo_date"][:4] for d in pool}),
    }


def sheet_analogs(wb, deals, an):
    ws = wb.create_sheet("Analogs")
    put(ws, "A1", "DAY-1 ANALOGS — what deals like this actually did", TITLE)
    put(ws, "A2", f"built from {an['pool_n']} deals with a listing-day price, {', '.join(an['years'])}. "
                  "'Hit-rate' = % that closed above the offer price. Read n before trusting a cell.", SUB)
    r = 4
    put(ws, f"A{r}", "BY SUBSCRIPTION LEVEL (all deals)", SECT); r += 1
    for j, h in enumerate(["Subscription", "n", "Median day-1", "Hit-rate", "P25", "P75"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    for label, _lo, _hi in SUB_BUCKETS:
        r += 1
        st = an["bucket"].get(label)
        put(ws, f"A{r}", label, BODY, border=BOX)
        for j, k, fmt in ((2, "n", "0"), (3, "median", PCT), (4, "hit", '0"%"'),
                          (5, "p25", PCT), (6, "p75", PCT)):
            put(ws, f"{get_column_letter(j)}{r}", (st or {}).get(k), fmt=fmt, border=BOX)
    r += 2
    r0_sb = r
    put(ws, f"A{r}", "BY SUBSECTOR x SUBSCRIPTION", SECT); r += 1
    # the machine key ("subsector|bucket") the Screener MATCHes on lives in a
    # HIDDEN helper column G — 120 rows of "Biotech 18A|10-100x" read as
    # clutter, and the visible table needs none of it
    for j, h in enumerate(["Subsector", "Subscription", "n", "Median day-1", "Hit-rate"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    first_sb = r + 1
    for (sub, label), st in sorted(an["sub_bucket"].items()):
        r += 1
        put(ws, f"A{r}", sub, BODY, border=BOX)
        put(ws, f"B{r}", label, BODY, border=BOX)
        put(ws, f"C{r}", st["n"], fmt="0", border=BOX)
        put(ws, f"D{r}", st["median"], fmt=PCT, border=BOX)
        put(ws, f"E{r}", st["hit"], fmt='0"%"', border=BOX)
        put(ws, f"G{r}", f"{sub}|{label}", NOTE)
    ws.column_dimensions["G"].hidden = True
    last_sb = r
    r += 2
    put(ws, f"A{r}", "BY SUBSECTOR (any subscription)", SECT); r += 1
    for j, h in enumerate(["Subsector", "n", "Median day-1", "Hit-rate",
                           "Median P/E", "Median P/S", "n P/E", "n P/S"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    first_s = r + 1
    for sub, st in sorted(an["subsector"].items()):
        r += 1
        put(ws, f"A{r}", sub, BODY, border=BOX)
        put(ws, f"B{r}", st["n"], fmt="0", border=BOX)
        put(ws, f"C{r}", st["median"], fmt=PCT, border=BOX)
        put(ws, f"D{r}", st["hit"], fmt='0"%"', border=BOX)
        vm = an["val"].get(sub, {})
        put(ws, f"E{r}", vm.get("pe"), fmt="0.0", border=BOX)
        put(ws, f"F{r}", vm.get("ps"), fmt="0.0", border=BOX)
        put(ws, f"G{r}", vm.get("n_pe"), fmt="0", border=BOX)
        put(ws, f"H{r}", vm.get("n_ps"), fmt="0", border=BOX)
    last_s = r
    r += 2
    put(ws, f"A{r}", "BY YEAR x SUBSCRIPTION — does the edge survive a bad tape?", SECT); r += 1
    for j, h in enumerate(["Year", "Subscription", "n", "Median day-1", "Hit-rate"], 1):
        hdr(ws, f"{get_column_letter(j)}{r}", h)
    for (y, label), st in sorted(an["year_bucket"].items()):
        r += 1
        for j, v, fmt in ((1, y, None), (2, label, None), (3, st["n"], "0"),
                          (4, st["median"], PCT), (5, st["hit"], '0"%"')):
            put(ws, f"{get_column_letter(j)}{r}", v, fmt=fmt, border=BOX)
    for c, w in (("A", 30), ("B", 26), ("C", 14), ("D", 8), ("E", 13), ("F", 10)):
        ws.column_dimensions[c].width = w
    define(wb, "AnalogKey", f"Analogs!$G${first_sb}:$G${last_sb}")
    define(wb, "AnalogN", f"Analogs!$C${first_sb}:$C${last_sb}")
    define(wb, "AnalogMed", f"Analogs!$D${first_sb}:$D${last_sb}")
    define(wb, "AnalogHit", f"Analogs!$E${first_sb}:$E${last_sb}")
    define(wb, "SubsecKey", f"Analogs!$A${first_s}:$A${last_s}")
    define(wb, "SubsecN", f"Analogs!$B${first_s}:$B${last_s}")
    define(wb, "SubsecMed", f"Analogs!$C${first_s}:$C${last_s}")
    define(wb, "SubsecHit", f"Analogs!$D${first_s}:$D${last_s}")
    define(wb, "SubsecPE", f"Analogs!$E${first_s}:$E${last_s}")
    define(wb, "SubsecPS", f"Analogs!$F${first_s}:$F${last_s}")
    return ws




if __name__ == "__main__":
    # Without this the file is importable but does NOTHING when run as a script,
    # so `ipo.py export` quietly shipped whatever .xlsx was already on disk.
    main()
